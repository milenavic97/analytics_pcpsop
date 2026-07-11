"""
ETL — processa cada base e insere no Supabase.
"""

import pandas as pd
import re
from typing import Tuple
from app.database import supabase
import io
import uuid
from datetime import datetime

TUBETES_POR_CAIXA = 500

GRUPOS_ANEST = {
    "0101", "0102", "0103", "0104", "0105", "0106",
    "0107", "0108", "0109", "0110", "0111", "0112",
    "0113", "0114", "0115", "0116",
}
GRUPOS_ANEST_NORM = {
    "101", "102", "103", "104", "105", "106",
    "107", "108", "109", "110", "111", "112",
    "113", "114", "115", "116",
    "0101", "0102", "0103", "0104", "0105", "0106",
    "0107", "0108", "0109", "0110", "0111", "0112",
    "0113", "0114", "0115", "0116",
}

GRUPOS_VALIDOS = {
    "ALPHACAINE", "ALPHACAINE 80", "ARTICAINE", "ARTICAINE 200",
    "MEPIADRE", "MEPISV", "PRILONEST"
}

PRODUTOS_ANEST = {
    '50997','50305','50757','50807','50975','50979','51577','52469','52756','52762','52763','52851',
    '40295','40327','50993','50137','50989','51451','51515','51569','51585','52470','52750','52759',
    '52764','52783','52787','52816','52842','52852','52765','40319','50687','52823','40323','50999',
    '50131','50811','51581','52767','40303','51001','50135','50745','50809','50991','51579','52473',
    '52815','52766','52853','40299','51003','40315',
}

PK_MAP: dict[str, tuple[str, str]] = {
    "d_produtos": ("cod_produto", "VAZIO"),
    "d_lead_time_estoque": ("codigo", "VAZIO"),
    "d_qtd_minima_estoque": ("codigo", "VAZIO"),
    "d_custo_unitario": ("codigo", "VAZIO"),
}

MES_MAP = {
    "Jan": 1, "Fev": 2, "Mar": 3, "Abr": 4, "Mai": 5, "Jun": 6,
    "Jul": 7, "Ago": 8, "Set": 9, "Out": 10, "Nov": 11, "Dez": 12,
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _normaliza_grupo(val: str) -> str:
    v = str(val).strip()
    try:
        return str(int(v))
    except ValueError:
        return v


def _normaliza_armazem(val: str) -> str:
    v = str(val).strip()
    try:
        return str(int(v))
    except ValueError:
        return v


def _ajustar_quantidade_bom_por_tipo(quantidade: float, tipo_pai: str) -> float:
    """
    Regra de estrutura/FIFO:
      - Quando o produto pai for PI, a quantidade necessária da BOM deve ser dividida por 100.
      - Para PA e demais tipos, mantém a quantidade original.

    Isso corrige o gargalo de PI sem alterar saldo de estoque, quarentena ou quantidade da OP.
    """
    tipo = str(tipo_pai or "").strip().upper()

    if tipo == "PI":
        return quantidade / 100

    return quantidade


def _limpar_tabela(table: str):
    if table in PK_MAP:
        col, val = PK_MAP[table]
        supabase.table(table).delete().neq(col, val).execute()
    else:
        supabase.table(table).delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()


def _chunk_insert(table: str, records: list[dict], chunk_size: int = 500) -> list[str]:
    erros = []

    for i in range(0, len(records), chunk_size):
        chunk = records[i : i + chunk_size]

        try:
            supabase.table(table).insert(chunk).execute()

        except Exception as e:
            msg = str(e)[:300]

            if "duplicate" in msg.lower() or "unique" in msg.lower():
                erros.append("Registro duplicado encontrado — verifique se o arquivo já foi carregado.")
            elif "foreign key" in msg.lower():
                erros.append("Produto referenciado não existe na dimensão de produtos.")
            elif "not null" in msg.lower():
                erros.append("Campo obrigatório vazio em um ou mais registros.")
            else:
                erros.append(f"Erro ao inserir: {msg}")

    return erros


def _wide_to_long_agregado(
    df: pd.DataFrame,
    fixed_cols: list[str],
    value_col: str,
    group_cols: list[str],
) -> list[dict]:
    df.columns = [str(c).strip() for c in df.columns]

    month_cols = [c for c in df.columns if c not in fixed_cols]
    records = []

    for _, row in df.iterrows():
        for col in month_cols:
            mes = None
            ano = 2026

            col_str = str(col).strip()

            if col_str in MES_MAP:
                mes = MES_MAP[col_str]
            else:
                try:
                    date = pd.to_datetime(col, errors="coerce")

                    if pd.isna(date):
                        date = pd.to_datetime(float(col), unit="D", origin="1899-12-30")

                    mes = date.month
                    ano = date.year

                except Exception:
                    continue

            try:
                qtd = float(row[col] or 0)
            except Exception:
                qtd = 0

            if qtd == 0:
                continue

            base = {k: str(row.get(k, "")).strip() for k in fixed_cols}
            base["mes"] = mes
            base["ano"] = ano
            base[value_col] = qtd
            records.append(base)

    if not records:
        return []

    agg_df = pd.DataFrame(records)
    agg_df = agg_df.groupby(group_cols, as_index=False)[value_col].sum()

    return agg_df.to_dict("records")




# ─── Parâmetros de Estoque: Lead Time / Qtd Mínima / Custo Unitário ───────────

def _process_parametro_estoque(
    conteudo: bytes,
    filename: str,
    tabela_destino: str,
    campo_valor_db: str,
    nomes_valor: list[str],
    nome_base: str,
) -> Tuple[int, list]:
    """
    Processador genérico para bases mestre simples por código de item.

    Usado por:
      - d_lead_time_estoque
      - d_qtd_minima_estoque
      - d_custo_unitario

    Layout flexível:
      Código | Descrição | Tipo | Unidade | <campo de valor> | Ativo | Observação

    Regras:
      - Substitui a base inteira a cada upload.
      - Mantém código como texto com zero à esquerda quando numérico.
      - Aceita número em formato brasileiro.
      - Ignora linhas sem código.
      - Em código duplicado, mantém a última ocorrência do arquivo.
    """
    from io import BytesIO
    from datetime import datetime, timezone
    import unicodedata

    def normaliza_coluna(valor: str) -> str:
        texto = str(valor or "").strip().lower()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = (
            texto
            .replace("ç", "c")
            .replace("º", "")
            .replace("ª", "")
            .replace("?", "")
            .replace("/", "_")
            .replace("-", "_")
            .replace(".", "")
            .replace("(", "")
            .replace(")", "")
            .replace("$", "")
            .replace("r$", "")
            .replace("+", "_")
            .replace("%", "")
            .replace(" ", "_")
        )
        while "__" in texto:
            texto = texto.replace("__", "_")
        return texto.strip("_")

    def parse_text(valor) -> str | None:
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        texto = str(valor).strip()
        if texto.endswith(".0"):
            texto = texto[:-2]

        if not texto or texto.lower() in {"nan", "none", "nat", "null"}:
            return None

        return texto

    def parse_codigo(valor) -> str | None:
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()

        texto = texto.strip()
        if not texto or texto.lower() in {"nan", "none", "nat", "null"}:
            return None

        return texto.zfill(5) if texto.isdigit() else texto

    def parse_num(valor) -> float:
        if valor is None:
            return 0.0

        try:
            if pd.isna(valor):
                return 0.0
        except Exception:
            pass

        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return 0.0

                v = (
                    v
                    .replace("\xa0", "")
                    .replace("R$", "")
                    .replace("r$", "")
                    .replace(" ", "")
                )

                # Padrão brasileiro: 1.234,56 -> 1234.56
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def parse_bool(valor, default: bool = True) -> bool:
        if valor is None:
            return default

        try:
            if pd.isna(valor):
                return default
        except Exception:
            pass

        texto = str(valor).strip().upper()

        if texto in {"SIM", "S", "YES", "Y", "TRUE", "1", "ATIVO", "ATIVA"}:
            return True

        if texto in {"NAO", "NÃO", "N", "NO", "FALSE", "0", "INATIVO", "INATIVA"}:
            return False

        return default

    def ler_excel_com_header(conteudo_bytes: bytes) -> pd.DataFrame:
        excel = pd.ExcelFile(BytesIO(conteudo_bytes))

        # Preferência por aba com nome da base; se não achar, usa a primeira com cabeçalho válido.
        abas_preferidas = []
        for aba in excel.sheet_names:
            aba_norm = normaliza_coluna(aba)
            if any(p in aba_norm for p in ["lead", "min", "qtd", "custo", "param"]):
                abas_preferidas.append(aba)

        abas = abas_preferidas + [a for a in excel.sheet_names if a not in abas_preferidas]

        for aba in abas:
            for h in range(0, 10):
                try:
                    df_test = pd.read_excel(
                        BytesIO(conteudo_bytes),
                        sheet_name=aba,
                        header=h,
                        nrows=0,
                    )

                    cols_norm = [normaliza_coluna(c) for c in df_test.columns]

                    tem_codigo = any(
                        c in {
                            "codigo", "cod", "cod_produto", "codigo_produto",
                            "codproduto", "produto_codigo", "cod_item", "item"
                        }
                        for c in cols_norm
                    )

                    valor_aliases_norm = {normaliza_coluna(n) for n in nomes_valor}
                    tem_valor = any(c in valor_aliases_norm for c in cols_norm)

                    if tem_codigo and tem_valor:
                        return pd.read_excel(
                            BytesIO(conteudo_bytes),
                            sheet_name=aba,
                            header=h,
                        )

                except Exception:
                    continue

        # Fallback: primeira aba com header 0. A validação abaixo informará colunas faltantes.
        return pd.read_excel(BytesIO(conteudo_bytes), sheet_name=0, header=0)

    try:
        df = ler_excel_com_header(conteudo)
    except Exception as e:
        return 0, [f"Erro ao ler arquivo de {nome_base}: {str(e)[:200]}"]

    if df.empty:
        return 0, [f"Arquivo de {nome_base} vazio."]

    df.columns = [str(c).strip() for c in df.columns]
    col_norm_map = {normaliza_coluna(c): c for c in df.columns}

    def col(*nomes: str):
        for nome in nomes:
            nome_norm = normaliza_coluna(nome)
            if nome_norm in col_norm_map:
                return col_norm_map[nome_norm]
        return None

    c_codigo = col(
        "codigo", "código", "cod", "cod_produto", "codigo_produto",
        "código produto", "cod produto", "cód. produto", "produto_codigo",
        "cod_item", "item"
    )
    c_descricao = col(
        "descricao", "descrição", "produto", "desc_produto",
        "descricao_produto", "descrição produto", "desc item", "item descricao"
    )
    c_tipo = col("tipo", "tp", "tipo_produto", "tipo produto")
    c_unidade = col("unidade", "unid", "un", "um")
    c_valor = col(*nomes_valor)
    c_ativo = col("ativo", "status", "habilitado")
    c_observacao = col("observacao", "observação", "obs", "comentario", "comentário")

    faltando = []
    if c_codigo is None:
        faltando.append("Código")
    if c_valor is None:
        faltando.append(nomes_valor[0])

    if faltando:
        return 0, [
            f"Colunas não encontradas na base {nome_base}: {', '.join(faltando)}. "
            f"Colunas aceitas para o valor: {', '.join(nomes_valor)}."
        ]

    atualizado_em = datetime.now(timezone.utc).isoformat()

    records_por_codigo: dict[str, dict] = {}
    duplicados: list[str] = []

    for _, row in df.iterrows():
        codigo = parse_codigo(row.get(c_codigo))

        if not codigo:
            continue

        valor = parse_num(row.get(c_valor))

        record = {
            "codigo": codigo,
            "descricao": parse_text(row.get(c_descricao)) if c_descricao else None,
            "tipo": parse_text(row.get(c_tipo)) if c_tipo else None,
            "unidade": parse_text(row.get(c_unidade)) if c_unidade else None,
            campo_valor_db: valor,
            "ativo": parse_bool(row.get(c_ativo), default=True) if c_ativo else True,
            "observacao": parse_text(row.get(c_observacao)) if c_observacao else None,
            "atualizado_em": atualizado_em,
        }

        if codigo in records_por_codigo:
            duplicados.append(codigo)

        records_por_codigo[codigo] = record

    records = list(records_por_codigo.values())

    if not records:
        return 0, [f"Nenhum registro válido encontrado na base {nome_base}."]

    erros: list[str] = []

    if duplicados:
        erros.append(
            "Códigos duplicados encontrados; mantida a última ocorrência do arquivo: "
            + ", ".join(sorted(set(duplicados))[:20])
        )

    try:
        _limpar_tabela(tabela_destino)
    except Exception as e:
        return 0, [f"Erro ao limpar {tabela_destino}: {str(e)[:300]}"]

    erros_insert = _chunk_insert(tabela_destino, records)
    erros.extend(erros_insert)

    return len(records) - len(erros_insert), erros


def process_lead_time_estoque(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return _process_parametro_estoque(
        conteudo=conteudo,
        filename=filename,
        tabela_destino="d_lead_time_estoque",
        campo_valor_db="lead_time_dias",
        nomes_valor=[
            "lead_time_dias",
            "lead time dias",
            "lead time",
            "lead_time",
            "lead time total",
            "lead time total dias",
            "lead time total (dias)",
            "lead_time_total",
            "lead_time_total_dias",
            "lt",
            "lt dias",
            "prazo",
            "prazo dias",
        ],
        nome_base="Lead Time de Estoque",
    )


def process_qtd_minima_estoque(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return _process_parametro_estoque(
        conteudo=conteudo,
        filename=filename,
        tabela_destino="d_qtd_minima_estoque",
        campo_valor_db="qtd_minima",
        nomes_valor=[
            "qtd_minima",
            "qtd mínima",
            "qtd minima",
            "quantidade minima",
            "quantidade mínima",
            "pedido minimo",
            "pedido mínimo",
            "qtd min",
            "moq",
            "lote minimo",
            "lote mínimo",
        ],
        nome_base="Quantidade Mínima de Estoque",
    )




def process_parametros_estoque(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Processa uma base única de parâmetros de estoque contendo Lead Time e MOQ.

    Layout esperado/flexível:
      Cod. Produto | Descrição | Tipo | UN | Moq | Lead time total (dias)

    Experiência do usuário:
      - o usuário sobe um único arquivo no modal da Gestão de Estoque;
      - o processador grava o MOQ em d_qtd_minima_estoque;
      - o processador grava o Lead Time em d_lead_time_estoque.

    Observação técnica:
      - mantemos as duas tabelas separadas no banco para simplificar os cálculos
        do Aging e preservar compatibilidade com as funções antigas.
    """
    total_lt, erros_lt = process_lead_time_estoque(conteudo, filename)
    total_moq, erros_moq = process_qtd_minima_estoque(conteudo, filename)

    erros: list[str] = []

    # Evita duplicar mensagens idênticas e identifica de qual parte veio o erro.
    for erro in erros_lt or []:
        erros.append(f"Lead Time: {erro}")

    for erro in erros_moq or []:
        erros.append(f"Quantidade Mínima/MOQ: {erro}")

    # Para o upload único, o total exibido deve ser a quantidade de códigos
    # processados, não a soma das duas tabelas.
    total = max(total_lt or 0, total_moq or 0)

    if total_lt != total_moq:
        erros.append(
            f"Aviso: foram carregados {total_lt} registros de Lead Time e "
            f"{total_moq} registros de Quantidade Mínima/MOQ. Verifique se há "
            "códigos com Lead Time ou MOQ em branco."
        )

    return total, erros


# Alias opcional para nomes de rota/tabela usados no router de upload.
def process_parametros_estoque_file(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return process_parametros_estoque(conteudo, filename)


def process_custo_unitario(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return _process_parametro_estoque(
        conteudo=conteudo,
        filename=filename,
        tabela_destino="d_custo_unitario",
        campo_valor_db="custo_unitario",
        nomes_valor=[
            "custo_unitario",
            "custo unitario",
            "custo unitário",
            "c unitario",
            "c. unitario",
            "c. unitário",
            "custo",
            "valor unitario",
            "valor unitário",
            "preco unitario",
            "preço unitário",
        ],
        nome_base="Custo Unitário",
    )


# Aliases opcionais, caso algum router use o nome físico da tabela.
def process_d_lead_time_estoque(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return process_lead_time_estoque(conteudo, filename)


def process_d_qtd_minima_estoque(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return process_qtd_minima_estoque(conteudo, filename)


def process_d_custo_unitario(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return process_custo_unitario(conteudo, filename)



# ─── d_produtos ──────────────────────────────────────────────────────────────

def process_d_produtos(df: pd.DataFrame) -> Tuple[int, list]:
    """
    Processa a dimensão de produtos.

    Compatível com:
      1) Layout antigo da ferramenta:
         CodProduto | DescProduto | Grupo | Mercado

      2) Nova estrutura gerencial:
         cod_produto | desc_produto | macro_negocio | tipo_negocio |
         tipo_produto_erp | familia | segmento | grupo | mercado | abc_ytm |
         linha | status_original | status_portfolio | transferencia_bravi |
         fornecedor_terceiro | modelo_fornecimento | grupo_gerencial |
         incluir_overview_anestesicos | ativo_analise | observacao |
         concatenado_produto

    Regra de segurança:
      - Mantém cod_produto, desc_produto, grupo e mercado para não quebrar as telas atuais.
      - As novas colunas são apenas camadas gerenciais adicionais.
      - A Overview de anestésicos continua dependendo do grupo original / códigos anestésicos.
    """
    import unicodedata

    df.columns = [str(c).strip() for c in df.columns]

    def normaliza_coluna(valor: str) -> str:
        texto = str(valor or "").strip().lower()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = (
            texto
            .replace("ç", "c")
            .replace("º", "")
            .replace("ª", "")
            .replace("?", "")
            .replace("/", "_")
            .replace("-", "_")
            .replace(".", "")
            .replace(" ", "_")
        )
        while "__" in texto:
            texto = texto.replace("__", "_")
        return texto.strip("_")

    col_norm_map = {normaliza_coluna(c): c for c in df.columns}

    def col(*nomes: str):
        for nome in nomes:
            nome_norm = normaliza_coluna(nome)
            if nome_norm in col_norm_map:
                return col_norm_map[nome_norm]
        return None

    def parse_text(valor) -> str | None:
        if valor is None:
            return None
        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        texto = str(valor).strip()
        if texto.endswith(".0"):
            texto = texto[:-2]

        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return None

        return texto.strip()

    def parse_codigo(valor) -> str | None:
        texto = parse_text(valor)
        if not texto:
            return None

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
        except Exception:
            pass

        texto = texto.strip()
        if texto.endswith(".0"):
            texto = texto[:-2]

        return texto.zfill(5) if texto.isdigit() else texto

    def texto_upper(valor) -> str:
        return str(valor or "").strip().upper()

    def parse_bool(valor, default: bool = False) -> bool:
        texto = texto_upper(valor)
        if texto in {"SIM", "S", "YES", "Y", "TRUE", "1", "ATIVO", "ATIVA"}:
            return True
        if texto in {"NAO", "NÃO", "N", "NO", "FALSE", "0", "INATIVO", "INATIVA"}:
            return False
        return default

    def parse_sim_nao(valor, default: str = "Não") -> str:
        return "Sim" if parse_bool(valor, default=(default == "Sim")) else "Não"

    def contem(texto: str | None, termo: str) -> bool:
        return termo.upper() in texto_upper(texto)

    def normaliza_status(texto: str | None) -> str | None:
        if not texto:
            return None
        t = texto_upper(texto)
        if "DESCONT" in t:
            return "Descontinuado"
        if "BRAVI" in t or "TRANSFER" in t:
            return "Transferido Bravi"
        if "OBSOLE" in t:
            return "Obsoleto"
        if "TROCA" in t and "COD" in t:
            return "Troca de Código"
        if "AMOST" in t:
            return "Amostra"
        if "ATIV" in t:
            return "Ativo"
        return str(texto).strip()

    def normaliza_grupo_str(valor: str | None) -> str:
        return str(valor or "").strip().upper()

    def normaliza_mercado(valor: str | None, is_anest_injetavel: bool = False) -> str:
        """
        Normaliza o campo antigo `mercado` para respeitar a constraint da tabela.

        Hoje `d_produtos.mercado` é uma coluna antiga, obrigatória e usada pelas
        visões já existentes. Por isso ela NÃO deve receber valores gerenciais
        como "A classificar", "Validar" etc.

        Regra segura:
          - Anestésicos injetáveis ficam como PI;
          - Demais itens, quando não houver mercado válido, ficam como NACIONAL.
        """
        texto = parse_text(valor)

        if texto:
            t = unicodedata.normalize("NFKD", texto)
            t = "".join(ch for ch in t if not unicodedata.combining(ch))
            t = t.strip().upper()

            if t in {"PI", "P.I"}:
                return "PI"

            if "NACION" in t or t in {"NAC", "BR", "BRASIL"}:
                return "NACIONAL"

            # Se aparecer algo ligado a exportação, mantemos em NACIONAL por
            # segurança até a constraint/estrutura de mercado ser revisada.
            # O detalhe comercial deve ficar nas colunas gerenciais novas.
            if "EXPORT" in t or t in {"EXP", "EXTERNO", "INTERNACIONAL"}:
                return "NACIONAL"

            if t not in {"", "A CLASSIFICAR", "VALIDAR", "NAO CLASSIFICADO", "NÃO CLASSIFICADO", "NONE", "NAN"}:
                # Qualquer valor fora da lista aceita pelo banco volta para o
                # padrão seguro, em vez de quebrar o upload inteiro.
                return "PI" if is_anest_injetavel else "NACIONAL"

        return "PI" if is_anest_injetavel else "NACIONAL"

    # Colunas — aceita nomes antigos, novos e variações comuns.
    c_codigo = col(
        "cod_produto", "codigo", "código", "codproduto", "cod produto",
        "cód. produto", "cod. produto", "cód produto", "produto_codigo"
    )
    c_desc = col(
        "desc_produto", "descricao", "descrição", "produto", "descproduto",
        "descricao_produto", "descrição produto", "desc produto"
    )
    c_tipo_produto = col("tipo_produto_erp", "tipo produto erp", "tipo_produto", "tipo produto", "tipo", "tp")
    c_familia = col("familia", "família")
    c_segmento = col("segmento")
    c_grupo = col("grupo", "grupo produto", "grupo de produto")
    c_mercado = col("mercado")
    c_abc = col("abc_ytm", "abc ytm", "abc", "curva", "curva abc")
    c_linha = col("linha", "linha_original", "linha original")
    c_status_original = col("status_original", "status original", "status")
    c_macro = col("macro_negocio", "macro negócio", "macro negocio", "macro")
    c_tipo_negocio = col("tipo_negocio", "tipo negócio", "tipo negocio", "negocio", "negócio")
    c_status_portfolio = col("status_portfolio", "status portfólio", "status portfolio", "status gerencial")
    c_bravi = col(
        "transferencia_bravi", "transferência bravi", "transferencia bravi",
        "transferencia bravi?", "transferência bravi?", "transferido para bravi",
        "transferido_bravi", "bravi"
    )
    c_fornecedor = col("fornecedor_terceiro", "fornecedor terceiro", "terceiro", "fornecedor")
    c_modelo = col("modelo_fornecimento", "modelo fornecimento", "modelo de fornecimento")
    c_grupo_gerencial = col("grupo_gerencial", "grupo gerencial")
    c_overview = col("incluir_overview_anestesicos", "incluir overview anestesicos", "incluir overview anestésicos", "overview anestesicos")
    c_ativo = col("ativo_analise", "ativo analise", "ativo análise", "ativo")
    c_obs = col("observacao", "observação", "obs")
    c_concat = col("concatenado_produto", "concatenado produto", "produto concatenado")

    if c_codigo is None:
        return 0, ["Coluna de código do produto não encontrada. Use cod_produto/CÓDIGO/CodProduto."]

    records_por_codigo: dict[str, dict] = {}
    duplicados: set[str] = set()

    for _, row in df.iterrows():
        cod_produto = parse_codigo(row.get(c_codigo))
        if not cod_produto:
            continue

        desc_produto = parse_text(row.get(c_desc)) if c_desc else None
        grupo = parse_text(row.get(c_grupo)) if c_grupo else None
        mercado = parse_text(row.get(c_mercado)) if c_mercado else None
        familia = parse_text(row.get(c_familia)) if c_familia else None
        segmento = parse_text(row.get(c_segmento)) if c_segmento else None
        linha = parse_text(row.get(c_linha)) if c_linha else None
        tipo_produto_erp = parse_text(row.get(c_tipo_produto)) if c_tipo_produto else None
        abc_ytm = parse_text(row.get(c_abc)) if c_abc else None
        status_original = parse_text(row.get(c_status_original)) if c_status_original else None

        macro_negocio = parse_text(row.get(c_macro)) if c_macro else None
        tipo_negocio = parse_text(row.get(c_tipo_negocio)) if c_tipo_negocio else None
        status_portfolio = parse_text(row.get(c_status_portfolio)) if c_status_portfolio else None
        transferencia_bravi = parse_sim_nao(row.get(c_bravi), default="Não") if c_bravi else "Não"
        fornecedor_terceiro = parse_text(row.get(c_fornecedor)) if c_fornecedor else None
        modelo_fornecimento = parse_text(row.get(c_modelo)) if c_modelo else None
        grupo_gerencial = parse_text(row.get(c_grupo_gerencial)) if c_grupo_gerencial else None
        observacao = parse_text(row.get(c_obs)) if c_obs else None
        concatenado_produto = parse_text(row.get(c_concat)) if c_concat else None

        grupo_norm = normaliza_grupo_str(grupo)
        desc_norm = texto_upper(desc_produto)
        familia_norm = texto_upper(familia)
        segmento_norm = texto_upper(segmento)
        status_base = status_portfolio or status_original

        # Inferências leves apenas para preencher campos gerenciais quando o arquivo não trouxer.
        # Não altera o campo grupo, que segue sendo a classificação original usada pela Overview.
        is_anest_injetavel = (
            grupo_norm in {g.upper() for g in GRUPOS_VALIDOS}
            or cod_produto in PRODUTOS_ANEST
        )
        is_benzotop = "BENZOTOP" in desc_norm or "BENZOTOP" in familia_norm or "BENZOTOP" in segmento_norm
        is_pps = "PPS" in familia_norm or "PPS" in segmento_norm or "PPS" in texto_upper(linha)

        # `mercado` é coluna antiga com NOT NULL + CHECK no Supabase.
        # Não pode receber vazio/A classificar.
        mercado = normaliza_mercado(mercado, is_anest_injetavel=is_anest_injetavel)

        if not macro_negocio:
            if is_anest_injetavel or is_benzotop:
                macro_negocio = "Anestésicos"
            elif is_pps:
                macro_negocio = "PPS"

        if not tipo_negocio:
            if is_benzotop:
                tipo_negocio = "Benzotop"
            elif is_anest_injetavel:
                tipo_negocio = "Anestésicos Injetáveis"
            elif (macro_negocio or "").upper() == "PPS" or is_pps:
                tipo_negocio = "PPS"

        if not status_portfolio:
            status_portfolio = normaliza_status(status_base)

        if transferencia_bravi == "Sim" and not status_portfolio:
            status_portfolio = "Transferido Bravi"

        if not status_portfolio:
            status_portfolio = "Ativo"

        if not fornecedor_terceiro and transferencia_bravi == "Sim":
            fornecedor_terceiro = "Bravi"

        if not modelo_fornecimento:
            tipo_erp_norm = texto_upper(tipo_produto_erp)
            status_norm = texto_upper(status_portfolio)
            if transferencia_bravi == "Sim":
                modelo_fornecimento = "Terceirizado"
            elif "DESCONT" in status_norm:
                modelo_fornecimento = "Sem produção futura"
            elif tipo_erp_norm == "MR":
                modelo_fornecimento = "Revenda"
            elif tipo_negocio in {"Anestésicos Injetáveis", "Benzotop"}:
                modelo_fornecimento = "Produção interna"

        if not grupo_gerencial:
            status_norm = texto_upper(status_portfolio)
            if tipo_negocio == "Anestésicos Injetáveis":
                grupo_gerencial = "Anestésicos Injetáveis"
            elif tipo_negocio == "Benzotop":
                grupo_gerencial = "Benzotop"
            elif tipo_negocio == "PPS" and transferencia_bravi == "Sim":
                grupo_gerencial = "PPS - Transferência Bravi"
            elif tipo_negocio == "PPS" and "DESCONT" in status_norm:
                grupo_gerencial = "PPS - Descontinuado"
            elif tipo_negocio == "PPS":
                grupo_gerencial = "PPS - Ativo terceirizado/revenda"
            elif "OBSOLE" in status_norm:
                grupo_gerencial = "Obsoleto"
            elif "TROCA" in status_norm:
                grupo_gerencial = "Troca de Código"
            elif "AMOST" in status_norm:
                grupo_gerencial = "Amostra"
            else:
                grupo_gerencial = "A classificar"

        if c_overview:
            incluir_overview_anestesicos = parse_bool(row.get(c_overview), default=is_anest_injetavel)
        else:
            incluir_overview_anestesicos = bool(is_anest_injetavel)

        if c_ativo:
            ativo_analise = parse_bool(row.get(c_ativo), default=True)
        else:
            # Mantém true por padrão para permitir acompanhar também descontinuados/Bravi em estoque.
            ativo_analise = True

        if not concatenado_produto:
            concatenado_produto = f"{cod_produto} - {desc_produto}" if desc_produto else cod_produto

        record = {
            # Colunas antigas — NÃO remover / NÃO trocar significado.
            "cod_produto": cod_produto,
            "desc_produto": desc_produto or "",
            "grupo": grupo or "",
            "mercado": mercado,

            # Novas camadas gerenciais.
            "tipo_produto_erp": tipo_produto_erp,
            "familia": familia,
            "segmento": segmento,
            "abc_ytm": abc_ytm,
            "linha": linha,
            "status_original": status_original,
            "macro_negocio": macro_negocio,
            "tipo_negocio": tipo_negocio,
            "status_portfolio": status_portfolio,
            "transferencia_bravi": transferencia_bravi,
            "fornecedor_terceiro": fornecedor_terceiro,
            "modelo_fornecimento": modelo_fornecimento,
            "grupo_gerencial": grupo_gerencial,
            "incluir_overview_anestesicos": incluir_overview_anestesicos,
            "ativo_analise": ativo_analise,
            "observacao": observacao,
            "concatenado_produto": concatenado_produto,
        }

        if cod_produto in records_por_codigo:
            duplicados.add(cod_produto)

        # Em caso de duplicidade, mantém a última ocorrência do arquivo.
        records_por_codigo[cod_produto] = record

    records = list(records_por_codigo.values())

    if not records:
        return 0, ["Nenhum produto válido encontrado na base d_produtos."]

    def _chunk_upsert_d_produtos(records_upsert: list[dict], chunk_size: int = 500) -> list[str]:
        """
        Atualiza/insere a dimensão sem apagar a tabela.

        Motivo:
          d_produtos é referenciada por bases fato como f_forecast_sop,
          f_orcado_faturamento etc. Se apagarmos a dimensão antes de inserir,
          o Supabase bloqueia por foreign key.

        Regra:
          - código existente: atualiza as colunas;
          - código novo: insere;
          - código antigo que não veio no arquivo: permanece no banco.

        Segurança:
          - tenta primeiro em lote;
          - se o lote falhar, tenta linha a linha para identificar exatamente
            quais produtos precisam de correção sem perder os demais.
        """
        erros_upsert: list[str] = []

        for i in range(0, len(records_upsert), chunk_size):
            chunk = records_upsert[i : i + chunk_size]

            try:
                (
                    supabase
                    .table("d_produtos")
                    .upsert(chunk, on_conflict="cod_produto")
                    .execute()
                )
                continue

            except Exception:
                # Se o lote falhar, tenta item a item para isolar o problema.
                pass

            for rec in chunk:
                try:
                    (
                        supabase
                        .table("d_produtos")
                        .upsert(rec, on_conflict="cod_produto")
                        .execute()
                    )
                except Exception as e:
                    msg = str(e)[:500]
                    erros_upsert.append(
                        f"Erro ao atualizar produto {rec.get('cod_produto')} - {rec.get('desc_produto')}: {msg}"
                    )

        return erros_upsert

    # IMPORTANTE:
    # Não limpar d_produtos. Esta tabela é dimensão mestre e possui foreign keys
    # em tabelas fato. O upload deve ser incremental por upsert.
    erros_insert = _chunk_upsert_d_produtos(records)

    return len(records) - len(erros_insert), erros_insert


# ─── f_orcado_liberacao ───────────────────────────────────────────────────────

HERANCA_2025 = {"L1": 3_000_000.0, "L2": 1_298_500.0}


def process_orcado_liberacao(df: pd.DataFrame) -> Tuple[int, list]:
    df.columns = [str(c).strip() for c in df.columns]
    df = df[~df.iloc[:, 0].astype(str).str.contains("Liberacao", case=False, na=False)]
    df = df.dropna(subset=[df.columns[0]])

    records = []

    for _, row in df.iterrows():
        mes_label = str(row.iloc[0]).strip()
        mes_num = MES_MAP.get(mes_label)

        if not mes_num:
            continue

        for linha_col, linha_key in [("L1", "L1"), ("L2", "L2")]:
            qtd = float(row.get(linha_col, 0) or 0)

            records.append({
                "mes": mes_num,
                "ano": 2026,
                "linha": linha_key,
                "qtd_tubetes": qtd,
                "heranca_2025": False,
            })

    for linha_key, qtd in HERANCA_2025.items():
        records.append({
            "mes": 1,
            "ano": 2026,
            "linha": linha_key,
            "qtd_tubetes": qtd,
            "heranca_2025": True,
        })

    _limpar_tabela("f_orcado_liberacao")
    erros = _chunk_insert("f_orcado_liberacao", records)

    return len(records) - len(erros), erros


# ─── f_orcado_faturamento ─────────────────────────────────────────────────────

def process_orcado_faturamento(df: pd.DataFrame) -> Tuple[int, list]:
    df.columns = [str(c).strip() for c in df.columns]

    fixed = ["Cód. Produto", "Descricao Produto", "Grupo de Produto", "Família"]
    group_cols = ["Cód. Produto", "Descricao Produto", "Grupo de Produto", "Família", "mes", "ano"]

    records = _wide_to_long_agregado(df, fixed, "qtd_caixas", group_cols)

    if not records:
        return 0, ["Nenhum dado encontrado. Verifique o formato do arquivo."]

    clean = []

    for r in records:
        clean.append({
            "cod_produto":  str(r.get("Cód. Produto", "")).strip(),
            "desc_produto": str(r.get("Descricao Produto", "")).strip(),
            "grupo":        str(r.get("Grupo de Produto", "")).strip(),
            "familia":      str(r.get("Família", "")).strip(),
            "mes":          int(r["mes"]),
            "ano":          int(r["ano"]),
            "qtd_caixas":   float(r["qtd_caixas"]),
        })

    _limpar_tabela("f_orcado_faturamento")
    erros = _chunk_insert("f_orcado_faturamento", clean)

    return len(clean) - len(erros), erros


# ─── f_forecast_sop ──────────────────────────────────────────────────────────

def process_forecast_sop(df: pd.DataFrame) -> Tuple[int, list]:
    df.columns = [str(c).strip() for c in df.columns]

    fixed = ["Cód. Produto", "Descricao Produto", "Grupo de Produto", "Família"]
    group_cols = ["Cód. Produto", "Descricao Produto", "Grupo de Produto", "Família", "mes", "ano"]

    records = _wide_to_long_agregado(df, fixed, "qtd_forecast", group_cols)

    if not records:
        return 0, ["Nenhum dado encontrado. Verifique o formato do arquivo."]

    clean = []

    for r in records:
        clean.append({
            "cod_produto":  str(r.get("Cód. Produto", "")).strip(),
            "desc_produto": str(r.get("Descricao Produto", "")).strip(),
            "grupo":        str(r.get("Grupo de Produto", "")).strip(),
            "familia":      str(r.get("Família", "")).strip(),
            "mes":          int(r["mes"]),
            "ano":          int(r["ano"]),
            "qtd_forecast": float(r["qtd_forecast"]),
        })

    _limpar_tabela("f_forecast_sop")
    erros = _chunk_insert("f_forecast_sop", clean)

    return len(clean) - len(erros), erros


# ─── f_sd2_saidas ────────────────────────────────────────────────────────────

def process_sd2_saidas(df: pd.DataFrame) -> Tuple[int, list]:
    """
    Processa a SD2 de saídas/vendas em visão corporativa.

    Nova regra:
      - NÃO filtra mais apenas grupos anestésicos 0101-0116.
      - A SD2 passa a alimentar vendas reais de todas as famílias/linhas.
      - Filtros de linha, família, grupo, segmento e mercado devem ser feitos
        nas telas usando a dimensão d_produtos.

    Filtros técnicos mantidos:
      - considera apenas armazéns 03, 04, 07, 27 e 88;
      - considera apenas Tipo Produto PA ou MR, quando a coluna existir;
      - exclui descrições com AVULSO;
      - exclui estornos;
      - considera apenas TES/Tipo Saída classificado como "Venda" no De/Para Tipo Saída;
      - salva em f_sd2_saidas já filtrado para venda real válida.
    """

    df.columns = [str(c).strip() for c in df.columns]

    def normaliza_coluna(valor: str) -> str:
        return (
            str(valor or "")
            .strip()
            .upper()
            .replace(" ", "")
            .replace(".", "")
            .replace("-", "")
            .replace("_", "")
            .replace("/", "")
            .replace("Ç", "C")
            .replace("Ã", "A")
            .replace("Á", "A")
            .replace("Â", "A")
            .replace("À", "A")
            .replace("É", "E")
            .replace("Ê", "E")
            .replace("Í", "I")
            .replace("Ó", "O")
            .replace("Ô", "O")
            .replace("Õ", "O")
            .replace("Ú", "U")
        )

    col_norm_map = {normaliza_coluna(c): c for c in df.columns}

    def col(*nomes):
        for nome in nomes:
            nome_norm = normaliza_coluna(nome)
            if nome_norm in col_norm_map:
                return col_norm_map[nome_norm]
        return None

    def parse_num(valor) -> float:
        if valor is None or pd.isna(valor):
            return 0.0

        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return 0.0

                # Padrão BR: 1.234,56 -> 1234.56
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def parse_codigo(valor, zfill: int | None = None) -> str:
        if valor is None or pd.isna(valor):
            return ""

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()

        texto = texto.strip()

        if not texto or texto.lower() in {"nan", "none"}:
            return ""

        return texto.zfill(zfill) if zfill else texto

    def normaliza_tipo_saida(valor) -> str:
        return parse_codigo(valor).strip().upper()

    def normaliza_texto(valor) -> str:
        return str(valor or "").strip()

    def buscar_tipos_saida_venda() -> set[str]:
        """
        Busca no Supabase a tabela de De/Para Tipo Saída.

        Como o nome físico pode variar conforme a criação da tela de Bases,
        tentamos alguns nomes comuns. A tabela esperada tem colunas semelhantes a:
        COD_TIPO, TIPO_TES, TXT_PADRAO, DESCRICAO.
        """
        tabelas_possiveis = [
            "d_depara_tipo_saida",
            "depara_tipo_saida",
            "d_tipo_saida",
            "f_depara_tipo_saida",
            "depara_tipos_saida",
        ]

        codigos_venda: set[str] = set()

        for tabela in tabelas_possiveis:
            try:
                res = supabase.table(tabela).select("*").execute()
                rows = res.data or []
            except Exception:
                continue

            if not rows:
                continue

            for row in rows:
                row_norm = {normaliza_coluna(k): v for k, v in row.items()}

                descricao = str(
                    row_norm.get("DESCRICAO")
                    or row_norm.get("DESCRICAOTIPO")
                    or row_norm.get("TXTCLASSIFICACAO")
                    or ""
                ).strip().upper()

                if descricao != "VENDA":
                    continue

                cod = (
                    row_norm.get("CODTIPO")
                    or row_norm.get("CODIGOTIPO")
                    or row_norm.get("TIPO")
                    or row_norm.get("TIPOTES")
                    or row_norm.get("TES")
                )

                cod_norm = normaliza_tipo_saida(cod)
                if cod_norm:
                    codigos_venda.add(cod_norm)

            if codigos_venda:
                return codigos_venda

        return codigos_venda

    c_armazem = col("Armazem", "Armazém")
    c_grupo = col("Grupo", "Grupo Produto", "Grupo de Produto")
    c_descricao = col("Descricao", "Descrição", "Descr. Prod", "Desc Produto", "Descricao Produto")
    c_quantidade = col("Quantidade", "Qtd", "Qtd Faturada")
    c_emissao = col("Emissao", "Emissão", "DT Emissao", "DT Emissão", "Data Emissao", "Data Emissão")
    c_produto = col("Produto", "Cod Produto", "Cód. Produto", "Codigo Produto", "Código Produto", "CodProduto")
    c_tipo_produto = col("Tipo Produto", "TipoProduto", "TP Produto", "TP")
    c_tipo_saida = col(
        "Tipo Saida",
        "Tipo Saída",
        "Tipo S",
        "Tipo S.",
        "Tipo",
        "TES",
        "Cod. do Tipo",
        "Cod do Tipo",
        "Cod Tipo",
        "Código Tipo",
        "Codigo Tipo",
    )
    c_estornado = col("Estornado", "Estorno")
    c_vlr_total = col("Vlr.Total", "Vlr Total", "Valor Total", "Total")
    c_cliente = col("Cliente", "Cod Cliente", "Codigo Cliente", "Código Cliente")

    faltando = [
        nome for nome, c in [
            ("Armazem", c_armazem),
            ("Grupo", c_grupo),
            ("Descricao", c_descricao),
            ("Quantidade", c_quantidade),
            ("Emissao", c_emissao),
            ("Produto", c_produto),
        ]
        if c is None
    ]

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o formato da SD2."]

    df["_arm"] = df[c_armazem].astype(str).str.strip().apply(_normaliza_armazem)
    df["_desc"] = df[c_descricao].astype(str).str.upper()

    # Regra corporativa:
    # mantém apenas filtros técnicos de venda válida.
    # O filtro fixo por grupos anestésicos foi removido.
    mask = (
        df["_arm"].isin(["3", "4", "7", "27", "88"])
        & ~df["_desc"].str.contains("AVULSO", na=False)
    )

    if c_tipo_produto is not None:
        mask &= df[c_tipo_produto].astype(str).str.strip().str.upper().isin(["PA", "MR"])

    if c_estornado is not None:
        est = df[c_estornado].astype(str).str.strip().str.upper()
        mask &= (
            df[c_estornado].isna()
            | est.isin(["", "N", "NAO", "NÃO", "NO", "0", "FALSE", "NAN", "NONE"])
        )

    # Operação/TES: apenas Tipo Saída classificado como Venda no De/Para.
    # Se a coluna de Tipo Saída existir mas o De/Para não estiver disponível, bloqueia o upload
    # para evitar inflar a venda com transferência/devolução/outras operações.
    if c_tipo_saida is not None:
        tipos_venda = buscar_tipos_saida_venda()

        if not tipos_venda:
            return 0, [
                "Não foi possível localizar no Supabase o De/Para Tipo Saída com DESCRICAO = 'Venda'. "
                "Verifique se a base De/Para Tipo Saída foi carregada antes da SD2."
            ]

        mask &= df[c_tipo_saida].apply(normaliza_tipo_saida).isin(tipos_venda)
    else:
        return 0, [
            "Coluna de Tipo Saída/TES não encontrada na SD2. "
            "Ela é necessária para filtrar apenas operações classificadas como Venda."
        ]

    df = df[mask].copy()

    if df.empty:
        return 0, [
            "Nenhum registro encontrado após aplicar os filtros técnicos da SD2: "
            "armazéns 03/04/07/27/88, Tipo Produto PA/MR quando existir, "
            "sem AVULSO, sem estorno e Tipo Saída classificado como Venda."
        ]

    records = []

    for _, row in df.iterrows():
        try:
            emissao = pd.to_datetime(row.get(c_emissao), errors="coerce")

            if pd.isna(emissao):
                continue

            quantidade = parse_num(row.get(c_quantidade))
            if quantidade == 0:
                continue

            records.append({
                "produto":    parse_codigo(row.get(c_produto)),
                "descricao":  normaliza_texto(row.get(c_descricao)),
                "quantidade": quantidade,
                "vlr_total":  parse_num(row.get(c_vlr_total)) if c_vlr_total is not None else 0.0,
                "armazem":    str(row.get(c_armazem, "")).strip(),
                "grupo":      str(row.get(c_grupo, "")).strip(),
                "cliente":    parse_codigo(row.get(c_cliente)) if c_cliente is not None else "",
                # IMPORTANTE:
                # f_sd2_saidas.mes e f_sd2_saidas.ano são colunas geradas no Supabase.
                # Por isso NÃO podem ser inseridas aqui; o banco calcula automaticamente
                # a partir da coluna emissao.
                "emissao":    emissao.date().isoformat(),
            })

        except Exception:
            continue

    if not records:
        return 0, ["Nenhum registro válido encontrado após converter data e quantidade."]

    _limpar_tabela("f_sd2_saidas")
    erros = _chunk_insert("f_sd2_saidas", records)

    return len(records) - len(erros), erros

# ─── f_sd3_entradas ──────────────────────────────────────────────────────────

def process_sd3_entradas(df: pd.DataFrame) -> Tuple[int, list]:
    df.columns = [str(c).strip() for c in df.columns]

    if "TP Movimento" in df.columns:
        df["_tp"] = (
            df["TP Movimento"]
            .astype(str)
            .str.strip()
            .str.replace(r"\.0$", "", regex=True)
        )
        mask_tp = df["_tp"] == "499"
    else:
        mask_tp = pd.Series([True] * len(df), index=df.index)

    colunas_necessarias = ["Armazem", "Grupo", "Descr. Prod"]
    faltando = [c for c in colunas_necessarias if c not in df.columns]

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o formato."]

    df["_arm"] = df["Armazem"].astype(str).str.strip().apply(_normaliza_armazem)
    df["_grp"] = df["Grupo"].astype(str).str.strip().apply(_normaliza_grupo)
    df["_descr_prod"] = df["Descr. Prod"].astype(str).str.upper()

    mask = (
        mask_tp
        & df["_arm"].isin(["4", "7"])
        & df["_grp"].isin(GRUPOS_ANEST_NORM)
        & ~df["_descr_prod"].str.contains("AVULSO", na=False)
        & ~df["_descr_prod"].str.contains(r"\bAG\b", regex=True, na=False)
    )

    if "Tipo Produto" in df.columns:
        mask &= df["Tipo Produto"].astype(str).str.strip().str.upper().eq("PA")

    if "Estornado" in df.columns:
        mask &= ~df["Estornado"].astype(str).str.strip().str.upper().eq("SIM")

    df = df[mask].copy()

    if df.empty:
        return 0, [
            "Nenhum registro encontrado após aplicar os filtros: "
            "TP Movimento 499, armazéns 04/07, grupos anestésicos, "
            "sem AVULSO, sem AG/amostra grátis, Tipo Produto PA quando existir, e sem estorno."
        ]

    records = []

    for _, row in df.iterrows():
        try:
            dt = pd.to_datetime(row.get("DT Emissao") or row.get("DT Emissão"), errors="coerce")

            if pd.isna(dt):
                continue

            quantidade = float(row.get("Quantidade", 0) or 0)

            if quantidade == 0:
                continue

            produto = str(row.get("Produto", "")).strip().replace(".0", "")
            armazem_norm = _normaliza_armazem(row.get("Armazem", ""))
            armazem_db = "04" if armazem_norm == "4" else "07"

            records.append({
                "produto":    produto,
                "descr_prod": str(row.get("Descr. Prod", "")).strip(),
                "lote":       str(row.get("Lote", "")).strip(),
                "quantidade": quantidade,
                "armazem":    armazem_db,
                "grupo":      str(row.get("Grupo", "")).strip(),
                "dt_emissao": dt.date().isoformat(),
                "custo":      float(row.get("Custo", 0) or 0),
            })

        except Exception:
            continue

    if not records:
        return 0, ["Nenhum registro válido encontrado após converter data e quantidade."]

    _limpar_tabela("f_sd3_entradas")
    erros = _chunk_insert("f_sd3_entradas", records)

    return len(records) - len(erros), erros


# ─── f_estoque ────────────────────────────────────────────────────────────────

def process_estoque(df: pd.DataFrame) -> Tuple[int, list]:
    df.columns = [str(c).strip() for c in df.columns]

    colunas_necessarias = ["Produto", "Armazem", "Data Saldo", "Qtd.Inic.Mes"]
    faltando = [c for c in colunas_necessarias if c not in df.columns]

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o formato."]

    df = df.dropna(subset=["Produto", "Armazem", "Data Saldo", "Qtd.Inic.Mes"]).copy()

    df["Produto"] = (
        df["Produto"]
        .astype(str)
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)
    )

    df["_arm"] = df["Armazem"].astype(str).str.strip().apply(_normaliza_armazem)

    mask = (
        df["_arm"].isin(["4", "7"])
        & df["Produto"].isin(PRODUTOS_ANEST)
    )

    df = df[mask].copy()

    if df.empty:
        return 0, ["Nenhum produto anestésico encontrado nos armazéns 04 e 07."]

    def parse_data_saldo(value):
        if pd.isna(value):
            return pd.NaT

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return pd.to_datetime(float(value), unit="D", origin="1899-12-30", errors="coerce")

        return pd.to_datetime(value, errors="coerce")

    records = []

    for _, row in df.iterrows():
        try:
            data_saldo = parse_data_saldo(row["Data Saldo"])

            if pd.isna(data_saldo):
                continue

            qtd_caixas = float(row["Qtd.Inic.Mes"] or 0)

            if qtd_caixas == 0:
                continue

            armazem_db = "04" if row["_arm"] == "4" else "07"

            records.append({
                "mes":        int(data_saldo.month),
                "ano":        int(data_saldo.year),
                "produto":    str(row["Produto"]).strip(),
                "armazem":    armazem_db,
                "qtd_caixas": qtd_caixas,
            })

        except Exception:
            continue

    if not records:
        return 0, ["Nenhum registro válido encontrado no arquivo de estoque."]

    agg_df = pd.DataFrame(records)
    agg_df = agg_df.groupby(["mes", "ano", "produto", "armazem"], as_index=False)["qtd_caixas"].sum()

    records = agg_df.to_dict("records")

    _limpar_tabela("f_estoque")
    erros = _chunk_insert("f_estoque", records)

    return len(records) - len(erros), erros


# ─── f_producao_real ─────────────────────────────────────────────────────────

def process_producao_real(df: pd.DataFrame) -> Tuple[int, list]:
    df.columns = [str(c).strip() for c in df.columns]

    if "EQUIPAMENTO" not in df.columns:
        return 0, ["Coluna EQUIPAMENTO não encontrada. Verifique se está usando o arquivo correto."]

    df = df.dropna(subset=["EQUIPAMENTO"])

    records = []

    for _, row in df.iterrows():
        try:
            data_ini = pd.to_datetime(row.get("DATA INICIAL"), errors="coerce")
            data_fim = pd.to_datetime(row.get("DATA FINAL"), errors="coerce")

            records.append({
                "equipamento":   str(row.get("EQUIPAMENTO",        "")).strip(),
                "tipo_evento":   str(row.get("TIPO DE EVENTO",     "")).strip(),
                "evento":        str(row.get("EVENTO",             "")).strip(),
                "produto":       str(row.get("PRODUTO",            "")).strip(),
                "lote":          str(row.get("LOTE",               "")).strip(),
                "data_inicial":  data_ini.isoformat() if not pd.isna(data_ini) else None,
                "data_final":    data_fim.isoformat() if not pd.isna(data_fim) else None,
                "duracao_h":     float(row.get("DURACAO",              0) or 0),
                "qtd_produzida": float(row.get("QUANTIDADE PRODUZIDA", 0) or 0),
                "qtd_rejeitada": float(row.get("QUANTIDADE REJEITADA", 0) or 0),
                "mes":           data_ini.month if not pd.isna(data_ini) else None,
                "ano":           data_ini.year if not pd.isna(data_ini) else None,
            })

        except Exception:
            continue

    if not records:
        return 0, ["Nenhum registro válido encontrado no arquivo de produção."]

    _limpar_tabela("f_producao_real")
    erros = _chunk_insert("f_producao_real", records)

    return len(records) - len(erros), erros


# ─── f_entradas_previstas ─────────────────────────────────────────────────────

def process_entradas_previstas(df: pd.DataFrame) -> Tuple[int, list]:
    records = []
    linha_atual = None
    meses_cols: list = []

    for _, row in df.iterrows():
        vals = list(row.values)
        primeiro = str(vals[0]).strip() if pd.notna(vals[0]) else ""

        if "LINHA 1" in primeiro.upper() and "LINHA 2" not in primeiro.upper():
            linha_atual = "L1"
            meses_cols = []

            for v in vals[1:]:
                try:
                    meses_cols.append(int(float(v)))
                except Exception:
                    meses_cols.append(None)

            continue

        if "LINHA 2" in primeiro.upper():
            linha_atual = "L2"
            meses_cols = []

            for v in vals[1:]:
                try:
                    meses_cols.append(int(float(v)))
                except Exception:
                    meses_cols.append(None)

            continue

        if primeiro not in GRUPOS_VALIDOS or linha_atual is None:
            continue

        grupo = primeiro

        for col_idx, mes_num in enumerate(meses_cols):
            if mes_num is None:
                continue

            data_idx = col_idx + 1

            if data_idx >= len(vals):
                continue

            try:
                qtd = float(vals[data_idx] or 0)
            except Exception:
                qtd = 0

            if qtd == 0:
                continue

            records.append({
                "linha":      linha_atual,
                "grupo":      grupo,
                "mes":        mes_num,
                "ano":        2026,
                "qtd_caixas": qtd,
            })

    if not records:
        return 0, ["Nenhum dado encontrado. Verifique o formato do arquivo."]

    _limpar_tabela("f_entradas_previstas")
    erros = _chunk_insert("f_entradas_previstas", records)

    return len(records) - len(erros), erros


# ─── f_mps_producao ──────────────────────────────────────────────────────────

def process_mps_producao_file(conteudo: bytes, filename: str) -> Tuple[int, list]:
    from io import BytesIO
    from datetime import datetime, date
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel

    wb = load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    def normalizar_header(h):
        if h is None:
            return ""
        return str(h).strip().lower()

    def normalizar_linha(valor):
        v = str(valor).strip().upper().replace(".0", "")

        if v in ["1", "L1", "LINHA 1", "LINHA1"]:
            return "L1"

        if v in ["2", "L2", "LINHA 2", "LINHA2"]:
            return "L2"

        return None

    header_norm = [normalizar_header(h) for h in headers]

    try:
        idx_mes = header_norm.index("mês")
    except ValueError:
        try:
            idx_mes = header_norm.index("mes")
        except ValueError:
            return 0, ["Coluna 'Mês' não encontrada."]

    try:
        idx_versao = header_norm.index("versão")
    except ValueError:
        try:
            idx_versao = header_norm.index("versao")
        except ValueError:
            return 0, ["Coluna 'Versão' não encontrada."]

    try:
        idx_linha = header_norm.index("linha")
    except ValueError:
        return 0, ["Coluna 'Linha' não encontrada."]

    def parse_data_header(value):
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return None

        try:
            parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")

            if pd.isna(parsed):
                return None

            return parsed.date()

        except Exception:
            return None

    date_cols = []

    for idx, h in enumerate(headers):
        data_col = parse_data_header(h)

        if data_col:
            date_cols.append((idx, data_col))

    if not date_cols:
        return 0, ["Nenhuma coluna de data encontrada no arquivo."]

    records = []

    for row in ws.iter_rows(min_row=2):
        mes_val = row[idx_mes].value
        versao_val = row[idx_versao].value
        linha_val = row[idx_linha].value

        if mes_val is None or versao_val is None or linha_val is None:
            continue

        try:
            mes_linha = int(float(mes_val))
        except Exception:
            continue

        linha = normalizar_linha(linha_val)

        if linha is None:
            continue

        try:
            versao = str(int(float(versao_val)))
        except Exception:
            versao = str(versao_val).strip()

        for col_idx, data_dia in date_cols:
            if data_dia.month != mes_linha:
                continue

            cell = row[col_idx]
            raw_value = cell.value

            try:
                horas_producao = float(raw_value or 0)
            except Exception:
                horas_producao = 0

            horas_parada = max(0, 24 - horas_producao)

            comentario = None

            if cell.comment:
                comentario = cell.comment.text.strip()

            records.append({
                "versao": versao,
                "data_dia": data_dia.isoformat(),
                "linha": linha,
                "horas_producao": horas_producao,
                "horas_parada": horas_parada,
                "comentario": comentario,
            })

    if not records:
        return 0, ["Nenhum registro válido encontrado no calendário de produção."]

    _limpar_tabela("f_mps_producao")
    erros = _chunk_insert("f_mps_producao", records)

    return len(records) - len(erros), erros


# ─── f_mps_liberacoes ────────────────────────────────────────────────────────

def process_mps_liberacoes_file(conteudo: bytes, filename: str) -> Tuple[int, list]:
    from io import BytesIO

    df = pd.read_excel(BytesIO(conteudo), sheet_name=0, header=None)

    records = []
    linha_atual = None
    versao_atual = None
    mes_revisao_atual = None
    meses_cols: list[tuple[int, int]] = []

    def normalizar_linha(valor):
        texto = str(valor or "").strip().upper()

        if "LINHA 1" in texto or texto in {"L1", "1"}:
            return "L1"

        if "LINHA 2" in texto or texto in {"L2", "2"}:
            return "L2"

        return None

    def normalizar_versao(valor):
        texto = str(valor or "").strip().upper().replace("V", "")

        try:
            return str(int(float(texto)))
        except Exception:
            return ""

    def normalizar_mes(valor):
        if valor is None or pd.isna(valor):
            return None

        texto = str(valor).strip()

        if texto in MES_MAP:
            return MES_MAP[texto]

        try:
            mes = int(float(texto))
            if 1 <= mes <= 12:
                return mes
        except Exception:
            return None

        return None

    for _, row in df.iterrows():
        vals = list(row.values)

        if len(vals) < 4:
            continue

        col_mes_revisao = vals[0] if len(vals) > 0 else None
        col_versao = vals[1] if len(vals) > 1 else None
        col_desc = vals[2] if len(vals) > 2 else None

        desc = str(col_desc or "").strip()
        desc_upper = desc.upper()

        linha_detectada = normalizar_linha(desc)

        if linha_detectada:
            linha_atual = linha_detectada
            versao_atual = normalizar_versao(col_versao)
            mes_revisao_atual = normalizar_mes(col_mes_revisao)
            meses_cols = []

            for idx in range(3, len(vals)):
                mes_planejado = normalizar_mes(vals[idx])

                if mes_planejado is not None:
                    meses_cols.append((idx, mes_planejado))

            continue

        if desc_upper != "TOTAL":
            continue

        if linha_atual not in {"L1", "L2"}:
            continue

        if not versao_atual:
            versao_atual = normalizar_versao(col_versao)

        if not versao_atual:
            continue

        if mes_revisao_atual is None:
            mes_revisao_atual = normalizar_mes(col_mes_revisao)

        if mes_revisao_atual is None:
            continue

        if not meses_cols:
            continue

        for idx, mes_planejado in meses_cols:
            if idx >= len(vals):
                continue

            try:
                qtd = float(vals[idx] or 0)
            except Exception:
                qtd = 0

            records.append({
                "versao": versao_atual,
                "mes_revisao": int(mes_revisao_atual),
                "mes": int(mes_planejado),
                "ano": 2026,
                "linha": linha_atual,
                "qtd_caixas": qtd,
            })

    if not records:
        return 0, ["Nenhum dado encontrado. Verifique se o arquivo possui blocos de LINHA 1/LINHA 2 e linhas TOTAL."]

    agg_df = pd.DataFrame(records)
    agg_df = agg_df.groupby(["versao", "mes_revisao", "mes", "ano", "linha"], as_index=False)["qtd_caixas"].sum()

    clean = []
    for r in agg_df.to_dict("records"):
        clean.append({
            "versao": str(r["versao"]),
            "mes_revisao": int(r["mes_revisao"]),
            "mes": int(r["mes"]),
            "ano": int(r["ano"]),
            "linha": str(r["linha"]).strip().upper(),
            "qtd_caixas": float(r["qtd_caixas"]),
        })

    _limpar_tabela("f_mps_liberacoes")
    erros = _chunk_insert("f_mps_liberacoes", clean)

    return len(clean) - len(erros), erros


# ─── d_bom_estrutura ─────────────────────────────────────────────────────────

def process_bom_estrutura(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Processa a Estrutura de Materiais/BOM e grava na tabela d_bom_estrutura.

    Aceita dois layouts:
      1) Layout Protheus original:
         Codigo | Descricao | Tipo | CODIGO | DESCRICAO | TP | QUANTIDADE | UM

      2) Layout atualizado/manual:
         Código | Descrição | Tipo | Componente | Descrição2 | UM | Quantidade

    Observações:
      - A base é mestre: substitui toda a d_bom_estrutura a cada upload.
      - Quando o layout novo não traz TP do componente, o TP é buscado na BOM atual
        por código do componente; se não existir, é inferido pela descrição/unidade.
      - CARTUCHO/TUBETE/LACRE/EMBOLO etc. são tratados como ME para entrarem
        como materiais gargalantes na página de Ordens.
      - Mantém a regra já existente de PI: quantidade da BOM do pai PI é dividida por 100.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    def normaliza_coluna(valor: str) -> str:
        return (
            str(valor or "")
            .strip()
            .upper()
            .replace(" ", "")
            .replace(".", "")
            .replace("-", "")
            .replace("_", "")
            .replace("/", "")
            .replace("Ç", "C")
            .replace("Ã", "A")
            .replace("Á", "A")
            .replace("Â", "A")
            .replace("À", "A")
            .replace("É", "E")
            .replace("Ê", "E")
            .replace("Í", "I")
            .replace("Ó", "O")
            .replace("Ô", "O")
            .replace("Õ", "O")
            .replace("Ú", "U")
        )

    def normaliza_texto(valor: str) -> str:
        return normaliza_coluna(valor)

    def parse_codigo(valor, zfill: int | None = 5) -> str | None:
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()

        texto = texto.strip()

        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return None

        return texto.zfill(zfill) if zfill and texto.isdigit() else texto

    def parse_text(valor) -> str:
        if valor is None:
            return ""

        try:
            if pd.isna(valor):
                return ""
        except Exception:
            pass

        texto = str(valor).strip()
        if texto.endswith(".0"):
            texto = texto[:-2]

        if texto.lower() in {"nan", "none", "nat"}:
            return ""

        return texto.strip()

    def parse_num(valor) -> float:
        if valor is None:
            return 0.0

        try:
            if pd.isna(valor):
                return 0.0
        except Exception:
            pass

        try:
            if isinstance(valor, str):
                v = valor.strip().replace("\xa0", "")
                if not v:
                    return 0.0

                # Padrão BR: 1.244,4 -> 1244.4
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def buscar_tp_existente() -> dict[str, str]:
        """Usa a BOM atualmente carregada como apoio para preencher TP no layout manual."""
        mapa: dict[str, str] = {}

        try:
            res = supabase.table("d_bom_estrutura").select("codigo_comp,tp").execute()
        except Exception:
            return mapa

        for row in res.data or []:
            codigo = parse_codigo(row.get("codigo_comp"), zfill=5)
            tp = parse_text(row.get("tp")).upper()

            if codigo and tp and codigo not in mapa:
                mapa[codigo] = tp

        return mapa

    def inferir_tp(codigo_comp: str, desc_comp: str, unidade: str, tp_existente: dict[str, str]) -> str:
        codigo = parse_codigo(codigo_comp, zfill=5) or str(codigo_comp or "").strip()

        if codigo in tp_existente:
            return tp_existente[codigo]

        desc_norm = normaliza_texto(desc_comp)
        un_norm = normaliza_texto(unidade)

        # Materiais de embalagem que precisam entrar como gargalo na página de Ordens.
        palavras_embalagem = [
            "CARTUCHO",
            "TUBETE",
            "TUBO",
            "LACRE",
            "EMBOLO",
            "EMBOLO",
            "ROTULO",
            "ROTULAGEM",
            "ETIQUETA",
            "BULA",
            "CAIXA",
            "DISPLAY",
            "BERCO",
            "BLISTER",
            "SACO",
            "SACHE",
            "TAMPA",
            "FRASCO",
        ]

        if any(p in desc_norm for p in palavras_embalagem):
            return "ME"

        # Unidades típicas de matéria-prima.
        if un_norm in {"KG", "G", "GR", "MG", "L", "ML"}:
            return "MP"

        palavras_mp = [
            "CLORIDRATO",
            "EPINEFRINA",
            "ACIDO",
            "SODIO",
            "POTASSIO",
            "ALCOOL",
            "AGUA",
            "SOLUCAO",
            "BASE",
            "INSUMO",
        ]

        if any(p in desc_norm for p in palavras_mp):
            return "MP"

        # Fallback de segurança: melhor tratar como material gargalante do que ignorar.
        return "ME"

    def achar_indices(header: list) -> dict[str, int | None]:
        header_txt = [str(h or "").strip() for h in header]
        header_norm = [normaliza_coluna(h) for h in header_txt]

        def idx_exact(*nomes: str) -> int | None:
            nomes_set = {str(n).strip() for n in nomes}
            for i, h in enumerate(header_txt):
                if h in nomes_set:
                    return i
            return None

        def idx_norm(*nomes: str, after: int | None = None) -> int | None:
            nomes_norm = {normaliza_coluna(n) for n in nomes}
            for i, h in enumerate(header_norm):
                if after is not None and i <= after:
                    continue
                if h in nomes_norm:
                    return i
            return None

        def idx_norm_last(*nomes: str) -> int | None:
            nomes_norm = {normaliza_coluna(n) for n in nomes}
            achados = [i for i, h in enumerate(header_norm) if h in nomes_norm]
            return achados[-1] if achados else None

        # Código do pai: primeiro código da linha.
        i_codigo_pai = (
            idx_exact("Codigo", "Código", "Cod Produto", "Cód. Produto", "Codigo Pai", "Código Pai")
            or idx_norm("Codigo", "Código", "Cod Produto", "Codigo Pai", "CodPai", "Produto")
        )

        # Componente: no layout novo vem como Componente; no Protheus vem como CODIGO.
        i_codigo_comp = idx_exact("Componente", "COMPONENTE", "Codigo Componente", "Código Componente")

        if i_codigo_comp is None:
            i_codigo_comp = idx_exact("CODIGO")

        if i_codigo_comp is None and i_codigo_pai is not None:
            i_codigo_comp = idx_norm("Codigo", "Código", "CODIGO", after=i_codigo_pai)

        # Se ainda houver ambiguidade por nomes duplicados, pega o último código da linha como componente.
        if i_codigo_comp is None:
            i_codigo_comp = idx_norm_last("Codigo", "Código", "CODIGO")

        i_desc_pai = (
            idx_exact("Descricao", "Descrição", "Desc Produto", "Descrição Produto", "Descricao Pai", "Descrição Pai")
            or idx_norm("Descricao", "Descrição", "Desc Produto", "Descricao Pai")
        )

        i_desc_comp = idx_exact("Descrição2", "Descricao2", "DESCRICAO", "Descrição Componente", "Descricao Componente")

        if i_desc_comp is None and i_desc_pai is not None:
            i_desc_comp = idx_norm("Descricao", "Descrição", "DESCRICAO", after=i_desc_pai)

        if i_desc_comp is None:
            i_desc_comp = idx_norm_last("Descricao", "Descrição", "DESCRICAO")

        i_tipo_pai = idx_exact("Tipo", "TIPO", "Tipo Pai") or idx_norm("Tipo", "Tipo Pai")
        i_tp = idx_exact("TP", "Tipo Componente", "TP Componente") or idx_norm("TP", "Tipo Componente", "TP Componente")
        i_quantidade = idx_exact("QUANTIDADE", "Quantidade", "Qtd", "Qtd.") or idx_norm("Quantidade", "QUANTIDADE", "Qtd", "Qtd.")
        i_um = idx_exact("UM", "Unidade", "UN") or idx_norm("UM", "Unidade", "UN")

        return {
            "codigo_pai": i_codigo_pai,
            "desc_pai": i_desc_pai,
            "tipo_pai": i_tipo_pai,
            "codigo_comp": i_codigo_comp,
            "desc_comp": i_desc_comp,
            "tp": i_tp,
            "quantidade": i_quantidade,
            "um": i_um,
        }

    def achar_header(rows: list[tuple]) -> tuple[int | None, dict[str, int | None]]:
        for idx_row, row in enumerate(rows[:20]):
            indices = achar_indices(list(row))

            if (
                indices.get("codigo_pai") is not None
                and indices.get("codigo_comp") is not None
                and indices.get("quantidade") is not None
            ):
                return idx_row, indices

        return None, {}

    try:
        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as e:
        return 0, [f"Erro ao ler arquivo de estrutura de materiais: {str(e)[:200]}"]

    # Preferência por abas com nome relacionado a estrutura/BOM; senão usa a primeira.
    ws = wb.active
    for nome in wb.sheetnames:
        nome_norm = normaliza_coluna(nome)
        if any(p in nome_norm for p in ["ESTRUTURA", "BOM", "MATERIAL"]):
            ws = wb[nome]
            break

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, ["Arquivo vazio ou sem dados na planilha."]

    header_idx, indices = achar_header(rows)

    if header_idx is None:
        return 0, [
            "Cabeçalho da BOM não encontrado. Formatos aceitos: "
            "Protheus (Codigo, Descricao, Tipo, CODIGO, DESCRICAO, TP, QUANTIDADE, UM) "
            "ou estrutura manual (Código, Descrição, Tipo, Componente, Descrição2, UM, Quantidade)."
        ]

    data = rows[header_idx + 1:]

    i_codigo_pai = indices.get("codigo_pai")
    i_desc_pai = indices.get("desc_pai")
    i_tipo_pai = indices.get("tipo_pai")
    i_codigo_comp = indices.get("codigo_comp")
    i_desc_comp = indices.get("desc_comp")
    i_tp = indices.get("tp")
    i_quantidade = indices.get("quantidade")
    i_um = indices.get("um")

    faltando = []
    if i_codigo_pai is None:
        faltando.append("Codigo/Código do produto pai")
    if i_codigo_comp is None:
        faltando.append("CODIGO/Componente")
    if i_quantidade is None:
        faltando.append("QUANTIDADE/Quantidade")

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o formato do arquivo."]

    tp_existente = buscar_tp_existente()

    records = []
    codigo_pai_atual = None
    desc_pai_atual = ""
    tipo_pai_atual = ""

    for row in data:
        if not row:
            continue

        codigo_pai_val = parse_codigo(row[i_codigo_pai] if i_codigo_pai is not None and i_codigo_pai < len(row) else None, zfill=5)

        if codigo_pai_val:
            codigo_pai_atual = codigo_pai_val
            desc_pai_atual = parse_text(row[i_desc_pai]) if i_desc_pai is not None and i_desc_pai < len(row) else desc_pai_atual
            tipo_pai_atual = parse_text(row[i_tipo_pai]).upper() if i_tipo_pai is not None and i_tipo_pai < len(row) else tipo_pai_atual

        codigo_comp = parse_codigo(row[i_codigo_comp] if i_codigo_comp is not None and i_codigo_comp < len(row) else None, zfill=5)

        if not codigo_comp or not codigo_pai_atual:
            continue

        quantidade_original = parse_num(row[i_quantidade] if i_quantidade is not None and i_quantidade < len(row) else None)

        if quantidade_original <= 0:
            continue

        unidade = parse_text(row[i_um]) if i_um is not None and i_um < len(row) else ""
        desc_comp = parse_text(row[i_desc_comp]) if i_desc_comp is not None and i_desc_comp < len(row) else ""

        if i_tp is not None and i_tp < len(row):
            tp = parse_text(row[i_tp]).upper()
        else:
            tp = ""

        if not tp:
            tp = inferir_tp(codigo_comp, desc_comp, unidade, tp_existente)

        quantidade = _ajustar_quantidade_bom_por_tipo(quantidade_original, tipo_pai_atual)

        records.append({
            "codigo_pai": codigo_pai_atual,
            "descricao_pai": desc_pai_atual,
            "tipo_pai": tipo_pai_atual,
            "codigo_comp": codigo_comp,
            "descricao_comp": desc_comp,
            "tp": tp,
            "quantidade": quantidade,
            "unidade": unidade,
        })

    if not records:
        return 0, ["Nenhum componente encontrado no arquivo de estrutura de materiais."]

    try:
        _limpar_tabela("d_bom_estrutura")
    except Exception as e:
        return 0, [f"Erro ao limpar d_bom_estrutura: {str(e)[:300]}"]

    erros = _chunk_insert("d_bom_estrutura", records)

    return len(records) - len(erros), erros




# ─── d_lotes_teoricos ────────────────────────────────────────────────────────

def process_lotes_teoricos(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê a base mestre de lotes teóricos por produto + linha.

    Objetivo:
      - Cadastrar a quantidade teórica exigida pelo Protheus para abertura da OP.
      - A programação pode vir com quantidade ajustada por rendimento.
      - A viabilidade de material deve usar a quantidade teórica quando existir.

    Tabela destino:
      d_lotes_teoricos

    Colunas esperadas no Excel, com nomes flexíveis:
      - codigo_produto / código / codigo / cod_produto
      - descricao_produto / descrição / produto
      - letra_lote / letra
      - linha
      - qtd_teorica_abertura / qtd_teorica / quantidade_teorica / lote_teorico
      - ativo
      - observacao / observação

    Regra:
      - Substitui tudo a cada upload, pois é uma base mestre.
      - Chave lógica usada depois na OP: codigo_produto + linha.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    def normaliza_coluna(valor: str) -> str:
        return (
            str(valor or "")
            .strip()
            .lower()
            .replace(" ", "_")
            .replace(".", "")
            .replace("-", "_")
            .replace("/", "_")
            .replace("__", "_")
            .replace("ç", "c")
            .replace("ã", "a")
            .replace("á", "a")
            .replace("à", "a")
            .replace("â", "a")
            .replace("é", "e")
            .replace("ê", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ô", "o")
            .replace("õ", "o")
            .replace("ú", "u")
        )

    def parse_codigo(valor) -> str | None:
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()

        texto = texto.strip()

        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return None

        # Produtos DFL costumam ter 5 dígitos. Se vier numérico menor, completa à esquerda.
        return texto.zfill(5) if texto.isdigit() else texto

    def parse_text(valor) -> str | None:
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        texto = str(valor).strip()

        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return None

        if texto.endswith(".0"):
            texto = texto[:-2]

        return texto.strip()

    def parse_num(valor) -> float:
        if valor is None:
            return 0.0

        try:
            if pd.isna(valor):
                return 0.0
        except Exception:
            pass

        try:
            if isinstance(valor, str):
                v = valor.strip()

                if not v:
                    return 0.0

                # Remove espaços e separadores invisíveis.
                v = v.replace("\xa0", "").replace(" ", "")

                # Padrão brasileiro: 1.234,56 -> 1234.56
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")
                    return float(v)

                # Para a base de lote teórico, valores como 305.000 / 350.000
                # representam tubetes inteiros, não casas decimais.
                # Então, se houver ponto e a parte final tiver 3 dígitos,
                # interpreta como separador de milhar.
                if "." in v:
                    partes = v.split(".")
                    if len(partes) > 1 and all(len(p) == 3 for p in partes[1:]):
                        v = "".join(partes)
                        return float(v)

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def parse_bool_ativo(valor) -> bool:
        if valor is None:
            return True

        try:
            if pd.isna(valor):
                return True
        except Exception:
            pass

        texto = str(valor).strip().upper()

        if texto in {"", "SIM", "S", "YES", "Y", "TRUE", "1", "ATIVO", "ATIVA"}:
            return True

        if texto in {"NAO", "NÃO", "N", "NO", "FALSE", "0", "INATIVO", "INATIVA"}:
            return False

        return True

    def normalizar_linha(valor) -> str | None:
        texto = str(valor or "").strip().upper()

        if not texto:
            return None

        texto_sem_espaco = texto.replace(" ", "").replace("_", "").replace("-", "")

        if texto_sem_espaco in {"L1", "1", "LINHA1", "ENVASEL1", "ENVASELINHA1"}:
            return "L1"

        if texto_sem_espaco in {"L2", "2", "LINHA2", "ENVASEL2", "ENVASELINHA2"}:
            return "L2"

        if "L1" in texto or "LINHA 1" in texto or "ENVASE 1" in texto:
            return "L1"

        if "L2" in texto or "LINHA 2" in texto or "ENVASE 2" in texto:
            return "L2"

        return texto

    try:
        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as e:
        return 0, [f"Erro ao ler arquivo de lotes teóricos: {str(e)[:200]}"]

    # Preferência pela aba lotes_teoricos; se não existir, usa a primeira aba.
    sheet_name = None
    for nome in wb.sheetnames:
        if normaliza_coluna(nome) in {"lotes_teoricos", "lote_teorico", "d_lotes_teoricos"}:
            sheet_name = nome
            break

    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return 0, ["Arquivo de lotes teóricos vazio."]

    # Acha o cabeçalho nas primeiras linhas.
    header_idx = None
    for i, row in enumerate(rows[:10]):
        cols_norm = [normaliza_coluna(v) for v in row]

        tem_codigo = any(c in {
            "codigo_produto", "cod_produto", "codigo", "cod", "produto_codigo", "codigoproduto"
        } for c in cols_norm)

        tem_qtd = any(c in {
            "qtd_teorica_abertura", "qtd_teorica", "quantidade_teorica",
            "lote_teorico", "qtd_lote_teorico", "quantidade_lote_teorico",
            "qtd_teorica_protheus", "qtd_abertura_protheus"
        } for c in cols_norm)

        if tem_codigo and tem_qtd:
            header_idx = i
            break

    if header_idx is None:
        return 0, [
            "Cabeçalho não encontrado. A base precisa ter pelo menos as colunas "
            "codigo_produto e qtd_teorica_abertura."
        ]

    header = [str(h or "").strip() for h in rows[header_idx]]
    header_norm = [normaliza_coluna(h) for h in header]
    data_rows = rows[header_idx + 1:]

    def idx(*nomes):
        for nome in nomes:
            nome_norm = normaliza_coluna(nome)
            if nome_norm in header_norm:
                return header_norm.index(nome_norm)
        return None

    i_codigo = idx("codigo_produto", "cod_produto", "codigo", "cod", "produto_codigo", "Codigo Produto", "Código Produto")
    i_desc = idx("descricao_produto", "desc_produto", "descricao", "descrição", "produto")
    i_letra = idx("letra_lote", "letra", "letra_produto")
    i_linha = idx("linha", "linha_envase", "linha_producao", "linha produção")
    i_qtd = idx(
        "qtd_teorica_abertura",
        "qtd_teorica",
        "quantidade_teorica",
        "lote_teorico",
        "qtd_lote_teorico",
        "quantidade_lote_teorico",
        "qtd_teorica_protheus",
        "qtd_abertura_protheus",
    )
    i_ativo = idx("ativo", "status")
    i_obs = idx("observacao", "observação", "obs")

    faltando = []
    if i_codigo is None:
        faltando.append("codigo_produto")
    if i_linha is None:
        faltando.append("linha")
    if i_qtd is None:
        faltando.append("qtd_teorica_abertura")

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o template de lotes teóricos."]

    records = []
    chaves_vistas: set[tuple[str, str]] = set()
    duplicados: list[str] = []

    for row in data_rows:
        if not row:
            continue

        codigo = parse_codigo(row[i_codigo] if i_codigo < len(row) else None)
        linha = normalizar_linha(row[i_linha] if i_linha < len(row) else None)
        qtd_teorica = parse_num(row[i_qtd] if i_qtd < len(row) else None)

        # Permite deixar linhas do template em branco.
        if not codigo and not linha and qtd_teorica <= 0:
            continue

        if not codigo:
            continue

        if not linha:
            continue

        if qtd_teorica <= 0:
            continue

        chave = (codigo, linha)

        if chave in chaves_vistas:
            duplicados.append(f"{codigo} / {linha}")
            continue

        chaves_vistas.add(chave)

        descricao = parse_text(row[i_desc] if i_desc is not None and i_desc < len(row) else None)
        letra = parse_text(row[i_letra] if i_letra is not None and i_letra < len(row) else None)
        observacao = parse_text(row[i_obs] if i_obs is not None and i_obs < len(row) else None)
        ativo = parse_bool_ativo(row[i_ativo] if i_ativo is not None and i_ativo < len(row) else None)

        records.append({
            "codigo_produto": codigo,
            "descricao_produto": descricao,
            "letra_lote": letra.upper() if letra else None,
            "linha": linha,
            "qtd_teorica_abertura": qtd_teorica,
            "ativo": ativo,
            "observacao": observacao,
        })

    if not records:
        return 0, [
            "Nenhum lote teórico válido encontrado. Preencha código, linha e qtd_teorica_abertura."
        ]

    erros = []

    if duplicados:
        erros.append(
            "Linhas duplicadas ignoradas para a mesma combinação código + linha: "
            + ", ".join(duplicados[:20])
        )

    # Base mestre: substitui tudo a cada upload.
    try:
        _limpar_tabela("d_lotes_teoricos")
    except Exception as e:
        return 0, [f"Erro ao limpar d_lotes_teoricos: {str(e)[:300]}"]

    erros_insert = _chunk_insert("d_lotes_teoricos", records)
    erros.extend(erros_insert)

    return len(records) - len(erros_insert), erros


# Alias opcional para facilitar mapeamento no router de upload, caso ele use
# o nome físico da tabela como padrão.
def process_d_lotes_teoricos(conteudo: bytes, filename: str) -> Tuple[int, list]:
    return process_lotes_teoricos(conteudo, filename)



# ─── f_estoque_saldo ─────────────────────────────────────────────────────────

def process_estoque_saldo(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê a base de saldo por lote e salva no Supabase o saldo DISPONÍVEL.

    NOVA REGRA:
      - Mantém TODOS os armazéns da SB8.
      - Ao subir novamente no mesmo dia, substitui apenas o snapshot daquele data_ref.
      - Isso evita duplicar a SB8 do dia e mantém histórico dos dias anteriores.

    saldo_lote salvo = Saldo Lote - Emp. do Lote
    """
    from io import BytesIO
    from datetime import date as date_cls
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=3, values_only=True))
    if not rows:
        return 0, ["Arquivo vazio."]

    header = [str(h or "").strip() for h in rows[0]]
    data = rows[1:]

    snapshot_id = str(uuid.uuid4())
    upload_id = str(uuid.uuid4())

    def normaliza_coluna(valor: str) -> str:
        return (
            str(valor or "")
            .strip()
            .upper()
            .replace(" ", "")
            .replace(".", "")
            .replace("Ç", "C")
            .replace("Ã", "A")
            .replace("Á", "A")
            .replace("Â", "A")
            .replace("É", "E")
            .replace("Ê", "E")
            .replace("Í", "I")
            .replace("Ó", "O")
            .replace("Ô", "O")
            .replace("Ú", "U")
        )

    header_norm = [normaliza_coluna(h) for h in header]

    def idx(*nomes):
        for nome in nomes:
            nome_norm = normaliza_coluna(nome)

            if nome_norm in header_norm:
                return header_norm.index(nome_norm)

        return None

    def parse_num(valor) -> float:
        if valor is None:
            return 0.0

        try:
            if isinstance(valor, str):
                v = valor.strip()

                if not v:
                    return 0.0

                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    i_produto = idx("Produto")
    i_descricao = idx("Descricao", "Descrição")
    i_armazem = idx("Armazem", "Armazém")
    i_lote = idx("Lote")
    i_saldo = idx("Saldo Lote")

    i_empenho = idx(
        "Emp. do Lote",
        "Emp do Lote",
        "Emp.Lote",
        "Emp Lote",
        "Emp. Qd",
        "Emp Qd",
        "Empenho",
        "Qtd Empenhada",
        "Quantidade Empenhada"
    )

    i_validade = idx("Data Validad", "Data Validade")

    faltando = [n for n, i in [
        ("Produto", i_produto),
        ("Armazem", i_armazem),
        ("Saldo Lote", i_saldo),
    ] if i is None]

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o formato."]

    data_ref = date_cls.today().isoformat()

    records = []

    for row in data:
        try:
            armazem_raw = str(row[i_armazem] or "").strip()

            try:
                armazem_norm = str(int(float(armazem_raw))).zfill(2)
            except Exception:
                armazem_norm = armazem_raw.zfill(2)

            # Mantém TODOS os armazéns da SB8.
            # Cada página/endpoint filtra depois conforme a necessidade:
            # - OP/MRP: insumos/armazéns de matéria-prima
            # - Overview: PA anestésicos/armazéns de PA
            codigo = str(row[i_produto] or "").strip()

            if not codigo:
                continue

            codigo = codigo.replace(".0", "").zfill(5)

            saldo_bruto = parse_num(row[i_saldo])

            empenhado = (
                parse_num(row[i_empenho])
                if i_empenho is not None
                else 0.0
            )

            saldo_disponivel = max(saldo_bruto - empenhado, 0.0)

            # Antes, um lote 100% empenhado (saldo_disponivel = 0) era
            # descartado inteiro -- inclusive o saldo BRUTO dele, que é
            # estoque físico real só que reservado. Isso fazia o total
            # bruto da SB8 vir menor do que o arquivo de origem (ex.:
            # Benzotop 52749: arquivo tinha 85.071 no armazém 04, só
            # 84.751 chegavam no banco -- a diferença eram lotes
            # totalmente empenhados sendo descartados aqui). Agora só
            # descarta lote com bruto zerado/negativo (vazio de verdade).
            if saldo_bruto <= 0:
                continue

            data_validade = None

            if i_validade is not None and row[i_validade]:
                try:
                    val_raw = row[i_validade]

                    if isinstance(val_raw, date_cls):
                        data_validade = val_raw.isoformat()
                    else:
                        dt = pd.to_datetime(val_raw, errors="coerce")

                        if not pd.isna(dt):
                            data_validade = dt.date().isoformat()

                except Exception:
                    pass

            lote = (
                str(row[i_lote] or "").strip()
                if i_lote is not None
                else ""
            )

            descricao = (
                str(row[i_descricao] or "").strip()
                if i_descricao is not None
                else ""
            )

            records.append({
                "snapshot_id": snapshot_id,
                "upload_id": upload_id,

                "data_ref": data_ref,

                "codigo": codigo,
                "descricao": descricao,

                "armazem": armazem_norm,
                "lote": lote,

                "saldo_lote": saldo_disponivel,
                "saldo_bruto": saldo_bruto,
                "empenho_lote": empenhado,

                "data_validade": data_validade,
            })

        except Exception:
            continue

    if not records:
        return 0, ["Nenhum registro válido encontrado na SB8."]

    # Substitui apenas o snapshot do mesmo dia.
    # Assim, se você subir a SB8 de hoje novamente, ela NÃO duplica: apaga hoje e reinsere.
    # Os snapshots de dias anteriores continuam preservados para histórico.
    try:
        supabase.table("f_estoque_saldo").delete().eq("data_ref", data_ref).execute()
    except Exception as e:
        return 0, [f"Erro ao limpar snapshot de estoque do dia {data_ref}: {str(e)[:300]}"]

    erros = _chunk_insert("f_estoque_saldo", records)

    return len(records) - len(erros), erros




# ─── f_programacao_ops ───────────────────────────────────────────────────────

def process_programacao_ops(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê a planilha de programação mensal de anestésicos.

    Abas esperadas:
      - ENVASE LINHA 1
      - ENVASE LINHA 2
      - EMBALAGEM

    Regra principal da página Produção:
      - O planejado do dashboard NÃO deve ser a soma das OPs.
      - O planejado oficial deve vir do KPI do topo da programação:
          célula L4 = META MÊS - TUBETES
      - Esse valor é gravado em f_programacao_ops_resumo por mês/linha.
      - A produção.py usa meta_mes_tubetes / 500 para montar o gráfico.

    Regras da página Ordens:
      - Continua gravando f_programacao_ops linha a linha.
      - A coluna G continua sendo a quantidade de cada OP.
    """
    from io import BytesIO
    from datetime import date as date_cls, datetime as datetime_cls
    from openpyxl import load_workbook

    # Otimização segura:
    # a Programação mensal tem muita formatação. read_only=True evita carregar
    # estilos/células fora do necessário e acelera a abertura do Excel.
    # keep_links=False evita tentar resolver vínculos externos.
    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True, keep_links=False)

    ABAS = {
        "ENVASE LINHA 1": "ENVASE_L1",
        "ENVASE LINHA 2": "ENVASE_L2",
        "EMBALAGEM":      "EMBALAGEM",
    }

    def parse_data(val):
        if val is None:
            return None

        try:
            if isinstance(val, datetime_cls):
                return val.date().isoformat()

            if isinstance(val, date_cls):
                return val.isoformat()

            dt = pd.to_datetime(val, dayfirst=True, errors="coerce")
            if not pd.isna(dt):
                return dt.date().isoformat()

        except Exception:
            pass

        return None

    def parse_float(val) -> float:
        if val is None:
            return 0.0

        try:
            if pd.isna(val):
                return 0.0
        except Exception:
            pass

        try:
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                return float(val)

            texto = str(val).strip()
            if not texto or texto.lower() in {"nan", "none", "nat", "null", "-"}:
                return 0.0

            texto = (
                texto
                .replace("\xa0", "")
                .replace("R$", "")
                .replace("r$", "")
                .replace(" ", "")
            )

            # Formato brasileiro:
            # 7.943.663      -> 7943663
            # 7.943.663,00   -> 7943663.00
            # 123,45         -> 123.45
            if "," in texto:
                texto = texto.replace(".", "").replace(",", ".")
            else:
                # Se houver ponto e a última parte tiver 3 dígitos, trata como separador de milhar.
                partes = texto.split(".")
                if len(partes) > 1 and all(len(p) == 3 for p in partes[1:]):
                    texto = "".join(partes)

            return float(texto)

        except Exception:
            return 0.0

    def parse_codigo(val) -> str:
        if val is None:
            return ""

        try:
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                texto = str(int(float(val)))
            else:
                texto = str(val).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(val).strip()

        texto = texto.strip()
        if not texto or texto.lower() in {"nan", "none", "nat", "null"}:
            return ""

        return texto.zfill(5) if texto.isdigit() else texto

    def parse_op(val):
        if val is None:
            return None

        try:
            if pd.isna(val):
                return None
        except Exception:
            pass

        try:
            return str(int(float(str(val)))).strip()
        except Exception:
            texto = str(val).strip()
            if texto.endswith(".0"):
                texto = texto[:-2]
            return texto or None

    def normalizar_linha_resumo(linha_tag: str) -> str | None:
        if linha_tag == "ENVASE_L1":
            return "L1"
        if linha_tag == "ENVASE_L2":
            return "L2"
        return None

    def detectar_mes_ref() -> str | None:
        """
        Detecta o mês pela DATA FIM da própria programação.
        Nas abas de envase/embalagem, DATA FIM fica na coluna L, índice 11.
        """
        for nome_aba in ABAS.keys():
            if nome_aba not in wb.sheetnames:
                continue

            ws_temp = wb[nome_aba]

            for row in ws_temp.iter_rows(min_row=7, values_only=True):
                if not row:
                    continue

                data_ref = row[11] if len(row) > 11 else None
                data_iso = parse_data(data_ref)

                if not data_iso:
                    continue

                try:
                    dt = pd.to_datetime(data_iso, errors="coerce")
                    if pd.isna(dt):
                        continue
                    return f"{dt.year}-{str(dt.month).zfill(2)}"
                except Exception:
                    continue

        return None

    mes_ref = detectar_mes_ref() or "2026-00"

    records = []
    resumo_records = []
    erros_gerais = []

    for nome_aba, linha_tag in ABAS.items():
        if nome_aba not in wb.sheetnames:
            continue

        ws = wb[nome_aba]
        is_embalagem = linha_tag == "EMBALAGEM"
        quantidade_aba = 0.0

        # Lê a partir da linha 7. A linha 6 é cabeçalho.
        for row in ws.iter_rows(min_row=7, values_only=True):
            if not row or len(row) < 7:
                continue

            lote    = str(row[0] or "").strip()
            codigo  = parse_codigo(row[1])
            produto = str(row[2] or "").strip()
            tempo   = parse_float(row[3])
            un_h    = parse_float(row[4])
            op_numero = parse_op(row[5])
            quantidade = parse_float(row[6])

            if not lote or lote.upper() == "LOTE":
                continue

            if not codigo:
                continue

            # Evita capturar cabeçalhos, textos soltos e linhas decorativas.
            try:
                int(float(str(codigo)))
            except Exception:
                continue

            codigo = str(int(float(codigo))).zfill(5)

            if quantidade <= 0:
                continue

            quantidade_aba += quantidade

            if is_embalagem:
                # H=Destino, I=OBS, J=Previsão Bausch, K=Data Início, L=Data Fim
                observacoes      = str(row[7] or "").strip() if len(row) > 7 else ""
                obs2             = str(row[8] or "").strip() if len(row) > 8 else ""
                observacoes      = observacoes or obs2
                data_inicio_fab  = parse_data(row[10]) if len(row) > 10 else None
                data_fim         = parse_data(row[11]) if len(row) > 11 else None
                data_lav_emb     = None
                data_lav_pesagem = None
                data_termino     = None
            else:
                # H=Obs, I=Data Lav.Êmb, J=Data Lav.Pes, K=Data Início, L=Data Fim, M=Término
                observacoes      = str(row[7] or "").strip() if len(row) > 7 else ""
                data_lav_emb     = parse_data(row[8])  if len(row) > 8  else None
                data_lav_pesagem = parse_data(row[9])  if len(row) > 9  else None
                data_inicio_fab  = parse_data(row[10]) if len(row) > 10 else None
                data_fim         = parse_data(row[11]) if len(row) > 11 else None
                data_termino     = parse_data(row[12]) if len(row) > 12 else None

            records.append({
                "mes_ref":                mes_ref,
                "linha":                  linha_tag,
                "lote":                   lote,
                "codigo":                 codigo,
                "produto":                produto,
                "op_numero":              op_numero,
                "quantidade":             quantidade,
                "tempo_horas":            tempo,
                "un_h":                   un_h,
                "observacoes":            observacoes or None,
                "data_lavagem_emb":       data_lav_emb,
                "data_lavagem_pesagem":   data_lav_pesagem,
                "data_inicio_fabricacao": data_inicio_fab,
                "data_fim":               data_fim,
                "data_termino":           data_termino,
            })

        linha_resumo = normalizar_linha_resumo(linha_tag)

        if linha_resumo:
            # KPI oficial da programação:
            # L4 = META MÊS - TUBETES.
            # N4 = PROG. MÊS - TUBETES, quando existir.
            # O4 = DIF. MÊS - TUBETES, quando existir.
            meta_mes_tubetes = parse_float(ws["L4"].value)
            prog_mes_tubetes = parse_float(ws["N4"].value)
            dif_mes_tubetes = parse_float(ws["O4"].value)

            # Se N4 vier vazio no layout, usa a soma operacional como fallback informativo.
            if prog_mes_tubetes <= 0:
                prog_mes_tubetes = quantidade_aba

            # Se O4 vier vazio, calcula diferença só para deixar preenchido.
            if dif_mes_tubetes == 0 and (meta_mes_tubetes or prog_mes_tubetes):
                dif_mes_tubetes = prog_mes_tubetes - meta_mes_tubetes

            resumo_records.append({
                "mes_ref": mes_ref,
                "linha": linha_resumo,
                "meta_mes_tubetes": meta_mes_tubetes,
                "prog_mes_tubetes": prog_mes_tubetes,
                "dif_mes_tubetes": dif_mes_tubetes,
                "arquivo_origem": filename,
            })

    if not records:
        return 0, ["Nenhuma OP encontrada nas abas ENVASE LINHA 1, ENVASE LINHA 2 ou EMBALAGEM."]

    try:
        supabase.table("f_programacao_ops").delete().eq("mes_ref", mes_ref).execute()
    except Exception as e:
        erros_gerais.append(f"Aviso ao limpar programação do mês anterior: {str(e)[:150]}")

    try:
        supabase.table("f_programacao_ops_resumo").delete().eq("mes_ref", mes_ref).execute()
    except Exception as e:
        erros_gerais.append(f"Aviso ao limpar resumo da programação: {str(e)[:150]}")

    erros = []
    erros.extend(_chunk_insert("f_programacao_ops", records))

    if resumo_records:
        erros.extend(_chunk_insert("f_programacao_ops_resumo", resumo_records))

    erros = erros_gerais + erros

    return len(records) - len(erros), erros


# ─── f_liberacoes_previstas_sku ───────────────────────────────────────────────

def process_liberacoes_previstas_sku(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê a planilha de liberações previstas por SKU e linha.
    Estrutura:
      Linha 0: grupos (DEMANDA, LINHA 1, LINHA 2, ESTOQUE)
      Linha 3: anos
      Linha 4: header — colunas fixas (A-F) + meses por seção
      Linha 5+: dados por produto
    Salva por SKU + linha (L1/L2) + mês + ano.
    Estoque inicial por SKU/mês salvo na linha L1 (evita duplicação).
    """
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 6:
        return 0, ["Arquivo vazio ou fora do formato esperado."]

    grupo_row = rows[0]
    ano_row   = rows[3]
    mes_row   = rows[4]

    l1_cols:  list[tuple[int, int, int]] = []
    l2_cols:  list[tuple[int, int, int]] = []
    est_cols: list[tuple[int, int, int]] = []

    for col_idx, (grupo, ano, mes) in enumerate(zip(grupo_row, ano_row, mes_row)):
        if col_idx < 6:
            continue
        if not isinstance(mes, (int, float)):
            continue
        if grupo == "LINHA 1" and isinstance(ano, (int, float)):
            l1_cols.append((col_idx, int(ano), int(mes)))
        elif grupo == "LINHA 2" and isinstance(ano, (int, float)):
            l2_cols.append((col_idx, int(ano), int(mes)))
        elif grupo == "ESTOQUE" and isinstance(ano, (int, float)):
            est_cols.append((col_idx, int(ano), int(mes)))

    if not l1_cols and not l2_cols:
        return 0, ["Nenhuma coluna de LINHA 1 ou LINHA 2 encontrada. Verifique o formato."]

    records = []

    for row in rows[5:]:
        if not row or not row[0]:
            continue

        try:
            cod_produto = str(row[0]).strip().replace(".0", "")
            if not cod_produto or cod_produto in ("None", ""):
                continue
        except Exception:
            continue

        descricao = str(row[1] or "").strip()
        tipo      = str(row[2] or "").strip()
        grupo     = str(row[3] or "").strip()
        mercado   = str(row[4] or "").strip()

        # Mapa de estoque por (ano, mes)
        estoque_map: dict[tuple[int, int], float] = {}
        for col_idx, ano, mes in est_cols:
            try:
                qtd = float(row[col_idx] or 0)
            except (TypeError, ValueError):
                qtd = 0.0
            estoque_map[(ano, mes)] = max(qtd, 0.0)

        # Verifica se tem liberação em L1 ou L2
        tem_l1 = any(
            float(row[col_idx] or 0) > 0
            for col_idx, ano, mes in l1_cols
        )
        tem_l2 = any(
            float(row[col_idx] or 0) > 0
            for col_idx, ano, mes in l2_cols
        )

        # L1 — inclui estoque_inicial
        for col_idx, ano, mes in l1_cols:
            try:
                qtd = float(row[col_idx] or 0)
            except (TypeError, ValueError):
                qtd = 0.0
            est = estoque_map.get((ano, mes), 0.0)
            # Salva se tem liberação OU tem estoque
            if qtd <= 0 and est <= 0:
                continue
            records.append({
                "ano": ano, "mes": mes, "linha": "L1",
                "cod_produto": cod_produto,
                "descricao": descricao,
                "tipo": tipo,
                "grupo": grupo,
                "mercado": mercado,
                "qtd_caixas": qtd,
                "estoque_inicial": est,
            })

        # L2 — estoque_inicial fica 0 (evita duplicar com L1)
        for col_idx, ano, mes in l2_cols:
            try:
                qtd = float(row[col_idx] or 0)
            except (TypeError, ValueError):
                qtd = 0.0
            if qtd <= 0:
                continue
            records.append({
                "ano": ano, "mes": mes, "linha": "L2",
                "cod_produto": cod_produto,
                "descricao": descricao,
                "tipo": tipo,
                "grupo": grupo,
                "mercado": mercado,
                "qtd_caixas": qtd,
                "estoque_inicial": 0.0,
            })

        # Produto sem linha — salva só o estoque (linha "EST")
        if not tem_l1 and not tem_l2:
            for col_idx, ano, mes in est_cols:
                est = estoque_map.get((ano, mes), 0.0)
                if est <= 0:
                    continue
                records.append({
                    "ano": ano, "mes": mes, "linha": "EST",
                    "cod_produto": cod_produto,
                    "descricao": descricao,
                    "tipo": tipo,
                    "grupo": grupo,
                    "mercado": mercado,
                    "qtd_caixas": 0.0,
                    "estoque_inicial": est,
                })

    if not records:
        return 0, ["Nenhum registro com quantidade > 0 encontrado."]

    _limpar_tabela("f_liberacoes_previstas_sku")
    erros = _chunk_insert("f_liberacoes_previstas_sku", records)

    return len(records) - len(erros), erros


# ─── f_apontamentos ───────────────────────────────────────────────────────────

def process_apontamentos(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê o relatório de apontamentos da Cogtive e insere em f_apontamentos.

    IMPORTANTE:
      - Esta função NÃO limpa mais a f_apontamentos inteira.
      - A limpeza mensal, quando usada, fica no router de upload:
          /upload/apontamentos?modo=replace_month
      - Isso permite subir só o mês atual sem apagar histórico.
      - A referência de data é DATA INICIAL.
      - DURAÇÃO do Cogtive costuma vir como fração de dia; salvamos em horas.
    """
    from io import BytesIO
    from datetime import datetime as datetime_cls, date as date_cls
    from openpyxl import load_workbook
    import unicodedata

    def normaliza_coluna(valor: str) -> str:
        texto = str(valor or "").strip().upper()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = (
            texto
            .replace("Ç", "C")
            .replace(".", "")
            .replace("-", " ")
            .replace("_", " ")
            .replace("/", " ")
        )
        texto = " ".join(texto.split())
        return texto

    def parse_text(valor) -> str:
        if valor is None:
            return ""

        try:
            if pd.isna(valor):
                return ""
        except Exception:
            pass

        texto = str(valor).strip()

        if texto.endswith(".0"):
            texto = texto[:-2]

        if texto.lower() in {"nan", "none", "nat", "null"}:
            return ""

        return texto

    def parse_num(valor) -> float:
        if valor is None:
            return 0.0

        try:
            if pd.isna(valor):
                return 0.0
        except Exception:
            pass

        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return 0.0

                v = v.replace("\xa0", "").replace(" ", "")

                # Formato brasileiro: 1.234,56 -> 1234.56
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def parse_dt(valor):
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        if isinstance(valor, datetime_cls):
            return valor.isoformat()

        if isinstance(valor, date_cls):
            return datetime_cls.combine(valor, datetime_cls.min.time()).isoformat()

        # Serial Excel com decimal, ex.: 46024.6550462963
        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            try:
                numero = float(valor)
                if 20000 <= numero <= 70000:
                    dt = pd.to_datetime(numero, unit="D", origin="1899-12-30")
                    if not pd.isna(dt):
                        return dt.to_pydatetime().isoformat()
            except Exception:
                return None

        texto = str(valor).strip()
        if not texto:
            return None

        # String contendo serial Excel.
        try:
            numero = float(texto.replace(",", "."))
            if 20000 <= numero <= 70000:
                dt = pd.to_datetime(numero, unit="D", origin="1899-12-30")
                if not pd.isna(dt):
                    return dt.to_pydatetime().isoformat()
        except Exception:
            pass

        try:
            dt = pd.to_datetime(texto, dayfirst=True, errors="coerce")
            if pd.isna(dt):
                return None
            return dt.to_pydatetime().isoformat()
        except Exception:
            return None

    def parse_duracao_h(valor, data_inicial=None, data_final=None) -> float:
        """
        Duração em horas do relatório Cogtive.

        Correção v91:
          - no arquivo atual, a coluna DURAÇÃO já vem em horas;
          - exemplo: 0,058333 = 3,5 minutos, não 1,4 hora;
          - quando DATA INICIAL e DATA FINAL existem, usamos a diferença real;
          - isso evita inflar as horas em 24x.
        """
        dt_ini = parse_dt(data_inicial)
        dt_fim = parse_dt(data_final)

        try:
            if dt_ini and dt_fim:
                ini = pd.to_datetime(dt_ini, errors="coerce")
                fim = pd.to_datetime(dt_fim, errors="coerce")
                if not pd.isna(ini) and not pd.isna(fim) and fim > ini:
                    return max(0.0, (fim - ini).total_seconds() / 3600.0)
        except Exception:
            pass

        if valor is None:
            return 0.0

        # Horário/duração textual tipo 01:30:00
        if isinstance(valor, str) and ":" in valor:
            partes = valor.strip().split(":")
            try:
                h = float(partes[0] or 0)
                m = float(partes[1] or 0) if len(partes) > 1 else 0
                s = float(partes[2] or 0) if len(partes) > 2 else 0
                return h + (m / 60.0) + (s / 3600.0)
            except Exception:
                pass

        n = parse_num(valor)

        if n <= 0:
            return 0.0

        # Neste relatório, número puro já é hora.
        return n

    def classifica_etapa(equip: str) -> str:
        e = str(equip or "").upper()

        if "LAVADORA" in e:
            return "LAVAGEM"

        if "ENVASADORA" in e or "ENVASE" in e or " ENV" in f" {e}":
            return "ENVASE"

        if "FABRIMA" in e or "BAUSCH" in e or "EMBAL" in e:
            return "EMBALAGEM"

        return "OUTRO"

    try:
        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as e:
        return 0, [f"Erro ao ler arquivo de apontamentos: {str(e)[:250]}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return 0, ["Arquivo de apontamentos vazio."]

    # Acha linha de header procurando LOTE + DATA INICIAL/EQUIPAMENTO.
    header_idx = None

    for i, row in enumerate(rows[:30]):
        if not row:
            continue

        norm = [normaliza_coluna(v) for v in row]
        tem_lote = "LOTE" in norm
        tem_data = "DATA INICIAL" in norm or "DATA INICIO" in norm
        tem_equip = "EQUIPAMENTO" in norm

        if tem_lote and (tem_data or tem_equip):
            header_idx = i
            break

    if header_idx is None:
        return 0, [
            "Cabeçalho não encontrado. Verifique se o arquivo é o relatório de apontamentos "
            "com as colunas DATA INICIAL, EQUIPAMENTO, LOTE e QUANTIDADE PRODUZIDA."
        ]

    header = [str(h or "").strip() for h in rows[header_idx]]
    header_norm = [normaliza_coluna(h) for h in header]

    def idx(*nomes):
        nomes_norm = {normaliza_coluna(n) for n in nomes}

        for i, h in enumerate(header_norm):
            if h in nomes_norm:
                return i

        return None

    i_data_inicial = idx("DATA INICIAL", "DATA INICIO", "INICIO", "DATA HORA INICIAL")
    i_data_final = idx("DATA FINAL", "FIM", "DATA HORA FINAL")
    i_duracao = idx("DURAÇÃO", "DURACAO", "DURAÇÃO H", "DURACAO H")
    i_tag = idx("ID", "TAG")
    i_equipamento = idx("EQUIPAMENTO", "MAQUINA", "MÁQUINA")
    i_ordem = idx("ORDEM", "OP", "ORDEM PRODUCAO", "ORDEM DE PRODUCAO")
    i_lote = idx("LOTE")
    i_produto = idx("PRODUTO", "DESCRICAO PRODUTO", "DESCRIÇÃO PRODUTO")
    i_sku = idx("SKU", "CODIGO", "CÓDIGO", "COD PRODUTO", "CODIGO PRODUTO")
    i_qtd_produzida = idx("QUANTIDADE PRODUZIDA", "QTD PRODUZIDA", "QTD PRODUCAO", "QUANTIDADE PRODUCAO")
    i_qtd_rejeitada = idx("QUANTIDADE REJEITADA", "QTD REJEITADA", "REFUGO")
    i_tipo_evento = idx("TIPO DE EVENTO", "TIPO EVENTO", "TIPO")
    i_evento = idx("EVENTO", "MOTIVO", "DESCRICAO EVENTO", "DESCRIÇÃO EVENTO")
    i_situacao = idx("SITUAÇÃO", "SITUACAO", "STATUS")

    faltando = []
    if i_data_inicial is None:
        faltando.append("DATA INICIAL")
    if i_equipamento is None:
        faltando.append("EQUIPAMENTO")
    if i_lote is None:
        faltando.append("LOTE")
    if i_qtd_produzida is None:
        faltando.append("QUANTIDADE PRODUZIDA")
    if i_tipo_evento is None:
        faltando.append("TIPO DE EVENTO")

    if faltando:
        return 0, [
            "Colunas obrigatórias não encontradas em apontamentos: "
            + ", ".join(faltando)
            + "."
        ]

    def get(row, i):
        if i is None:
            return None
        if i >= len(row):
            return None
        return row[i]

    records = []
    registros_ignorados_sem_data = 0

    for row in rows[header_idx + 1:]:
        if not row:
            continue

        lote = parse_text(get(row, i_lote))
        if not lote or lote.upper() == "LOTE":
            continue

        data_inicial = parse_dt(get(row, i_data_inicial))

        if not data_inicial:
            registros_ignorados_sem_data += 1
            continue

        equipamento = parse_text(get(row, i_equipamento))
        etapa = classifica_etapa(equipamento)

        records.append({
            "data_inicial":   data_inicial,
            "data_final":     parse_dt(get(row, i_data_final)),
            "duracao_h":      parse_duracao_h(get(row, i_duracao), get(row, i_data_inicial), get(row, i_data_final)),
            "tag":            parse_text(get(row, i_tag)),
            "equipamento":    equipamento,
            "etapa":          etapa,
            "ordem":          parse_text(get(row, i_ordem)),
            "lote":           lote,
            "produto":        parse_text(get(row, i_produto)),
            "sku":            parse_text(get(row, i_sku)),
            "qtd_produzida":  parse_num(get(row, i_qtd_produzida)),
            "qtd_rejeitada":  parse_num(get(row, i_qtd_rejeitada)),
            "tipo_evento":    parse_text(get(row, i_tipo_evento)),
            "evento":         parse_text(get(row, i_evento)),
            "situacao":       parse_text(get(row, i_situacao)),
        })

    if not records:
        return 0, ["Nenhum apontamento válido encontrado. Verifique o arquivo."]

    erros = []

    if registros_ignorados_sem_data:
        erros.append(
            f"Aviso: {registros_ignorados_sem_data} linhas foram ignoradas por falta de DATA INICIAL válida."
        )

    # NÃO limpar aqui.
    # A limpeza por mês é feita no router upload.py antes de chamar este processador.
    erros_insert = _chunk_insert("f_apontamentos", records)
    erros.extend(erros_insert)

    return len(records) - len(erros_insert), erros


# ─── f_liberacao_diaria ───────────────────────────────────────────────────────

def process_liberacao_diaria(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê o Gantt de liberação diária (liberacaodia.xlsx) e agrega por
    grupo_produto + data_lib + linha para alimentar o "Previsto até Hoje".

    Estrutura do Excel:
      LOTE, CÓDIGO, PRODUTO, TEMPO (Horas.), UN / HORA, QTD. (Tubetes),
      MÊS PRODUÇÃO, ANO PRODUÇÃO, DATA INÍCIO, DATA FIM,
      DATA LIB., MÊS LIBERAÇÃO, ANO LIBERAÇÃO

    Regras:
      - qtd_prevista salva em CAIXAS (QTD Tubetes / 500)
      - Linha detectada pelo lote: 2605F1026 → L1 (dígito após letra = linha)
      - Agrupa por (ano, mes, data_lib, grupo_produto, linha)
      - Substitui tudo a cada upload
    """
    import re
    from io import BytesIO
    from collections import defaultdict
    from openpyxl import load_workbook

    TUBETES_POR_CAIXA = 500

    PRODUTO_GRUPO = {
        "40295": "ALPHACAINE", "40327": "ALPHACAINE 80",
        "40319": "ARTICAINE",  "40323": "ARTICAINE 200",
        "40303": "MEPIADRE",
        "40299": "MEPISV",
        "40315": "PRILONEST",
    }

    def detecta_linha(lote: str) -> str | None:
        m = re.search(r"\d{4}[A-Za-z](\d)", str(lote))
        if m:
            return f"L{m.group(1)}"
        return None

    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return 0, ["Arquivo vazio."]

    header = [str(h or "").strip().upper() for h in rows[0]]

    def col(nome):
        try:
            return header.index(nome.upper())
        except ValueError:
            return None

    i_lote    = col("LOTE")
    i_codigo  = col("CÓDIGO")
    i_produto = col("PRODUTO")
    i_qtd     = col("QTD. (Tubetes)")
    i_mes_lib = col("MÊS LIBERAÇÃO")
    i_ano_lib = col("ANO LIBERAÇÃO")
    i_data_lib = col("DATA LIB.")
    i_data_ini = col("DATA INÍCIO")
    i_data_fim = col("DATA FIM")

    faltando = [n for n, i in [
        ("LOTE", i_lote), ("QTD. (Tubetes)", i_qtd),
        ("MÊS LIBERAÇÃO", i_mes_lib), ("ANO LIBERAÇÃO", i_ano_lib),
        ("DATA LIB.", i_data_lib),
    ] if i is None]

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}"]

    records = []

    for row in rows[1:]:
        if not row:
            continue

        # Lote pode ser vazio (lotes futuros sem número definido)
        lote_raw = row[i_lote]
        lote = str(lote_raw).strip() if lote_raw else ""
        if lote.upper() == "LOTE":
            continue

        try:
            qtd_tb = float(row[i_qtd] or 0)
        except Exception:
            qtd_tb = 0.0

        try:
            mes_lib = int(row[i_mes_lib] or 0)
            ano_lib = int(row[i_ano_lib] or 0)
        except Exception:
            continue

        if mes_lib <= 0 or ano_lib <= 0:
            continue

        # Código e grupo
        try:
            codigo = str(int(float(row[i_codigo] or 0))).zfill(5) if i_codigo is not None and row[i_codigo] else ""
        except Exception:
            codigo = ""
        produto = str(row[i_produto] or "").strip() if i_produto is not None else ""
        grupo = PRODUTO_GRUPO.get(codigo, produto)

        # Linha pelo lote
        linha = detecta_linha(lote) if lote else None

        # Data lib
        data_lib_raw = row[i_data_lib]
        if hasattr(data_lib_raw, "date"):
            data_lib = data_lib_raw.date().isoformat()
        else:
            continue  # sem data_lib, pula

        # Data início/fim
        data_ini_raw = row[i_data_ini] if i_data_ini is not None else None
        data_ini = data_ini_raw.date().isoformat() if hasattr(data_ini_raw, "date") else None

        data_fim_raw = row[i_data_fim] if i_data_fim is not None else None
        data_fim = data_fim_raw.date().isoformat() if hasattr(data_fim_raw, "date") else None

        qtd_cx = qtd_tb / TUBETES_POR_CAIXA

        records.append({
            "ano":           ano_lib,
            "mes":           mes_lib,
            "lote":          lote or None,
            "data_lib":      data_lib,
            "grupo_produto": grupo,
            "linha":         linha,
            "qtd_prevista":  qtd_cx,
            "data_inicio":   data_ini,
            "data_fim":      data_fim,
        })

    if not records:
        return 0, ["Nenhum registro válido encontrado."]

    # Substitui tudo
    supabase.table("f_liberacao_diaria").delete().not_.is_("data_lib", "null").execute()
    erros = _chunk_insert("f_liberacao_diaria", records)

    return len(records) - len(erros), erros


# ─── f_compras_abertas ────────────────────────────────────────────────────────

def process_compras_abertas(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê o relatório de compras do Protheus:
      RELATORIO ENTREGAS DE PC E SC SEM PC (DFLRELPC) - EXPORTAR EXCEL

    Objetivo:
      alimentar a tabela f_compras_abertas para cruzar com gargalos/FIFO da página de OP.

    Regras:
      - considera SCs e PCs em aberto;
      - calcula quantidade_pendente = QUANTIDADE_PC - QUANTIDADE_ENTREGUE;
      - quando não houver PC, usa QUANTIDADE DA SA como pendente;
      - ignora registros sem produto;
      - salva apenas registros com quantidade pendente > 0 ou SC sem pedido;
      - substitui tudo a cada upload.
    """
    from io import BytesIO
    from openpyxl import load_workbook
    from datetime import date as date_cls, datetime as datetime_cls

    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, ["Arquivo vazio."]

    def normaliza_coluna(valor: str) -> str:
        return (
            str(valor or "")
            .strip()
            .upper()
            .replace(" ", "")
            .replace(".", "")
            .replace("-", "")
            .replace("_", "")
            .replace("/", "")
            .replace("Ç", "C")
            .replace("Ã", "A")
            .replace("Á", "A")
            .replace("Â", "A")
            .replace("À", "A")
            .replace("É", "E")
            .replace("Ê", "E")
            .replace("Í", "I")
            .replace("Ó", "O")
            .replace("Ô", "O")
            .replace("Ú", "U")
        )

    def achar_header_idx() -> int | None:
        for i, row in enumerate(rows[:20]):
            row_norm = [normaliza_coluna(v) for v in row]
            if "PRODUTOCODIGO" in row_norm and (
                "SCNUMERO" in row_norm or "PEDIDONUMERO" in row_norm
            ):
                return i
        return None

    header_idx = achar_header_idx()
    if header_idx is None:
        return 0, [
            "Cabeçalho não encontrado. Verifique se é o relatório de compras do Protheus "
            "com as colunas PRODUTO_CODIGO, SC_NUMERO e/ou PEDIDO_NUMERO."
        ]

    header = [str(h or "").strip() for h in rows[header_idx]]
    header_norm = [normaliza_coluna(h) for h in header]
    data_rows = rows[header_idx + 1:]

    def idx(*nomes):
        for nome in nomes:
            nome_norm = normaliza_coluna(nome)
            if nome_norm in header_norm:
                return header_norm.index(nome_norm)
        return None

    def parse_text(valor) -> str | None:
        if valor is None:
            return None

        texto = str(valor).strip()
        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return None

        if texto.endswith(".0"):
            texto = texto[:-2]

        return texto

    def parse_codigo(valor, zfill: int | None = None) -> str | None:
        if valor is None:
            return None

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()

        if not texto or texto.lower() in {"nan", "none"}:
            return None

        return texto.zfill(zfill) if zfill else texto

    def parse_num(valor) -> float:
        if valor is None:
            return 0.0

        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return 0.0

                # Padrão BR: 1.234,56 -> 1234.56
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def parse_int(valor) -> int | None:
        if valor is None:
            return None

        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return None
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")
                return int(float(v))

            return int(float(valor))

        except Exception:
            return None

    def parse_date(valor):
        if valor is None:
            return None

        try:
            if isinstance(valor, datetime_cls):
                return valor.date().isoformat()

            if isinstance(valor, date_cls):
                return valor.isoformat()

            dt = pd.to_datetime(valor, dayfirst=True, errors="coerce")
            if pd.isna(dt):
                return None

            return dt.date().isoformat()

        except Exception:
            return None

    i_sc_status           = idx("SC_STATUS")
    i_sc_solicitante      = idx("SC_SOLICITANTE")
    i_sc_numero           = idx("SC_NUMERO")
    i_sc_item             = idx("SC_ITEM")
    i_sc_emissao          = idx("SC_EMISSAO")
    i_sc_aprovador        = idx("SC_APROVADOR")
    i_sc_aprovacao        = idx("SC_APROVACAO")
    i_cotacao             = idx("COTACAO")
    i_cod_fornecedor      = idx("CODIGO_FORNECEDOR")
    i_fornecedor          = idx("RAZAO_SOCIAL_FORNECEDOR")
    i_pedido_emissao      = idx("PEDIDO_EMISSAO")
    i_pedido_numero       = idx("PEDIDO_NUMERO")
    i_pedido_item         = idx("PEDIDO_ITEM")
    i_produto_codigo      = idx("PRODUTO_CODIGO")
    i_qtd_sa              = idx("QUANTIDADE DA SA", "QUANTIDADE_DA_SA", "QUANTIDADE SA")
    i_produto_descricao   = idx("PRODUTO_DESCRICAO")
    i_produto_tipo        = idx("PRODUTO_TIPO")
    i_produto_grupo       = idx("PRODUTO_GRUPO")
    i_produto_grupo_desc  = idx("PRODUTO_GRUPO_DESC")
    i_pedido_comprador    = idx("PEDIDO_COMPRADOR")
    i_pedido_dt_aprov     = idx("PEDIDO_DATA_APROVACAO")
    i_cc_codigo           = idx("CC_CODIGO")
    i_cc_descricao        = idx("CC_DESCRICAO")
    i_data_prev_entrega   = idx("DATA_PREVISTA_ENTREGA")
    i_data_recebimento    = idx("DATA_RECEBIMENTO")
    i_qtd_pc              = idx("QUANTIDADE_PC")
    i_qtd_entregue        = idx("QUANTIDADE_ENTREGUE")
    i_mes_necessidade     = idx("MES NECESSIDADE", "MES_NECESSIDADE")
    i_ano_necessidade     = idx("ANO NECESSIDADE", "ANO_NECESSIDADE")
    i_data_prev_necess    = idx("DATA_PREVISAO_NECESSIDADE")
    i_id_comprador        = idx("ID DO COMPRADOR", "ID_COMPRADOR")
    i_comprador           = idx("COMPRADOR")
    i_atraso_entrega      = idx("ATRASO NA ENTREGA", "ATRASO_ENTREGA")
    i_situacao_data_exata = idx("SITUACAO - DATA EXATA", "SITUACAO_DATA_EXATA")
    i_entrega_status      = idx("ENTREGA STATUS", "ENTREGA_STATUS")
    i_tempo_aprov_sc      = idx("TEMPO APROVACAO SC", "TEMPO_APROVACAO_SC")
    i_atraso_range        = idx("ATRASO NA ENTREGA - COM RANGE", "ATRASO_ENTREGA_RANGE")
    i_situacao_range      = idx("SITUACAO - RANGE", "SITUACAO_RANGE")

    faltando = [n for n, i in [
        ("PRODUTO_CODIGO", i_produto_codigo),
    ] if i is None]

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o formato."]

    records = []

    for row in data_rows:
        if not row:
            continue

        produto_codigo = parse_codigo(row[i_produto_codigo], zfill=5) if i_produto_codigo is not None else None
        if not produto_codigo:
            continue

        qtd_sa = parse_num(row[i_qtd_sa]) if i_qtd_sa is not None else 0.0
        qtd_pc = parse_num(row[i_qtd_pc]) if i_qtd_pc is not None else 0.0
        qtd_entregue = parse_num(row[i_qtd_entregue]) if i_qtd_entregue is not None else 0.0

        pedido_numero = parse_codigo(row[i_pedido_numero]) if i_pedido_numero is not None else None
        sc_numero = parse_codigo(row[i_sc_numero]) if i_sc_numero is not None else None

        # Se tem PC, pendente = pedido - entregue.
        # Se ainda não tem PC, usa a quantidade da SA/SC para mostrar que existe solicitação.
        if qtd_pc > 0:
            quantidade_pendente = max(qtd_pc - qtd_entregue, 0.0)
        else:
            quantidade_pendente = max(qtd_sa - qtd_entregue, 0.0)

        tem_sc_sem_pedido = bool(sc_numero) and not bool(pedido_numero)

        # Ignora linhas completamente encerradas, mas mantém SC sem pedido
        # mesmo que a quantidade venha zerada, para rastrear necessidade sem PC.
        if quantidade_pendente <= 0 and not tem_sc_sem_pedido:
            continue

        comprador = parse_text(row[i_comprador]) if i_comprador is not None else None

        records.append({
            "sc_status":              parse_text(row[i_sc_status]) if i_sc_status is not None else None,
            "sc_solicitante":         parse_text(row[i_sc_solicitante]) if i_sc_solicitante is not None else None,
            "sc_numero":              sc_numero,
            "sc_item":                parse_codigo(row[i_sc_item]) if i_sc_item is not None else None,
            "sc_emissao":             parse_date(row[i_sc_emissao]) if i_sc_emissao is not None else None,
            "sc_aprovador":           parse_text(row[i_sc_aprovador]) if i_sc_aprovador is not None else None,
            "sc_aprovacao":           parse_date(row[i_sc_aprovacao]) if i_sc_aprovacao is not None else None,
            "cotacao":                parse_codigo(row[i_cotacao]) if i_cotacao is not None else None,

            "codigo_fornecedor":      parse_codigo(row[i_cod_fornecedor]) if i_cod_fornecedor is not None else None,
            "razao_social_fornecedor": parse_text(row[i_fornecedor]) if i_fornecedor is not None else None,

            "pedido_emissao":         parse_date(row[i_pedido_emissao]) if i_pedido_emissao is not None else None,
            "pedido_numero":          pedido_numero,
            "pedido_item":            parse_codigo(row[i_pedido_item]) if i_pedido_item is not None else None,

            "produto_codigo":         produto_codigo,
            "produto_descricao":      parse_text(row[i_produto_descricao]) if i_produto_descricao is not None else None,
            "produto_tipo":           parse_text(row[i_produto_tipo]) if i_produto_tipo is not None else None,
            "produto_grupo":          parse_codigo(row[i_produto_grupo]) if i_produto_grupo is not None else None,
            "produto_grupo_desc":     parse_text(row[i_produto_grupo_desc]) if i_produto_grupo_desc is not None else None,

            "quantidade_sa":          qtd_sa,
            "quantidade_pc":          qtd_pc,
            "quantidade_entregue":    qtd_entregue,
            "quantidade_pendente":    quantidade_pendente,

            "pedido_comprador":       parse_text(row[i_pedido_comprador]) if i_pedido_comprador is not None else None,
            "pedido_data_aprovacao":  parse_date(row[i_pedido_dt_aprov]) if i_pedido_dt_aprov is not None else None,

            "cc_codigo":              parse_codigo(row[i_cc_codigo]) if i_cc_codigo is not None else None,
            "cc_descricao":           parse_text(row[i_cc_descricao]) if i_cc_descricao is not None else None,

            "data_prevista_entrega":  parse_date(row[i_data_prev_entrega]) if i_data_prev_entrega is not None else None,
            "data_recebimento":       parse_date(row[i_data_recebimento]) if i_data_recebimento is not None else None,

            "mes_necessidade":        parse_int(row[i_mes_necessidade]) if i_mes_necessidade is not None else None,
            "ano_necessidade":        parse_int(row[i_ano_necessidade]) if i_ano_necessidade is not None else None,
            "data_previsao_necessidade": parse_date(row[i_data_prev_necess]) if i_data_prev_necess is not None else None,

            "id_comprador":           parse_codigo(row[i_id_comprador]) if i_id_comprador is not None else None,
            "comprador":              comprador,
            "comprador_nome":         comprador,
            "comprador_email":        None,

            "atraso_entrega":         parse_int(row[i_atraso_entrega]) if i_atraso_entrega is not None else None,
            "situacao_data_exata":    parse_text(row[i_situacao_data_exata]) if i_situacao_data_exata is not None else None,
            "entrega_status":         parse_text(row[i_entrega_status]) if i_entrega_status is not None else None,
            "tempo_aprovacao_sc":     parse_int(row[i_tempo_aprov_sc]) if i_tempo_aprov_sc is not None else None,
            "atraso_entrega_range":   parse_int(row[i_atraso_range]) if i_atraso_range is not None else None,
            "situacao_range":         parse_text(row[i_situacao_range]) if i_situacao_range is not None else None,
        })

    if not records:
        return 0, ["Nenhuma compra/SC em aberto encontrada no arquivo."]

    # Substitui tudo a cada upload.
    try:
        supabase.table("f_compras_abertas").delete().not_.is_("produto_codigo", "null").execute()
    except Exception:
        # Fallback caso o PostgREST implique com not_.is_
        supabase.table("f_compras_abertas").delete().neq("produto_codigo", "000000000000").execute()

    erros = _chunk_insert("f_compras_abertas", records)

    return len(records) - len(erros), erros



# ─── f_compras_fup ────────────────────────────────────────────────────────────

def process_compras_fup(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Processa a planilha de reunião/FUP de compras.

    Regra:
      - lê automaticamente todas as abas cujo nome começa com "Detalhes";
      - a coluna Coluna1 vira comentario_fup;
      - tenta extrair uma nova previsão de chegada do comentário;
      - substitui o snapshot inteiro da tabela f_compras_fup a cada upload.

    Tabela destino esperada:
      public.f_compras_fup
    """
    from io import BytesIO
    from datetime import date as date_cls, datetime as datetime_cls
    import unicodedata

    def normaliza_coluna(valor: str) -> str:
        texto = str(valor or "").strip().upper()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = (
            texto
            .replace(" ", "")
            .replace(".", "")
            .replace("-", "")
            .replace("_", "")
            .replace("/", "")
            .replace("º", "")
            .replace("ª", "")
        )
        return texto

    def normaliza_aba(valor: str) -> str:
        texto = str(valor or "").strip().upper()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = re.sub(r"\s+", "", texto)
        return texto

    def parse_text(valor):
        if valor is None:
            return None
        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass
        texto = str(valor).strip()
        if texto.endswith(".0"):
            texto = texto[:-2]
        if not texto or texto.lower() in {"nan", "none", "nat", "null"}:
            return None
        return texto

    def parse_codigo(valor, zfill: int | None = None):
        texto = parse_text(valor)
        if not texto:
            return None
        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
        except Exception:
            pass
        texto = texto.strip()
        if texto.endswith(".0"):
            texto = texto[:-2]
        return texto.zfill(zfill) if zfill and texto.isdigit() else texto

    def parse_num(valor) -> float:
        if valor is None:
            return 0.0
        try:
            if pd.isna(valor):
                return 0.0
        except Exception:
            pass
        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return 0.0
                v = v.replace("\xa0", "").replace(" ", "")
                if "," in v:
                    v = v.replace(".", "").replace(",", ".")
                return float(v)
            return float(valor or 0)
        except Exception:
            return 0.0

    def parse_date(valor):
        if valor is None:
            return None
        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass
        try:
            if isinstance(valor, datetime_cls):
                return valor.date().isoformat()
            if isinstance(valor, date_cls):
                return valor.isoformat()
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                numero = float(valor)
                if 20000 <= numero <= 70000:
                    dt = pd.to_datetime(numero, unit="D", origin="1899-12-30", errors="coerce")
                    if not pd.isna(dt):
                        return dt.date().isoformat()
            texto = str(valor).strip()
            if not texto:
                return None
            try:
                numero = float(texto.replace(",", "."))
                if 20000 <= numero <= 70000:
                    dt = pd.to_datetime(numero, unit="D", origin="1899-12-30", errors="coerce")
                    if not pd.isna(dt):
                        return dt.date().isoformat()
            except Exception:
                pass
            dt = pd.to_datetime(texto, dayfirst=True, errors="coerce")
            if pd.isna(dt):
                return None
            return dt.date().isoformat()
        except Exception:
            return None

    def data_base_year(*datas):
        for data_iso in datas:
            try:
                if data_iso:
                    return int(str(data_iso)[:4])
            except Exception:
                continue
        return datetime.now().year

    def extrair_nova_previsao(comentario: str | None, ano_base: int) -> str | None:
        texto = str(comentario or "").strip().upper()
        if not texto:
            return None

        # "PREVISÃO ERA DIA 30/06" é histórico/atraso, não nova promessa.
        if "ERA DIA" in texto or "ERA" in texto and "AGUARD" in texto:
            return None

        m = re.search(r"\b(\d{1,2})[\/\.-](\d{1,2})(?:[\/\.-](\d{2,4}))?\b", texto)
        if not m:
            return None

        dia = int(m.group(1))
        mes = int(m.group(2))
        ano_txt = m.group(3)
        if ano_txt:
            ano = int(ano_txt)
            if ano < 100:
                ano += 2000
        else:
            ano = ano_base

        try:
            return date_cls(ano, mes, dia).isoformat()
        except Exception:
            return None

    def classificar_status_fup(comentario: str | None, nova_previsao: str | None, entrega_original: str | None) -> str | None:
        texto = str(comentario or "").strip().upper()
        entrega_dt = None
        try:
            if entrega_original:
                entrega_dt = datetime.strptime(str(entrega_original)[:10], "%Y-%m-%d").date()
        except Exception:
            entrega_dt = None

        if "AGUARD" in texto or "RETORNO" in texto:
            return "Aguardando retorno"
        if "FATUR" in texto and nova_previsao:
            return "Faturamento previsto"
        if nova_previsao:
            return "Com nova previsão"
        if entrega_dt and entrega_dt < date_cls.today():
            return "Atrasado"
        if texto:
            return "Comentário FUP"
        return None

    try:
        excel = pd.ExcelFile(BytesIO(conteudo))
    except Exception as e:
        return 0, [f"Erro ao ler arquivo de FUP de compras: {str(e)[:200]}"]

    abas_detalhes = [aba for aba in excel.sheet_names if normaliza_aba(aba).startswith("DETALHES")]
    if not abas_detalhes:
        return 0, ["Nenhuma aba Detalhes* encontrada. Ex.: Detalhes1, Detalhes2, Detalhes3."]

    registros: list[dict] = []
    avisos: list[str] = []

    for aba in abas_detalhes:
        df = None
        header_usado = None
        for h in range(0, 15):
            try:
                teste = pd.read_excel(BytesIO(conteudo), sheet_name=aba, header=h, nrows=0)
                cols_norm = {normaliza_coluna(c) for c in teste.columns}
                tem_produto = "PRODUTOCODIGO" in cols_norm
                tem_pedido = "PEDIDONUMERO" in cols_norm or "SCNUMERO" in cols_norm
                tem_comentario = "COLUNA1" in cols_norm or "COMENTARIOFUP" in cols_norm or "COMENTARIO" in cols_norm
                if tem_produto and tem_pedido and tem_comentario:
                    df = pd.read_excel(BytesIO(conteudo), sheet_name=aba, header=h)
                    header_usado = h
                    break
            except Exception:
                continue

        if df is None:
            avisos.append(f"Aba {aba}: cabeçalho Detalhes não encontrado; aba ignorada.")
            continue

        df = df.dropna(how="all").copy()
        col_norm_map = {normaliza_coluna(c): c for c in df.columns}

        def col(*nomes):
            for nome in nomes:
                c = col_norm_map.get(normaliza_coluna(nome))
                if c is not None:
                    return c
            return None

        c_produto = col("PRODUTO_CODIGO", "PRODUTO CODIGO", "CODIGO PRODUTO")
        c_desc = col("PRODUTO_DESCRICAO", "PRODUTO DESCRICAO", "DESCRICAO")
        c_tipo = col("PRODUTO_TIPO", "PRODUTO TIPO", "TIPO")
        c_grupo = col("PRODUTO_GRUPO", "PRODUTO GRUPO")
        c_pedido = col("PEDIDO_NUMERO", "PEDIDO NUMERO", "PEDIDO")
        c_pedido_item = col("PEDIDO_ITEM", "PEDIDO ITEM", "ITEM PEDIDO")
        c_sc = col("SC_NUMERO", "SC NUMERO", "SC")
        c_sc_item = col("SC_ITEM", "SC ITEM")
        c_qtd = col("QUANTIDADE DA SA", "QUANTIDADE_DA_SA", "QUANTIDADE SA", "QUANTIDADE", "QTD")
        c_qtd_pc = col("QUANTIDADE_PC", "QUANTIDADE PC")
        c_qtd_entregue = col("QUANTIDADE_ENTREGUE", "QUANTIDADE ENTREGUE")
        c_pedido_emissao = col("PEDIDO_EMISSAO", "PEDIDO EMISSAO", "EMISSAO PEDIDO")
        c_sc_emissao = col("SC_EMISSAO", "SC EMISSAO")
        c_entrega = col("DATA_PREVISTA_ENTREGA", "DATA PREVISTA ENTREGA", "ENTREGA")
        c_fornecedor = col("RAZAO_SOCIAL_FORNECEDOR", "RAZAO SOCIAL FORNECEDOR", "FORNECEDOR")
        c_comprador = col("COMPRADOR", "COMPRADOR_NOME", "PEDIDO_COMPRADOR")
        c_coluna1 = col("Coluna1", "COLUNA1", "COMENTARIO FUP", "COMENTARIO", "OBS")

        if c_produto is None or c_coluna1 is None:
            avisos.append(f"Aba {aba}: PRODUTO_CODIGO ou Coluna1 ausente; aba ignorada.")
            continue

        for idx, row in df.iterrows():
            produto_codigo = parse_codigo(row.get(c_produto), zfill=5)
            if not produto_codigo:
                continue

            comentario = parse_text(row.get(c_coluna1))
            if not comentario:
                # A planilha pode ter linhas de base sem atualização; elas não precisam ir para a camada FUP.
                continue

            qtd_sa = parse_num(row.get(c_qtd)) if c_qtd is not None else 0.0
            qtd_pc = parse_num(row.get(c_qtd_pc)) if c_qtd_pc is not None else 0.0
            qtd_entregue = parse_num(row.get(c_qtd_entregue)) if c_qtd_entregue is not None else 0.0
            quantidade_pendente = max((qtd_pc if qtd_pc > 0 else qtd_sa) - qtd_entregue, 0.0)
            if quantidade_pendente <= 0 and qtd_sa > 0:
                quantidade_pendente = qtd_sa

            pedido_emissao = parse_date(row.get(c_pedido_emissao)) if c_pedido_emissao is not None else None
            sc_emissao = parse_date(row.get(c_sc_emissao)) if c_sc_emissao is not None else None
            data_entrega = parse_date(row.get(c_entrega)) if c_entrega is not None else None
            nova_previsao = extrair_nova_previsao(comentario, data_base_year(data_entrega, pedido_emissao, sc_emissao))
            status_fup = classificar_status_fup(comentario, nova_previsao, data_entrega)

            registros.append({
                "arquivo_origem": filename,
                "aba_origem": str(aba),
                "linha_excel": int(idx) + int(header_usado or 0) + 2,
                "produto_codigo": produto_codigo,
                "produto_descricao": parse_text(row.get(c_desc)) if c_desc is not None else None,
                "produto_tipo": parse_text(row.get(c_tipo)) if c_tipo is not None else None,
                "produto_grupo": parse_codigo(row.get(c_grupo)) if c_grupo is not None else None,
                "sc_numero": parse_codigo(row.get(c_sc)) if c_sc is not None else None,
                "sc_item": parse_codigo(row.get(c_sc_item)) if c_sc_item is not None else None,
                "pedido_numero": parse_codigo(row.get(c_pedido)) if c_pedido is not None else None,
                "pedido_item": parse_codigo(row.get(c_pedido_item)) if c_pedido_item is not None else None,
                "quantidade_sa": qtd_sa,
                "quantidade_pendente": quantidade_pendente,
                "pedido_emissao": pedido_emissao,
                "sc_emissao": sc_emissao,
                "data_prevista_entrega_original": data_entrega,
                "nova_previsao_fup": nova_previsao,
                "comentario_fup": comentario,
                "status_fup": status_fup,
                "fornecedor": parse_text(row.get(c_fornecedor)) if c_fornecedor is not None else None,
                "comprador": parse_text(row.get(c_comprador)) if c_comprador is not None else None,
            })

    if not registros:
        return 0, avisos or ["Nenhum comentário FUP válido encontrado nas abas Detalhes*." ]

    try:
        supabase.table("f_compras_fup").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    except Exception as e:
        return 0, [f"Erro ao limpar f_compras_fup. Verifique se a tabela foi criada no Supabase: {str(e)[:300]}"]

    erros = _chunk_insert("f_compras_fup", registros)
    return len(registros) - len(erros), avisos + erros

def process_calendario_paradas(conteudo, nome_arquivo):
    erros = []
    total = 0

    abas = ["L1", "L2", "FABRIMA"]

    try:
        excel = pd.ExcelFile(io.BytesIO(conteudo))

        registros = []

        for aba in abas:
            if aba not in excel.sheet_names:
                continue

            df = pd.read_excel(
                io.BytesIO(conteudo),
                sheet_name=aba,
                header=None
            )

            for row_idx in range(len(df) - 1):

                linha_datas = df.iloc[row_idx]
                linha_eventos = df.iloc[row_idx + 1]

                for col_idx in range(len(df.columns)):

                    valor_data = linha_datas[col_idx]
                    valor_evento = linha_eventos[col_idx]

                    if pd.isna(valor_data):
                        continue

                    if pd.isna(valor_evento):
                        continue

                    if not isinstance(valor_evento, str):
                        continue

                    descricao = str(valor_evento).strip()

                    if not descricao:
                        continue

                    try:
                        data = pd.to_datetime(valor_data).date()
                    except Exception:
                        continue

                    registros.append({
                        "data": str(data),
                        "linha": aba,
                        "descricao": descricao,
                        "origem": nome_arquivo,
                    })

        # remove duplicados
        registros_unicos = []
        vistos = set()

        for r in registros:
            chave = (
                r["data"],
                r["linha"],
                r["descricao"]
            )

            if chave in vistos:
                continue

            vistos.add(chave)
            registros_unicos.append(r)

        # limpa tabela antes
        supabase.table("f_calendario_paradas")\
            .delete()\
            .neq("id", "00000000-0000-0000-0000-000000000000")\
            .execute()

        # insere
        if registros_unicos:
            supabase.table("f_calendario_paradas")\
                .insert(registros_unicos)\
                .execute()

        total = len(registros_unicos)

    except Exception as e:
        erros.append(str(e))

    return total, erros
# ─── f_consumo_materiais ─────────────────────────────────────────────────────

def process_consumo_materiais(conteudo: bytes, filename: str) -> Tuple[int, list]:
    """
    Lê a base de consumo de materiais e salva histórico em f_consumo_materiais.

    Importante:
      - NÃO limpa a tabela.
      - Cada upload vira um novo snapshot.
      - Usado na página Análise MRP para comparar saldo, consumo histórico e risco.
    """
    from io import BytesIO
    from datetime import datetime, timezone
    import re

    def normaliza_coluna(valor: str) -> str:
        return (
            str(valor or "")
            .strip()
            .lower()
            .replace(" ", "_")
            .replace(".", "")
            .replace("-", "_")
            .replace("/", "_")
            .replace("ç", "c")
            .replace("ã", "a")
            .replace("á", "a")
            .replace("à", "a")
            .replace("â", "a")
            .replace("é", "e")
            .replace("ê", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ô", "o")
            .replace("õ", "o")
            .replace("ú", "u")
        )

    def parse_num(valor) -> float:
        if valor is None or pd.isna(valor):
            return 0.0

        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return 0.0

                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def parse_codigo(valor) -> str | None:
        if valor is None or pd.isna(valor):
            return None

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()

        texto = texto.strip()

        if not texto or texto.lower() in {"nan", "none"}:
            return None

        if texto.isdigit():
            return texto.zfill(5)

        return texto

    def achar_header(conteudo_bytes: bytes) -> int:
        for h in range(0, 10):
            try:
                df_test = pd.read_excel(
                    BytesIO(conteudo_bytes),
                    sheet_name=0,
                    header=h,
                    nrows=0,
                )

                cols_norm = [normaliza_coluna(c) for c in df_test.columns]

                tem_codigo = any(c in {"codigo", "cod", "cod_produto", "produto_codigo"} for c in cols_norm)
                tem_produto = any(c in {"produto", "descricao", "descricao_produto"} for c in cols_norm)

                if tem_codigo and tem_produto:
                    return h

            except Exception:
                continue

        return 0

    header = achar_header(conteudo)

    df = pd.read_excel(
        BytesIO(conteudo),
        sheet_name=0,
        header=header,
    )

    df.columns = [str(c).strip() for c in df.columns]
    col_norm_map = {normaliza_coluna(c): c for c in df.columns}

    def col(*nomes):
        for nome in nomes:
            n = normaliza_coluna(nome)
            if n in col_norm_map:
                return col_norm_map[n]
        return None

    c_codigo = col("Código", "Codigo", "Cod", "Cod Produto", "Produto Codigo")
    c_produto = col("Produto", "Descrição", "Descricao", "Descrição Produto", "Descricao Produto")
    c_unid = col("Unid", "UN", "Unidade")
    c_armaz = col("Armaz", "Armazem", "Armazém")
    c_nome_2 = col("Nome 2", "Nome_2")
    c_tipo = col("Tipo")
    c_grupo = col("Grupo")
    c_grupo_desc = col("Grupo Descrição", "Grupo Descricao", "Grupo_Descricao")
    c_saldo = col("Saldo")
    c_media_3m = col("Média 3M", "Media 3M", "Média 3 meses", "Media 3 meses")
    c_media_6m = col("Média 6M", "Media 6M", "Média 6 meses", "Media 6 meses")
    c_media_9m = col("Média 9M", "Media 9M", "Média 9 meses", "Media 9 meses")
    c_maior_media = col("Maior Média", "Maior Media")
    c_giro = col("Giro Estoque", "Giro")
    c_cobertura = col("Cobertura", "Cobertura Dias")
    c_maior_media_50 = col("Maior Média +50%", "Maior Media +50%", "Maior Média 50", "Maior Media 50")
    c_saldo_gap = col(
        "Saldo - Maior Média +50%",
        "Saldo - Maior Media +50%",
        "Saldo Menos Maior Média +50%",
        "Saldo Menos Maior Media +50%",
    )

    faltando = []
    if c_codigo is None:
        faltando.append("Código")
    if c_produto is None:
        faltando.append("Produto/Descrição")

    if faltando:
        return 0, [f"Colunas não encontradas: {', '.join(faltando)}. Verifique o formato da base de consumo."]

    data_snapshot = datetime.now(timezone.utc).isoformat()

    month_regex = re.compile(r"^m[_\s\-]*(\d{1,2})[_\s\-]*(\d{4})$", re.IGNORECASE)

    def achar_coluna_mes(mes: int, ano: int):
        alvo = f"m_{mes:02d}_{ano}"
        for norm, original in col_norm_map.items():
            norm_clean = norm.lower()
            if norm_clean == alvo:
                return original

            m = month_regex.match(norm_clean)
            if m:
                mm = int(m.group(1))
                aa = int(m.group(2))
                if mm == mes and aa == ano:
                    return original

        return None

    # Detecta dinamicamente todas as colunas de consumo mensal do Excel.
    #
    # Antes a lista estava fixa até m_05_2026. Por isso, quando a base veio com
    # M_06_2026, o upload ignorou a coluna e a tela passou a mostrar Consumo mês = 0.
    #
    # Exemplo aceito:
    #   M_06_2026, M 06 2026, M-06-2026, m_06_26
    def detectar_colunas_mes_consumo() -> dict[str, str]:
        encontrados: dict[str, str] = {}

        for norm, original in col_norm_map.items():
            norm_clean = str(norm or "").strip().lower()

            match = month_regex.match(norm_clean)
            if not match:
                continue

            try:
                mes_detectado = int(match.group(1))
                ano_detectado = int(match.group(2))
            except Exception:
                continue

            if ano_detectado < 100:
                ano_detectado += 2000

            if mes_detectado < 1 or mes_detectado > 12:
                continue

            if ano_detectado < 2020 or ano_detectado > 2035:
                continue

            encontrados[f"m_{mes_detectado:02d}_{ano_detectado}"] = original

        return dict(
            sorted(
                encontrados.items(),
                key=lambda item: (int(item[0].split("_")[2]), int(item[0].split("_")[1])),
                reverse=True,
            )
        )

    colunas_mes = detectar_colunas_mes_consumo()

    records = []

    for _, row in df.iterrows():
        codigo = parse_codigo(row.get(c_codigo))

        if not codigo:
            continue

        rec = {
            "data_snapshot": data_snapshot,
            "codigo": codigo,
            "produto": str(row.get(c_produto, "") or "").strip(),
            "unid": str(row.get(c_unid, "") or "").strip() if c_unid else None,
            "armaz": str(row.get(c_armaz, "") or "").strip() if c_armaz else None,
            "nome_2": str(row.get(c_nome_2, "") or "").strip() if c_nome_2 else None,
            "tipo": str(row.get(c_tipo, "") or "").strip() if c_tipo else None,
            "grupo": str(row.get(c_grupo, "") or "").strip() if c_grupo else None,
            "grupo_descricao": str(row.get(c_grupo_desc, "") or "").strip() if c_grupo_desc else None,

            "saldo": parse_num(row.get(c_saldo)) if c_saldo else 0.0,

            "media_3m": parse_num(row.get(c_media_3m)) if c_media_3m else 0.0,
            "media_6m": parse_num(row.get(c_media_6m)) if c_media_6m else 0.0,
            "media_9m": parse_num(row.get(c_media_9m)) if c_media_9m else 0.0,
            "maior_media": parse_num(row.get(c_maior_media)) if c_maior_media else 0.0,
            "giro_estoque": parse_num(row.get(c_giro)) if c_giro else 0.0,
            "cobertura_dias": parse_num(row.get(c_cobertura)) if c_cobertura else 0.0,
            "maior_media_50": parse_num(row.get(c_maior_media_50)) if c_maior_media_50 else 0.0,
            "saldo_menos_maior_media_50": parse_num(row.get(c_saldo_gap)) if c_saldo_gap else 0.0,

            "arquivo_origem": filename,
        }

        for campo_db, coluna_excel in colunas_mes.items():
            rec[campo_db] = parse_num(row.get(coluna_excel)) if coluna_excel else 0.0

        records.append(rec)

    if not records:
        return 0, ["Nenhum registro válido encontrado na base de consumo."]

    # Substitui a posição de estoque/consumo anterior.
    # Esta base é usada como snapshot atual na tela de Aging/Cobertura.
    # Histórico por snapshot deve ficar apenas em f_estoque_saldo.
    try:
        supabase.table("f_consumo_materiais")\
            .delete()\
            .neq("id", 0)\
            .execute()
    except Exception:
        # Fallback para garantir limpeza caso o PostgREST reclame do filtro numérico.
        supabase.table("f_consumo_materiais")\
            .delete()\
            .not_.is_("codigo", "null")\
            .execute()

    erros = _chunk_insert("f_consumo_materiais", records)

    return len(records) - len(erros), erros
def process_mrp_demanda(conteudo, nome_arquivo):
    import re

    erros = []
    registros = []

    def _to_float(valor):
        if valor is None or pd.isna(valor):
            return 0.0

        try:
            if isinstance(valor, str):
                v = valor.strip()
                if not v:
                    return 0.0

                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor or 0)

        except Exception:
            return 0.0

    def _normalizar_coluna(valor):
        return (
            str(valor or "")
            .strip()
            .upper()
            .replace("Ç", "C")
            .replace("Ã", "A")
            .replace("Á", "A")
            .replace("Â", "A")
            .replace("À", "A")
            .replace("É", "E")
            .replace("Ê", "E")
            .replace("Í", "I")
            .replace("Ó", "O")
            .replace("Ô", "O")
            .replace("Õ", "O")
            .replace("Ú", "U")
        )

    def _parse_codigo(valor):
        if valor is None or pd.isna(valor):
            return None

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()

        if not texto or texto.lower() in {"nan", "none"}:
            return None

        return texto.zfill(5)

    try:
        xls = pd.ExcelFile(io.BytesIO(conteudo))

        aba = xls.sheet_names[0]

        for nome in xls.sheet_names:
            nome_upper = str(nome).upper()
            if "MRP" in nome_upper or "ANEST" in nome_upper:
                aba = nome
                break

        df = pd.read_excel(
            io.BytesIO(conteudo),
            sheet_name=aba,
            header=0,
        )

        df.columns = [str(c).strip() for c in df.columns]

        colunas_norm = {
            _normalizar_coluna(c): c
            for c in df.columns
        }

        def achar_coluna(*opcoes):
            for opcao in opcoes:
                opcao_norm = _normalizar_coluna(opcao)
                if opcao_norm in colunas_norm:
                    return colunas_norm[opcao_norm]

            return None

        col_codigo = achar_coluna(
            "Cod. Produto",
            "Cód. Produto",
            "Cod Produto",
            "Código",
            "Codigo"
        )

        col_desc = achar_coluna(
            "Descrição",
            "Descricao",
            "Desc Produto",
            "Produto"
        )

        col_tipo = achar_coluna("Tipo")
        col_un = achar_coluna("UN", "UN ", "UM", "Unid")
        col_moq = achar_coluna("MOQ", "Moq")
        col_lt = achar_coluna(
            "Lead time total (dias)",
            "Lead time total",
            "Lead Time",
            "LT"
        )

        if not col_codigo:
            return 0, ["Coluna de código do produto não encontrada na aba MRP."]

        mapa_meses = {
            "JAN": 1,
            "FEV": 2,
            "MAR": 3,
            "ABR": 4,
            "MAI": 5,
            "JUN": 6,
            "JUL": 7,
            "AGO": 8,
            "SET": 9,
            "OUT": 10,
            "NOV": 11,
            "DEZ": 12,
        }

        colunas_periodo = {}

        for coluna in df.columns:
            nome_col = _normalizar_coluna(coluna)

            match = re.match(
                r"^(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)[\-/ ]?(\d{2,4})\s+(ESTOQUE|DEMANDA|PEDIDOS|NECESSIDADE)",
                nome_col,
            )

            if not match:
                continue

            mes_txt = match.group(1)
            ano_txt = match.group(2)
            metrica = match.group(3)

            ano = int(ano_txt)
            if ano < 100:
                ano += 2000

            chave = (ano, mapa_meses[mes_txt])

            if chave not in colunas_periodo:
                colunas_periodo[chave] = {}

            colunas_periodo[chave][metrica] = coluna

        if not colunas_periodo:
            return 0, ["Nenhuma coluna mensal encontrada no formato 'mai-26 ESTOQUE', 'mai-26 DEMANDA', etc."]

        data_rodada = datetime.now().isoformat()

        for _, row in df.iterrows():
            codigo = _parse_codigo(row.get(col_codigo))

            if not codigo:
                continue

            descricao = str(row.get(col_desc) or "").strip() if col_desc else ""
            tipo = str(row.get(col_tipo) or "").strip() if col_tipo else ""
            un = str(row.get(col_un) or "").strip() if col_un else ""

            moq = _to_float(row.get(col_moq)) if col_moq else 0.0
            lt = _to_float(row.get(col_lt)) if col_lt else 0.0

            for (ano, mes), cols in colunas_periodo.items():
                estoque = _to_float(row.get(cols.get("ESTOQUE")))
                demanda = _to_float(row.get(cols.get("DEMANDA")))
                pedidos = _to_float(row.get(cols.get("PEDIDOS")))
                necessidade = _to_float(row.get(cols.get("NECESSIDADE")))

                if estoque == 0 and demanda == 0 and pedidos == 0 and necessidade == 0:
                    continue

                registros.append({
                    "data_rodada": data_rodada,
                    "arquivo_origem": nome_arquivo,

                    "codigo": codigo,
                    "descricao": descricao,
                    "tipo": tipo,
                    "un": un,

                    "moq": moq,
                    "lead_time_total": lt,

                    "mes": mes,
                    "ano": ano,
                    "mes_label": f"{str(mes).zfill(2)}/{ano}",

                    "estoque_mrp": estoque,
                    "demanda_mrp": demanda,
                    "pedidos_mrp": pedidos,
                    "necessidade_mrp": necessidade,
                })

        if not registros:
            return 0, ["Nenhum registro válido encontrado na aba MRP."]

        erros = _chunk_insert("f_mrp_demanda", registros)

        return len(registros) - len(erros), erros

    except Exception as e:
        raise Exception(f"Erro processando MRP Demanda: {str(e)}")

# ─── f_desvios_lotes ─────────────────────────────────────────────────────────

def process_desvios_lotes(conteudo: bytes, filename: str) -> Tuple[int, list]:
    from io import BytesIO
    from datetime import date as date_cls, datetime as datetime_cls

    snapshot_id = str(uuid.uuid4())

    def normaliza_coluna(valor: str) -> str:
        return (
            str(valor or "").strip().upper()
            .replace(" ", "").replace(".", "").replace("-", "")
            .replace("_", "").replace("/", "")
            .replace("Ç", "C").replace("Ã", "A").replace("Á", "A")
            .replace("Â", "A").replace("À", "A").replace("É", "E")
            .replace("Ê", "E").replace("Í", "I").replace("Ó", "O")
            .replace("Ô", "O").replace("Õ", "O").replace("Ú", "U")
        )

    def parse_text(valor):
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        texto = str(valor).strip()

        if texto.endswith(".0"):
            texto = texto[:-2]

        if not texto or texto.lower() in {"nan", "none", "nat", "null"}:
            return None

        return texto.strip()

    def parse_num(valor):
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        try:
            if isinstance(valor, str):
                v = valor.strip()

                if not v:
                    return None

                if "," in v:
                    v = v.replace(".", "").replace(",", ".")

                return float(v)

            return float(valor)

        except Exception:
            return None

    def parse_date(valor):
        if valor is None:
            return None

        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass

        try:
            if isinstance(valor, datetime_cls):
                return valor.date().isoformat()

            if isinstance(valor, date_cls):
                return valor.isoformat()

            dt = pd.to_datetime(valor, dayfirst=True, errors="coerce")

            if pd.isna(dt):
                return None

            return dt.date().isoformat()

        except Exception:
            return None

    def normaliza_lote(valor) -> str:
        lote = str(valor or "").strip().upper().replace(" ", "")

        if lote.endswith(".0"):
            lote = lote[:-2]

        # Alguns lotes PPS vêm como número puro, por exemplo 086626.
        # Se o Excel transformar 086626 em 86626, recompõe com zero à esquerda.
        if lote.isdigit() and 4 <= len(lote) <= 6:
            lote = lote.zfill(6)

        return lote

    def normaliza_serial(valor) -> str:
        serial = str(valor or "").strip()

        if serial.endswith(".0"):
            serial = serial[:-2]

        return serial

    def normaliza_chave_texto(valor) -> str:
        texto = str(valor or "").strip().upper()
        texto = re.sub(r"\s+", " ", texto)
        return texto

    def explode_lotes(valor) -> list[str]:
        texto = parse_text(valor)

        if not texto:
            return []

        texto = texto.upper()
        encontrados: list[str] = []

        # Lote padrão DFL: 2604F2026, 2605C2035 etc.
        encontrados.extend(re.findall(r"\b\d{4}[A-Z]\d{4}\b", texto))

        # Lote numérico de PPS: 086626, 086726 etc.
        # Também aceita 5 dígitos porque o Excel pode remover o zero à esquerda.
        encontrados.extend(re.findall(r"\b\d{5,8}\b", texto))

        # Caso manual eventual: F2026. O ideal é subir o lote completo
        # para conciliar com a programação, mas gravamos para não perder o desvio.
        encontrados.extend(re.findall(r"\b[A-Z]\d{4}\b", texto))

        lotes = []
        vistos = set()

        for lote in encontrados:
            lote = normaliza_lote(lote)

            if lote and lote not in vistos:
                vistos.add(lote)
                lotes.append(lote)

        return lotes

    def achar_header(conteudo_bytes: bytes) -> int:
        for h in range(0, 12):
            try:
                df_test = pd.read_excel(
                    BytesIO(conteudo_bytes),
                    sheet_name=0,
                    header=h,
                    nrows=0,
                )

                cols_norm = [normaliza_coluna(c) for c in df_test.columns]

                tem_lote = "LOTE" in cols_norm or "LOTES" in cols_norm
                tem_estado = "ESTADO" in cols_norm or "STATUS" in cols_norm or "STATE" in cols_norm
                tem_serial = "SERIAL" in cols_norm or "NUMERO" in cols_norm or "N" in cols_norm or "NC" in cols_norm
                tem_titulo = any(c in cols_norm for c in ["TITLE", "TITULO", "DESCRICAO", "OCORRENCIA", "ASSUNTO", "MOTIVO"])
                tem_destino = any(c in cols_norm for c in ["DESTINO", "DESTINOPRODUTOINSUMO", "DESTINOPRODUTO", "TRATATIVA"])

                if tem_lote and (tem_estado or tem_serial or tem_titulo or tem_destino):
                    return h

            except Exception:
                continue

        return 0

    def select_all(
        table: str,
        filters: dict | None = None,
        order_col: str | None = None,
        desc: bool = False,
        page_size: int = 1000,
    ):
        rows = []
        start = 0

        while True:
            q = supabase.table(table).select("*")

            if filters:
                for col, val in filters.items():
                    q = q.eq(col, val)

            if order_col:
                q = q.order(order_col, desc=desc)

            resp = q.range(start, start + page_size - 1).execute()
            batch = resp.data or []
            rows.extend(batch)

            if len(batch) < page_size:
                break

            start += page_size

        return rows

    try:
        header = achar_header(conteudo)

        df = pd.read_excel(
            BytesIO(conteudo),
            sheet_name=0,
            header=header,
        )

    except Exception as e:
        return 0, [f"Erro lendo arquivo de desvios: {str(e)[:200]}"]

    if df.empty:
        return 0, ["Arquivo de desvios vazio."]

    df.columns = [str(c).strip() for c in df.columns]
    col_norm_map = {normaliza_coluna(c): c for c in df.columns}

    def col(*nomes):
        for nome in nomes:
            n = normaliza_coluna(nome)

            if n in col_norm_map:
                return col_norm_map[n]

        return None

    c_serial = col("Serial", "Nº", "N°", "Numero", "Número", "ID", "NC", "Desvio")

    c_titulo = col(
        "Title", "Título", "Titulo", "Descrição", "Descricao",
        "Ocorrência", "Ocorrencia", "Assunto", "Motivo"
    )

    c_setor = col("Setor", "Área", "Area", "Departamento", "Sector")

    c_data_criacao = col(
        "Data de criação", "Data Criacao", "Data de abertura",
        "Data Abertura", "Criado em", "dt_create", "Data Criacao"
    )

    c_estado = col("Estado", "Status", "Situação", "Situacao", "State")

    c_dias_desvio = col(
        "Dias de Desvio",
        "Dias em Desvio",
        "Dias Desvio",
        "Dias",
        "Aging"
    )

    c_lote = col("Lote", "Lotes")

    c_destino = col(
        "Destino Produto/Insumo",
        "Destino Produto",
        "Destino",
        "Tratativa",
    )

    if c_lote is None:
        return 0, ["Coluna de lote não encontrada. Use uma coluna chamada Lote ou Lotes, mesmo que algumas linhas fiquem vazias."]

    registros = []

    for _, row in df.iterrows():
        lote_original = parse_text(row.get(c_lote))
        lotes = explode_lotes(lote_original)

        serial = normaliza_serial(
            parse_text(row.get(c_serial)) if c_serial else None
        )

        titulo = parse_text(row.get(c_titulo)) if c_titulo else None
        setor = parse_text(row.get(c_setor)) if c_setor else None
        data_criacao = parse_date(row.get(c_data_criacao)) if c_data_criacao else None
        estado = parse_text(row.get(c_estado)) if c_estado else None
        dias_desvio = parse_num(row.get(c_dias_desvio)) if c_dias_desvio else None
        destino = parse_text(row.get(c_destino)) if c_destino else None

        tem_algum_dado = any([
            serial,
            titulo,
            setor,
            data_criacao,
            estado,
            dias_desvio is not None,
            destino,
            lote_original,
        ])

        if not tem_algum_dado:
            continue

        # Antes, linhas sem lote eram descartadas. Agora elas entram com lote nulo,
        # para aparecerem na tela de Desvios atuais, como no caso NC 2026 164.
        if not lotes:
            lotes = [None]

        for lote in lotes:
            registros.append({
                "serial": serial or None,
                "titulo": titulo,
                "setor": setor,
                "data_criacao": data_criacao,
                "estado": estado,
                "dias_desvio": dias_desvio,
                "lote": lote,
                "lote_original": lote_original,
                "destino": destino,
                "arquivo_origem": filename,
            })

    if not registros:
        return 0, ["Nenhum desvio válido encontrado. Verifique se há lote, serial, título ou destino preenchidos."]

    # SNAPSHOT ANTERIOR = ÚLTIMO UPLOAD
    snapshots = select_all(
        "desvios_snapshots",
        order_col="data_upload",
        desc=True,
    )

    ultimo_snapshot_id = None

    if snapshots:
        ultimo_snapshot_id = snapshots[0].get("snapshot_id")

    anterior = []

    if ultimo_snapshot_id:
        anterior = select_all(
            "desvios_snapshots",
            filters={"snapshot_id": ultimo_snapshot_id},
        )

    def chave_reg(r):
        serial_norm = normaliza_serial(r.get("serial"))
        lote_norm = normaliza_lote(r.get("lote")) if r.get("lote") else ""
        titulo_norm = normaliza_chave_texto(r.get("titulo"))
        destino_norm = normaliza_chave_texto(r.get("destino"))

        # Para registros com lote, a chave principal segue sendo serial + lote.
        # Para registros sem lote, usa também título/destino para evitar colisão.
        return (
            serial_norm,
            lote_norm,
            titulo_norm if not lote_norm else "",
            destino_norm if not lote_norm else "",
        )

    mapa_ant = {chave_reg(r): r for r in anterior}
    mapa_novo = {chave_reg(r): r for r in registros}

    eventos = []

    # NOVOS LOTES / DESVIOS SEM LOTE
    for chave, novo in mapa_novo.items():
        if chave not in mapa_ant:
            lote_desc = novo.get("lote") or "sem lote"
            serial_desc = novo.get("serial") or "sem serial"
            eventos.append({
                "snapshot_id": snapshot_id,
                "tipo_evento": "NOVO_LOTE" if novo.get("lote") else "NOVO_DESVIO_SEM_LOTE",
                "serial": novo.get("serial"),
                "lote": novo.get("lote"),
                "descricao": f"Registro {lote_desc} adicionado no desvio {serial_desc}"
            })

    # LOTES / DESVIOS REMOVIDOS
    for chave, antigo in mapa_ant.items():
        if chave not in mapa_novo:
            lote_desc = antigo.get("lote") or "sem lote"
            serial_desc = antigo.get("serial") or "sem serial"
            eventos.append({
                "snapshot_id": snapshot_id,
                "tipo_evento": "LOTE_REMOVIDO" if antigo.get("lote") else "DESVIO_SEM_LOTE_REMOVIDO",
                "serial": antigo.get("serial"),
                "lote": antigo.get("lote"),
                "descricao": f"Registro {lote_desc} removido do desvio {serial_desc}"
            })

    # NOVOS DESVIOS
    seriais_ant = {r.get("serial") for r in anterior if r.get("serial")}
    seriais_novo = {r.get("serial") for r in registros if r.get("serial")}

    for serial in sorted(seriais_novo - seriais_ant):
        eventos.append({
            "snapshot_id": snapshot_id,
            "tipo_evento": "NOVO_DESVIO",
            "serial": serial,
            "lote": None,
            "descricao": f"Novo desvio {serial}"
        })

    # LIMPA posição atual inteira.
    # Antes limpava apenas onde lote não era nulo; isso mantinha resíduos de desvios sem lote.
    try:
        _limpar_tabela("f_desvios_lotes")
    except Exception:
        supabase.table("f_desvios_lotes")\
            .delete()\
            .neq("id", "00000000-0000-0000-0000-000000000000")\
            .execute()

    erros = _chunk_insert("f_desvios_lotes", registros)

    # SALVA snapshot atual
    snapshot_records = []

    for r in registros:
        snapshot_records.append({
            "snapshot_id": snapshot_id,
            "arquivo_origem": filename,
            "serial": r.get("serial"),
            "titulo": r.get("titulo"),
            "setor": r.get("setor"),
            "estado": r.get("estado"),
            "dias_desvio": r.get("dias_desvio"),
            "lote": r.get("lote"),
            "lote_original": r.get("lote_original"),
            "destino": r.get("destino"),
        })

    erros += _chunk_insert(
        "desvios_snapshots",
        snapshot_records,
    )

    # LIMPA eventos antigos
    try:
        supabase.table("desvios_eventos")\
            .delete()\
            .not_.is_("snapshot_id", "null")\
            .execute()
    except Exception:
        pass

    if eventos:
        erros += _chunk_insert(
            "desvios_eventos",
            eventos,
        )

    return len(registros) - len(erros), erros





# ─── d_clientes ───────────────────────────────────────────────────────────────

def process_d_clientes(df: pd.DataFrame) -> Tuple[int, list]:
    """
    Processa a dimensão dClientes do Protheus para a tabela d_clientes.

    Layout aceito:
      Codigo | Loja | Nome | Fisica/Jurid | N Fantasia | Tipo | Estado |
      Municipio | Regiao | Desc.Região | CNPJ/CPF

    Regras:
      - substitui a dimensão inteira a cada upload;
      - código e loja são preservados com zero à esquerda;
      - o join principal do faturamento usa f_sd2_saidas.cliente = d_clientes.codigo.
    """
    import unicodedata

    df.columns = [str(c).strip() for c in df.columns]

    def normaliza_coluna(valor: str) -> str:
        texto = str(valor or "").strip().upper()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = (
            texto
            .replace("Ç", "C")
            .replace(".", "")
            .replace("-", "")
            .replace("_", "")
            .replace("/", "")
            .replace(" ", "")
        )
        return texto

    col_norm_map = {normaliza_coluna(c): c for c in df.columns}

    def col(*nomes):
        for nome in nomes:
            chave = normaliza_coluna(nome)
            if chave in col_norm_map:
                return col_norm_map[chave]
        return None

    def parse_text(valor) -> str:
        if valor is None:
            return ""
        try:
            if pd.isna(valor):
                return ""
        except Exception:
            pass

        texto = str(valor).strip()
        if texto.endswith(".0"):
            texto = texto[:-2]
        if texto.lower() in {"nan", "none", "nat", "null"}:
            return ""
        return texto.strip()

    def parse_codigo(valor, zfill: int = 6) -> str:
        texto = parse_text(valor)
        if not texto:
            return ""

        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
        except Exception:
            pass

        texto = texto.strip()
        if texto.endswith(".0"):
            texto = texto[:-2]

        return texto.zfill(zfill) if texto.isdigit() else texto

    c_codigo = col("Codigo", "Código", "Cod Cliente", "Cliente", "Codigo Cliente", "Código Cliente")
    c_loja = col("Loja")
    c_nome = col("Nome", "Razao Social", "Razão Social")
    c_fantasia = col("N Fantasia", "Nome Fantasia", "Fantasia", "N Fantasia")
    c_pessoa = col("Fisica/Jurid", "Física/Juríd", "Pessoa")
    c_tipo = col("Tipo", "Tipo Cliente")
    c_estado = col("Estado", "UF")
    c_municipio = col("Municipio", "Município")
    c_regiao = col("Regiao", "Região")
    c_desc_regiao = col("Desc.Região", "Desc.Regiao", "Descricao Regiao", "Descrição Região")
    c_cnpj = col("CNPJ/CPF", "CNPJ", "CPF")

    if c_codigo is None:
        return 0, ["Coluna Codigo/Código não encontrada na dClientes."]

    records_por_chave: dict[tuple[str, str], dict] = {}

    for _, row in df.iterrows():
        codigo = parse_codigo(row.get(c_codigo), zfill=6)
        if not codigo:
            continue

        loja = parse_codigo(row.get(c_loja), zfill=2) if c_loja is not None else ""

        chave = (codigo, loja)

        records_por_chave[chave] = {
            "codigo": codigo,
            "loja": loja,
            "nome": parse_text(row.get(c_nome)) if c_nome is not None else "",
            "nome_fantasia": parse_text(row.get(c_fantasia)) if c_fantasia is not None else "",
            "pessoa": parse_text(row.get(c_pessoa)) if c_pessoa is not None else "",
            "tipo_cliente": parse_text(row.get(c_tipo)) if c_tipo is not None else "",
            "estado": parse_text(row.get(c_estado)).upper() if c_estado is not None else "",
            "municipio": parse_text(row.get(c_municipio)) if c_municipio is not None else "",
            "regiao": parse_text(row.get(c_regiao)) if c_regiao is not None else "",
            "desc_regiao": parse_text(row.get(c_desc_regiao)) if c_desc_regiao is not None else "",
            "cnpj_cpf": parse_text(row.get(c_cnpj)) if c_cnpj is not None else "",
        }

    records = list(records_por_chave.values())

    if not records:
        return 0, ["Nenhum cliente válido encontrado na dClientes."]

    # d_clientes usa id BIGSERIAL no SQL.
    # Por isso NÃO pode usar _limpar_tabela(), porque o fallback genérico tenta
    # comparar id com UUID texto e gera erro:
    # invalid input syntax for type bigint: "00000000-0000-0000-0000-000000000000"
    try:
        supabase.table("d_clientes")\
            .delete()\
            .not_.is_("codigo", "null")\
            .execute()
    except Exception as e:
        return 0, [f"Erro ao limpar d_clientes: {str(e)[:300]}"]

    erros = _chunk_insert("d_clientes", records)

    return len(records) - len(erros), erros


# ─── MATA010 / Cadastro Protheus leve ────────────────────────────────────────

def process_mata010(conteudo, filename: str = "mata010.xlsx") -> Tuple[int, list]:
    """
    Processa a MATA010 do Protheus em layout leve:
      Código | Descrição | Tipo | Unidade

    Uso no projeto:
      - atualizar somente informações cadastrais básicas na d_produtos;
      - apoiar a regra PA x PI da Gestão de Estoque;
      - preservar as classificações gerenciais já existentes.

    Importante:
      - NÃO limpa d_produtos;
      - NÃO altera grupo, mercado, macro_negocio, tipo_negocio, status_portfolio,
        transferencia_bravi, fornecedor_terceiro, modelo_fornecimento ou grupo_gerencial
        quando o produto já existe;
      - para produtos novos, cria cadastro mínimo seguro.
    """
    from io import BytesIO
    import unicodedata

    def normaliza_coluna(valor: str) -> str:
        texto = str(valor or "").strip().lower()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = (
            texto
            .replace("ç", "c")
            .replace("º", "")
            .replace("ª", "")
            .replace("?", "")
            .replace("/", "_")
            .replace("-", "_")
            .replace(".", "")
            .replace("(", "")
            .replace(")", "")
            .replace(" ", "_")
        )
        while "__" in texto:
            texto = texto.replace("__", "_")
        return texto.strip("_")

    def parse_text(valor) -> str | None:
        if valor is None:
            return None
        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass
        texto = str(valor).strip()
        if texto.endswith(".0"):
            texto = texto[:-2]
        if not texto or texto.lower() in {"nan", "none", "nat", "null"}:
            return None
        return texto

    def parse_codigo(valor) -> str | None:
        if valor is None:
            return None
        try:
            if pd.isna(valor):
                return None
        except Exception:
            pass
        try:
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                texto = str(int(float(valor)))
            else:
                texto = str(valor).strip()
                if texto.endswith(".0"):
                    texto = texto[:-2]
        except Exception:
            texto = str(valor).strip()
        texto = texto.strip()
        if not texto or texto.lower() in {"nan", "none", "nat", "null"}:
            return None
        return texto.zfill(5) if texto.isdigit() else texto

    def parse_tipo(valor) -> str | None:
        texto = parse_text(valor)
        if not texto:
            return None
        return texto.strip().upper()

    def ler_df_mata010(conteudo_base) -> pd.DataFrame:
        if isinstance(conteudo_base, pd.DataFrame):
            return conteudo_base.copy()

        # Procura cabeçalho nas primeiras linhas, porque o Protheus pode exportar
        # título/filtros antes da tabela.
        for h in range(0, 15):
            try:
                df_test = pd.read_excel(BytesIO(conteudo_base), sheet_name=0, header=h, nrows=0)
                cols_norm = [normaliza_coluna(c) for c in df_test.columns]
                tem_codigo = any(c in {"codigo", "cod", "cod_produto", "codigo_produto", "produto_codigo"} for c in cols_norm)
                tem_descricao = any(c in {"descricao", "desc", "desc_produto", "descricao_produto", "produto"} for c in cols_norm)
                tem_tipo = any(c in {"tipo", "tp", "tipo_produto", "tipo_produto_erp"} for c in cols_norm)
                if tem_codigo and tem_descricao and tem_tipo:
                    return pd.read_excel(BytesIO(conteudo_base), sheet_name=0, header=h)
            except Exception:
                continue

        return pd.read_excel(BytesIO(conteudo_base), sheet_name=0, header=0)

    def select_all_d_produtos(page_size: int = 1000) -> dict[str, dict]:
        existentes: dict[str, dict] = {}
        page = 0
        while True:
            try:
                res = (
                    supabase.table("d_produtos")
                    .select("*")
                    .range(page * page_size, ((page + 1) * page_size) - 1)
                    .execute()
                )
                data = res.data or []
            except Exception:
                data = []

            for row in data:
                codigo = parse_codigo(row.get("cod_produto"))
                if codigo:
                    existentes[codigo] = row

            if len(data) < page_size:
                break
            page += 1

        return existentes

    def valor_existente(atual: dict, campo: str, default=None):
        valor = atual.get(campo)
        if valor is None:
            return default
        if isinstance(valor, str) and not valor.strip():
            return default
        return valor

    try:
        df = ler_df_mata010(conteudo)
    except Exception as e:
        return 0, [f"Erro ao ler MATA010: {str(e)[:250]}"]

    if df.empty:
        return 0, ["Arquivo MATA010 vazio."]

    df.columns = [str(c).strip() for c in df.columns]
    col_norm_map = {normaliza_coluna(c): c for c in df.columns}

    def col(*nomes: str):
        for nome in nomes:
            nome_norm = normaliza_coluna(nome)
            if nome_norm in col_norm_map:
                return col_norm_map[nome_norm]
        return None

    c_codigo = col("codigo", "código", "cod", "cod_produto", "codigo_produto", "código produto", "produto_codigo")
    c_descricao = col("descricao", "descrição", "desc", "produto", "desc_produto", "descricao_produto", "descrição produto")
    c_tipo = col("tipo", "tp", "tipo_produto", "tipo produto", "tipo_produto_erp")
    c_unidade = col("unidade", "unid", "un", "um")

    faltando = []
    if c_codigo is None:
        faltando.append("Código")
    if c_descricao is None:
        faltando.append("Descrição")
    if c_tipo is None:
        faltando.append("Tipo")

    if faltando:
        return 0, [
            "Colunas não encontradas na MATA010: " + ", ".join(faltando) +
            ". Layout esperado: Código, Descrição, Tipo e Unidade."
        ]

    existentes = select_all_d_produtos()
    records_por_codigo: dict[str, dict] = {}
    duplicados: set[str] = set()

    for _, row in df.iterrows():
        codigo = parse_codigo(row.get(c_codigo))
        if not codigo:
            continue

        descricao = parse_text(row.get(c_descricao)) if c_descricao else None
        tipo = parse_tipo(row.get(c_tipo)) if c_tipo else None
        unidade = parse_text(row.get(c_unidade)) if c_unidade else None

        atual = existentes.get(codigo, {})

        record = {
            "cod_produto": codigo,
            "desc_produto": descricao or valor_existente(atual, "desc_produto", ""),
            "grupo": valor_existente(atual, "grupo", ""),
            "mercado": valor_existente(atual, "mercado", "NACIONAL"),
            "tipo_produto_erp": tipo or valor_existente(atual, "tipo_produto_erp"),
            "familia": valor_existente(atual, "familia"),
            "segmento": valor_existente(atual, "segmento"),
            "abc_ytm": valor_existente(atual, "abc_ytm"),
            "linha": valor_existente(atual, "linha"),
            "status_original": valor_existente(atual, "status_original"),
            "macro_negocio": valor_existente(atual, "macro_negocio"),
            "tipo_negocio": valor_existente(atual, "tipo_negocio"),
            "status_portfolio": valor_existente(atual, "status_portfolio", "Ativo"),
            "transferencia_bravi": valor_existente(atual, "transferencia_bravi", "Não"),
            "fornecedor_terceiro": valor_existente(atual, "fornecedor_terceiro"),
            "modelo_fornecimento": valor_existente(atual, "modelo_fornecimento"),
            "grupo_gerencial": valor_existente(atual, "grupo_gerencial", "A classificar"),
            "incluir_overview_anestesicos": valor_existente(atual, "incluir_overview_anestesicos", False),
            "ativo_analise": valor_existente(atual, "ativo_analise", True),
            "observacao": valor_existente(atual, "observacao"),
            "concatenado_produto": f"{codigo} - {descricao}" if descricao else valor_existente(atual, "concatenado_produto", codigo),
        }

        # A MATA010 leve não tem coluna própria na d_produtos para Unidade.
        # Para não quebrar schema, preservamos unidade apenas no retorno/debug via observação
        # quando for produto novo e não houver observação anterior.
        if unidade and not atual and not record.get("observacao"):
            record["observacao"] = f"Unidade MATA010: {unidade}"

        if codigo in records_por_codigo:
            duplicados.add(codigo)

        records_por_codigo[codigo] = record

    records = list(records_por_codigo.values())

    if not records:
        return 0, ["Nenhum produto válido encontrado na MATA010."]

    erros: list[str] = []
    if duplicados:
        erros.append(
            "Códigos duplicados na MATA010; mantida a última ocorrência do arquivo: "
            + ", ".join(sorted(duplicados)[:20])
        )

    for i in range(0, len(records), 500):
        chunk = records[i : i + 500]
        try:
            supabase.table("d_produtos").upsert(chunk, on_conflict="cod_produto").execute()
        except Exception:
            for rec in chunk:
                try:
                    supabase.table("d_produtos").upsert(rec, on_conflict="cod_produto").execute()
                except Exception as e:
                    erros.append(
                        f"Erro ao atualizar MATA010 produto {rec.get('cod_produto')} - "
                        f"{rec.get('desc_produto')}: {str(e)[:300]}"
                    )

    return len(records) - max(0, len([e for e in erros if e.startswith("Erro ao atualizar MATA010 produto")])), erros


# Aliases para facilitar o mapeamento no router de upload.
def process_mata010_file(conteudo: bytes, filename: str = "mata010.xlsx") -> Tuple[int, list]:
    return process_mata010(conteudo, filename)


def process_d_mata010(conteudo, filename: str = "mata010.xlsx") -> Tuple[int, list]:
    return process_mata010(conteudo, filename)


def process_cadastro_mata010(conteudo, filename: str = "mata010.xlsx") -> Tuple[int, list]:
    return process_mata010(conteudo, filename)