from fastapi import APIRouter, HTTPException, UploadFile, File, BackgroundTasks
from app.database import supabase
from pydantic import BaseModel
from typing import Optional, Dict, Any, List, Tuple
from datetime import date, datetime, timedelta
from io import BytesIO
import pandas as pd
import unicodedata
import math
import re
from openpyxl import load_workbook

router = APIRouter(prefix="/mrp", tags=["mrp"])


# ─── Schemas ──────────────────────────────────────────────────────────────────

class RodadaCreate(BaseModel):
    nome: str
    mes: int
    ano: int
    versao: int
    observacao: Optional[str] = None


class RodadaCopiarCreate(BaseModel):
    nome: Optional[str] = None
    mes: Optional[int] = None
    ano: Optional[int] = None
    versao: Optional[int] = None
    observacao: Optional[str] = None


class EtapaCreate(BaseModel):
    rodada_id: str
    lote: Optional[str] = None
    op: Optional[str] = None
    codigo_produto: Optional[str] = None
    descricao_produto: Optional[str] = None
    etapa: str
    recurso: str
    linha_origem: Optional[str] = None
    data_inicio: Optional[date] = None
    data_fim: Optional[date] = None
    data_pa: Optional[date] = None
    qtd_planejada: Optional[float] = 0
    duracao_horas: Optional[float] = 0
    sequencia: Optional[int] = None
    status: Optional[str] = "planejada"
    origem: Optional[str] = None
    observacao: Optional[str] = None
    embalado: Optional[str] = None
    un_hora: Optional[float] = None
    mes_producao: Optional[int] = None
    ano_producao: Optional[int] = None
    mes_liberacao: Optional[int] = None
    ano_liberacao: Optional[int] = None
    mes_lib_manual: Optional[bool] = False


class EtapaUpdate(BaseModel):
    lote: Optional[str] = None
    op: Optional[str] = None
    codigo_produto: Optional[str] = None
    descricao_produto: Optional[str] = None
    etapa: Optional[str] = None
    recurso: Optional[str] = None
    linha_origem: Optional[str] = None
    data_inicio: Optional[date] = None
    data_fim: Optional[date] = None
    data_pa: Optional[date] = None
    qtd_planejada: Optional[float] = None
    duracao_horas: Optional[float] = None
    sequencia: Optional[int] = None
    status: Optional[str] = None
    origem: Optional[str] = None
    observacao: Optional[str] = None
    embalado: Optional[str] = None
    un_hora: Optional[float] = None
    mes_producao: Optional[int] = None
    ano_producao: Optional[int] = None
    mes_liberacao: Optional[int] = None
    ano_liberacao: Optional[int] = None
    mes_lib_manual: Optional[bool] = None


MESES_PT = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}


# ─── Helpers gerais ───────────────────────────────────────────────────────────

def _date_to_str(value):
    return str(value) if value else None


def _str_to_date(value):
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except Exception:
        return None


def _limpar_nan(value):
    if value is None:
        return None
    try:
        if isinstance(value, float) and math.isnan(value):
            return None
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def _to_str(value):
    value = _limpar_nan(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_float(value, default=0):
    value = _limpar_nan(value)
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        texto = str(value).strip()
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
        return float(texto)
    except Exception:
        return default


def _to_int(value):
    value = _limpar_nan(value)
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        texto = str(value).strip()
        if not texto:
            return None
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
            return int(float(texto))
        if "." in texto:
            partes = texto.split(".")
            if len(partes) == 2 and len(partes[1]) == 3 and partes[0].isdigit() and partes[1].isdigit():
                texto = "".join(partes)
                return int(texto)
            return int(float(texto))
        return int(float(texto))
    except Exception:
        return None


def _normalizar_coluna(col):
    col = str(col).strip().upper()
    col = unicodedata.normalize("NFKD", col)
    col = "".join(c for c in col if not unicodedata.combining(c))
    col = col.replace("\n", " ").replace(".", "")
    return " ".join(col.split())


def _normalizar_texto(value):
    if value is None:
        return ""
    texto = str(value).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return " ".join(texto.split())


def _normalizar_lote_key(value):
    texto = _normalizar_texto(value)
    texto = re.sub(r"\.0$", "", texto)
    texto = re.sub(r"[^A-Z0-9]", "", texto)
    return texto


def _extrair_lote_do_texto(value):
    texto = _normalizar_lote_key(value)
    if not texto:
        return ""
    match = re.search(r"\d{4}[A-Z]\d{3,5}", texto)
    if match:
        return match.group(0)
    return texto


def _lote_core_key(value):
    texto = _normalizar_lote_key(value)
    if not texto:
        return ""
    match = re.search(r"[A-Z][12]\d{3,5}", texto)
    if match:
        return match.group(0)
    return ""


def _buscar_coluna(df, possibilidades):
    colunas_normalizadas = {_normalizar_coluna(col): col for col in df.columns}
    for possibilidade in possibilidades:
        chave = _normalizar_coluna(possibilidade)
        if chave in colunas_normalizadas:
            return colunas_normalizadas[chave]
    for col_norm, col_original in colunas_normalizadas.items():
        for possibilidade in possibilidades:
            if _normalizar_coluna(possibilidade) in col_norm:
                return col_original
    return None


def _mapear_mes_lib_manual_por_linha(excel_bytes, sheet_name):
    mapa = {}
    wb = None
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=False, read_only=True)
        if sheet_name not in wb.sheetnames:
            return mapa
        ws = wb[sheet_name]
        header_row = 6
        col_mes_lib = None
        # Em read_only, iteramos pelas linhas
        rows = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))
        if rows:
            for cell in rows[0]:
                header_norm = _normalizar_coluna(cell.value)
                if header_norm in [
                    _normalizar_coluna("MÊS LIBERAÇÃO"),
                    _normalizar_coluna("MES LIBERAÇÃO"),
                    _normalizar_coluna("MES LIBERACAO"),
                ]:
                    col_mes_lib = cell.column
                    break
        if not col_mes_lib:
            return mapa
        for row in ws.iter_rows(min_row=header_row + 1, min_col=col_mes_lib, max_col=col_mes_lib, values_only=False):
            cell = row[0]
            valor = cell.value
            if valor is None or str(valor).strip() == "":
                continue
            eh_formula = (isinstance(valor, str) and valor.strip().startswith("="))
            mapa[cell.row] = not eh_formula
    except Exception:
        pass
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
    return mapa



def _texto_comentario_cell(cell) -> str | None:
    """
    Lê comentário real de célula do Excel.
    openpyxl só preserva comentário quando o arquivo é aberto sem read_only.
    """
    try:
        comentario = getattr(cell, "comment", None)
        if not comentario:
            return None

        texto = str(getattr(comentario, "text", "") or "").strip()
        if not texto:
            return None

        texto = re.sub(r"\s+", " ", texto)
        return texto[:800]
    except Exception:
        return None


def _rotulo_coluna_gantt(ws, col_idx: int) -> str:
    """
    Identifica se a coluna é um campo fixo ou uma coluna de dia do Gantt.
    No layout atual:
      - linha 5 = datas do calendário;
      - linha 6 = cabeçalhos/capacidade;
      - dados começam na linha 7.
    """
    try:
        data_header = _parse_data(ws.cell(row=5, column=col_idx).value)
        if data_header:
            return data_header
    except Exception:
        pass

    try:
        header = _to_str(ws.cell(row=6, column=col_idx).value)
        if header:
            return header
    except Exception:
        pass

    return f"Coluna {col_idx}"


def _mapear_colunas_fixas_gantt(ws) -> dict[str, int]:
    mapa: dict[str, int] = {}

    try:
        for cell in ws[6]:
            header_norm = _normalizar_coluna(cell.value)

            if header_norm == _normalizar_coluna("LOTE"):
                mapa["lote"] = cell.column
            elif header_norm in {_normalizar_coluna("CÓDIGO"), _normalizar_coluna("CODIGO")}:
                mapa["codigo"] = cell.column
            elif header_norm == _normalizar_coluna("PRODUTO"):
                mapa["produto"] = cell.column
    except Exception:
        pass

    return mapa


def _mapear_comentarios_gantt_por_linha(excel_bytes, sheet_name) -> dict[int, list[str]]:
    """
    v31 performance:
    Não varre mais a planilha inteira para anexar comentários na observação
    da etapa. Os comentários oficiais do Gantt ficam em f_mrp_calendario_dia.
    """
    return {}


def _ler_calendario_dia_gantt(excel_bytes, sheet_name, recurso, rodada_id, wb_full=None):
    """
    v31 leve:
    Lê o calendário diário e comentários reais do Gantt em uma única varredura
    por aba, reutilizando o workbook completo quando possível.

    Grava em f_mrp_calendario_dia:
      - capacidade do dia;
      - horas indisponíveis;
      - comentários das células do Gantt agrupados por data/recurso.
    """
    registros: list[dict[str, Any]] = []
    wb = None
    fechar_wb = False

    try:
        if wb_full is not None:
            wb = wb_full
        else:
            wb = load_workbook(
                BytesIO(excel_bytes),
                data_only=False,
                read_only=False,
                keep_links=False,
            )
            fechar_wb = True

        if sheet_name not in wb.sheetnames:
            return registros

        ws = wb[sheet_name]

        # 1) Mapeia as colunas que representam dias pela linha 5.
        colunas_dia: dict[int, dict[str, Any]] = {}

        for col_idx in range(1, ws.max_column + 1):
            data_ref = _parse_data(ws.cell(row=5, column=col_idx).value)
            if not data_ref:
                continue

            capacidade = _to_float(ws.cell(row=6, column=col_idx).value, 0)

            colunas_dia[col_idx] = {
                "data": data_ref,
                "capacidade": capacidade,
                "comentarios": [],
            }

        if not colunas_dia:
            return registros

        colunas_fixas = _mapear_colunas_fixas_gantt(ws)
        col_lote = colunas_fixas.get("lote", 1)
        col_codigo = colunas_fixas.get("codigo", 2)
        col_produto = colunas_fixas.get("produto", 3)

        # 2) Varre as células uma única vez e captura somente as que têm comentário.
        # Isso evita o loop pesado data x linhas que travava o upload.
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                col_idx = cell.column

                if col_idx not in colunas_dia:
                    continue

                texto = _texto_comentario_cell(cell)
                if not texto:
                    continue

                if cell.row >= 7:
                    lote = _to_str(ws.cell(row=cell.row, column=col_lote).value)
                    codigo = _to_str(ws.cell(row=cell.row, column=col_codigo).value)
                    produto = _to_str(ws.cell(row=cell.row, column=col_produto).value)
                    prefixo = lote or codigo or produto or f"Linha Excel {cell.row}"
                    comentario = f"{prefixo}: {texto}"
                else:
                    comentario = texto

                colunas_dia[col_idx]["comentarios"].append(comentario)

        # 3) Monta um registro por data/recurso.
        for col_idx, info in sorted(colunas_dia.items(), key=lambda item: item[1]["data"]):
            vistos = set()
            comentarios_unicos = []

            for comentario in info.get("comentarios") or []:
                if not comentario or comentario in vistos:
                    continue
                vistos.add(comentario)
                comentarios_unicos.append(comentario)

            comentario_calendario = " ; ".join(comentarios_unicos) if comentarios_unicos else None

            if comentario_calendario and len(comentario_calendario) > 4000:
                comentario_calendario = comentario_calendario[:4000] + "..."

            capacidade = _to_float(info.get("capacidade"), 0)

            registros.append({
                "rodada_id": rodada_id,
                "recurso": recurso,
                "data": info.get("data"),
                "horas_disponiveis_dia": capacidade,
                "horas_indisponiveis_planejadas": max(0.0, 24.0 - capacidade),
                "comentario_calendario": comentario_calendario,
                "origem_aba": sheet_name,
            })

    except Exception:
        return registros

    finally:
        if fechar_wb and wb:
            try:
                wb.close()
            except Exception:
                pass

    return registros


def _parse_datetime(value, ano_ref=None):
    value = _limpar_nan(value)
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.to_pydatetime().replace(tzinfo=None)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        try:
            dt = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
            if pd.isna(dt):
                return None
            return dt.to_pydatetime().replace(tzinfo=None)
        except Exception:
            return None
    texto = str(value).strip()
    if not texto:
        return None
    for dayfirst in (True, False):
        try:
            dt = pd.to_datetime(texto, dayfirst=dayfirst, errors="coerce")
            if not pd.isna(dt):
                return dt.to_pydatetime().replace(tzinfo=None)
        except Exception:
            pass
    texto_lower = texto.lower().replace(".", "")
    partes = texto_lower.split()
    data_txt = partes[-1] if partes else texto_lower
    if "/" in data_txt:
        try:
            pedacos = data_txt.split("/")
            dia = int("".join(filter(str.isdigit, pedacos[0])))
            mes_txt = pedacos[1]
            ano_txt = pedacos[2] if len(pedacos) >= 3 else None
            if mes_txt.isdigit():
                mes = int(mes_txt)
            else:
                mes_key = unicodedata.normalize("NFKD", mes_txt.upper())
                mes_key = "".join(c for c in mes_key if not unicodedata.combining(c))
                mes = MESES_PT.get(mes_key[:3])
            if not mes:
                return None
            ano = int("".join(filter(str.isdigit, ano_txt))) if ano_txt else int(ano_ref or datetime.now().year)
            return datetime(ano, mes, dia)
        except Exception:
            return None
    return None


def _parse_data(value, ano_ref=None):
    dt = _parse_datetime(value, ano_ref)
    return dt.date().isoformat() if dt else None


def _hora_str(dt: Optional[datetime]) -> Optional[str]:
    if not dt:
        return None
    return dt.strftime("%H:%M:%S")


def _eh_coluna_data(col):
    return isinstance(col, (datetime, date, pd.Timestamp)) or "/" in str(col).strip()


def _select_all(query, page_size=1000):
    todos = []
    page = 0
    while True:
        res = query.range(page * page_size, (page + 1) * page_size - 1).execute()
        data = res.data or []
        todos.extend(data)
        if len(data) < page_size:
            break
        page += 1
    return todos


def _insert_em_lotes(tabela, registros, tamanho=500):
    total = 0
    for i in range(0, len(registros), tamanho):
        lote = registros[i:i + tamanho]
        if not lote:
            continue
        res = supabase.table(tabela).insert(lote).execute()
        total += len(res.data or [])
    return total


# ─── Parâmetros ───────────────────────────────────────────────────────────────

def _carregar_parametros_globais() -> Dict[str, str]:
    defaults = {
        "lead_time_liberacao_padrao": "15",
        "lead_time_prilonest": "22",
        "data_limite_liberacao": "2026-12-17",
        "qtd_padrao": "300000",
        "qtd_articaine_200": "200000",
    }
    try:
        res = supabase.table("d_mrp_parametros").select("*").execute()
        for item in res.data or []:
            chave = item.get("chave")
            valor = item.get("valor")
            if chave and valor is not None:
                defaults[chave] = str(valor)
    except Exception:
        pass
    return defaults


def _carregar_parametros_produto() -> Dict[str, Dict[str, Any]]:
    params: Dict[str, Dict[str, Any]] = {}
    try:
        res = supabase.table("d_mrp_parametros_produto").select("*").eq("ativo", True).execute()
        for item in res.data or []:
            codigo = _to_str(item.get("codigo_produto"))
            grupo = _to_str(item.get("grupo_produto"))
            registro = {
                "codigo_produto": codigo,
                "grupo_produto": grupo,
                "qtd_tubetes_padrao": _to_float(item.get("qtd_tubetes_padrao"), 0),
                "un_hora_padrao": _to_float(item.get("un_hora_padrao"), 0),
                "lead_time_liberacao_dias": _to_int(item.get("lead_time_liberacao_dias")),
            }
            if codigo:
                params[f"COD::{_normalizar_texto(codigo)}"] = registro
            if grupo:
                params[f"GRUPO::{_normalizar_texto(grupo)}"] = registro
    except Exception:
        pass
    return params


def _buscar_parametro_produto(codigo, produto, parametros_produto):
    codigo_key = f"COD::{_normalizar_texto(codigo)}"
    produto_key = f"GRUPO::{_normalizar_texto(produto)}"
    if codigo_key in parametros_produto:
        return parametros_produto[codigo_key]
    if produto_key in parametros_produto:
        return parametros_produto[produto_key]
    return None


def _param_float(parametros, chave, default=0):
    return _to_float(parametros.get(chave), default)


def _param_int(parametros, chave, default=0):
    valor = _to_int(parametros.get(chave))
    return valor if valor is not None else default


def _ajustar_para_dia_util(data_ref: date) -> date:
    if data_ref.weekday() == 5:
        return data_ref + timedelta(days=2)
    if data_ref.weekday() == 6:
        return data_ref + timedelta(days=1)
    return data_ref


def _calcular_data_liberacao(data_fim_iso, produto, codigo, parametros, parametros_produto):
    dt_fim = _str_to_date(data_fim_iso)
    if not dt_fim:
        return None
    prod_param = _buscar_parametro_produto(codigo, produto, parametros_produto)
    if prod_param and prod_param.get("lead_time_liberacao_dias"):
        lead_time = int(prod_param["lead_time_liberacao_dias"])
    elif _normalizar_texto(produto) == "PRILONEST":
        lead_time = _param_int(parametros, "lead_time_prilonest", 22)
    else:
        lead_time = _param_int(parametros, "lead_time_liberacao_padrao", 15)
    return _ajustar_para_dia_util(dt_fim + timedelta(days=lead_time)).isoformat()


def _calcular_mes_ano_liberacao(data_lib_iso, parametros):
    dt_lib = _str_to_date(data_lib_iso)
    if not dt_lib:
        return None, None
    data_limite = _str_to_date(parametros.get("data_limite_liberacao"))
    if data_limite and dt_lib > data_limite:
        return 1, data_limite.year + 1
    return dt_lib.month, dt_lib.year


def _eh_mes_liberacao_manual(etapa: Dict[str, Any], parametros: Dict[str, str]) -> bool:
    if not etapa:
        return False
    flag = etapa.get("mes_lib_manual")
    if flag is True or str(flag).strip().lower() in {"true", "1", "sim", "s", "yes"}:
        return True
    mes_atual = _to_int(etapa.get("mes_liberacao"))
    ano_atual = _to_int(etapa.get("ano_liberacao"))
    data_pa_atual = etapa.get("data_pa")
    if not mes_atual or not ano_atual or not data_pa_atual:
        return False
    mes_auto, ano_auto = _calcular_mes_ano_liberacao(data_pa_atual, parametros)
    if not mes_auto or not ano_auto:
        return False
    return mes_atual != mes_auto or ano_atual != ano_auto


def _resolver_mes_ano_liberacao_final(etapa: Dict[str, Any], data_pa_nova: Optional[str], parametros: Dict[str, str]) -> Tuple[Optional[int], Optional[int]]:
    if _eh_mes_liberacao_manual(etapa, parametros):
        return _to_int(etapa.get("mes_liberacao")), _to_int(etapa.get("ano_liberacao"))
    return _calcular_mes_ano_liberacao(data_pa_nova, parametros)


def _detectar_mes_lib_manual(data_lib_iso, mes_lib_excel, ano_lib_excel, parametros) -> bool:
    mes_excel = _to_int(mes_lib_excel)
    ano_excel = _to_int(ano_lib_excel)
    if not data_lib_iso or not mes_excel:
        return False
    mes_calc, ano_calc = _calcular_mes_ano_liberacao(data_lib_iso, parametros)
    if not mes_calc:
        return False
    if mes_excel != mes_calc:
        return True
    if ano_excel and ano_calc and ano_excel != ano_calc:
        return True
    return False


def _calcular_impacto_mudanca(data_fim_anterior, data_fim_nova, un_hora_anterior, un_hora_nova):
    dt_ant = _str_to_date(data_fim_anterior)
    dt_nova = _str_to_date(data_fim_nova)
    impacto_dias = None
    tipo_impacto = "sem_comparativo"
    if dt_ant and dt_nova:
        impacto_dias = (dt_nova - dt_ant).days
        if impacto_dias > 0:
            tipo_impacto = "atrasou"
        elif impacto_dias < 0:
            tipo_impacto = "antecipou"
        else:
            tipo_impacto = "sem_mudanca_data"
    un_ant = _to_float(un_hora_anterior, 0)
    un_nova = _to_float(un_hora_nova, 0)
    delta_un_hora = un_nova - un_ant if (un_ant or un_nova) else None
    delta_un_hora_pct = (delta_un_hora / un_ant * 100) if un_ant else None
    return {
        "impacto_dias": impacto_dias,
        "tipo_impacto": tipo_impacto,
        "delta_un_hora": delta_un_hora,
        "delta_un_hora_pct": delta_un_hora_pct,
    }


def _calcular_qtd_padrao(codigo, produto, parametros, parametros_produto, qtd_atual=0):
    prod_param = _buscar_parametro_produto(codigo, produto, parametros_produto)
    if prod_param and prod_param.get("qtd_tubetes_padrao"):
        return float(prod_param["qtd_tubetes_padrao"])
    if _normalizar_texto(codigo) == "40323":
        return _param_float(parametros, "qtd_articaine_200", 200000)
    default = _param_float(parametros, "qtd_padrao", 300000)
    return default if default else (qtd_atual or 0)


def _calcular_un_hora_padrao(codigo, produto, parametros_produto, un_hora_atual=0):
    prod_param = _buscar_parametro_produto(codigo, produto, parametros_produto)
    if prod_param and prod_param.get("un_hora_padrao"):
        return float(prod_param["un_hora_padrao"])
    return un_hora_atual or 0


def _calcular_tempo_horas(qtd, un_hora):
    qtd = _to_float(qtd, 0)
    un_hora = _to_float(un_hora, 0)
    if un_hora <= 0:
        return 0
    return qtd / un_hora


# ─── Cascata de datas ─────────────────────────────────────────────────────────

def _calcular_data_fim_por_alocacao(data_inicio: date, duracao_horas: float, horas_disponiveis_por_data: Dict[str, float], horas_ja_alocadas_por_data: Dict[str, float]) -> date:
    horas_restantes = duracao_horas
    cursor = data_inicio
    ultima_data_com_horas = data_inicio
    limite = data_inicio + timedelta(days=400)
    while horas_restantes > 0.001 and cursor <= limite:
        data_str = cursor.isoformat()
        if data_str in horas_disponiveis_por_data:
            capacidade_dia = horas_disponiveis_por_data[data_str]
        elif cursor.weekday() < 5:
            capacidade_dia = 21.0
        else:
            capacidade_dia = 0
        ja_alocado = horas_ja_alocadas_por_data.get(data_str, 0)
        saldo_disponivel = max(0, capacidade_dia - ja_alocado)
        if saldo_disponivel > 0:
            horas_usadas = min(horas_restantes, saldo_disponivel)
            horas_ja_alocadas_por_data[data_str] = ja_alocado + horas_usadas
            horas_restantes -= horas_usadas
            ultima_data_com_horas = cursor
        cursor += timedelta(days=1)
    return ultima_data_com_horas


def _proximo_dia_util(data_ref: date) -> date:
    return data_ref + timedelta(days=1)


def _aplicar_cascata_subsequentes(rodada_id: str, recurso: str, sequencia_inicio: int, horas_disponiveis_por_data: Dict[str, float], horas_alocadas_por_data: Dict[str, float], data_inicio_proximo: date, parametros: Dict[str, str], parametros_produto: Dict[str, Any]) -> List[Dict[str, Any]]:
    etapas_subsequentes = _select_all(
        supabase.table("f_mrp_etapas").select("*").eq("rodada_id", rodada_id).eq("recurso", recurso).gt("sequencia", sequencia_inicio).order("sequencia")
    )
    atualizadas = []
    data_inicio_cursor = data_inicio_proximo
    for etapa in etapas_subsequentes:
        qtd = _to_float(etapa.get("qtd_planejada"), 0)
        un_hora = _to_float(etapa.get("un_hora"), 0)
        if qtd <= 0 or un_hora <= 0:
            data_fim_nova = data_inicio_cursor
        else:
            duracao_horas = qtd / un_hora
            data_fim_nova = _calcular_data_fim_por_alocacao(data_inicio_cursor, duracao_horas, horas_disponiveis_por_data, horas_alocadas_por_data)
        data_pa = _calcular_data_liberacao(data_fim_nova.isoformat(), etapa.get("descricao_produto"), etapa.get("codigo_produto"), parametros, parametros_produto)
        mes_lib, ano_lib = _calcular_mes_ano_liberacao(data_pa, parametros)
        update = {
            "data_inicio": data_inicio_cursor.isoformat(),
            "data_fim": data_fim_nova.isoformat(),
            "data_pa": data_pa,
            "mes_producao": data_inicio_cursor.month,
            "ano_producao": data_inicio_cursor.year,
            "mes_liberacao": mes_lib,
            "ano_liberacao": ano_lib,
            "status": "cascata",
            "origem": "CASCATA_REAL",
            "observacao": f"{etapa.get('observacao') or ''} | Cascata: data início ajustada para {data_inicio_cursor.isoformat()} pela atualização do lote anterior.",
        }
        supabase.table("f_mrp_etapas").update(update).eq("id", etapa["id"]).execute()
        atualizadas.append({"etapa_id": etapa["id"], "lote": etapa.get("lote"), "recurso": recurso, "data_inicio_nova": data_inicio_cursor.isoformat(), "data_fim_nova": data_fim_nova.isoformat(), "data_lib_nova": data_pa, "mes_liberacao": mes_lib, "ano_liberacao": ano_lib})
        data_inicio_cursor = _proximo_dia_util(data_fim_nova)
    return atualizadas


# ─── Importação MPS ───────────────────────────────────────────────────────────

def _chave_lote_alocacao(item):
    return (_normalizar_texto(item.get("recurso")), _normalizar_texto(item.get("lote")), _normalizar_texto(item.get("codigo_produto")))


def _recalcular_registros_mrp(registros, alocacoes, parametros, parametros_produto):
    aloc_por_lote: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = {}
    for aloc in alocacoes:
        if not aloc.get("lote") and not aloc.get("codigo_produto"):
            continue
        horas = _to_float(aloc.get("horas_alocadas"), 0)
        if horas <= 0:
            continue
        chave = _chave_lote_alocacao(aloc)
        aloc_por_lote.setdefault(chave, []).append(aloc)

    # Guarda o registro anterior já processado para detectar blocos híbridos do Excel.
    # Existem casos em que a primeira linha do bloco tem MÊS LIB manual, e a linha
    # seguinte ainda aparece com fórmula, mas visualmente pertence ao mesmo bloco.
    # Nesses casos, se a sequência é contínua e a DATA LIB ficou no mesmo bucket,
    # a linha seguinte precisa herdar mes_lib_manual=True.
    anterior_processado = None

    for item in registros:
        codigo = item.get("codigo_produto")
        produto = item.get("descricao_produto")
        qtd_excel = _to_float(item.get("qtd_planejada"), 0)
        un_hora_excel = _to_float(item.get("un_hora"), 0)
        qtd = qtd_excel if qtd_excel > 0 else _calcular_qtd_padrao(codigo, produto, parametros, parametros_produto, item.get("qtd_planejada") or 0)
        un_hora = un_hora_excel if un_hora_excel > 0 else _calcular_un_hora_padrao(codigo, produto, parametros_produto, item.get("un_hora") or 0)
        item["qtd_planejada"] = qtd
        item["un_hora"] = un_hora
        item["duracao_horas"] = _calcular_tempo_horas(qtd, un_hora)
        alocs = aloc_por_lote.get(_chave_lote_alocacao(item), [])
        if alocs:
            datas = sorted([a["data"] for a in alocs if a.get("data")])
            if datas:
                if not item.get("data_inicio"):
                    item["data_inicio"] = datas[0]
                if not item.get("data_fim"):
                    item["data_fim"] = datas[-1]
        if item.get("data_inicio"):
            dt_inicio = _str_to_date(item["data_inicio"])
            if dt_inicio:
                if not item.get("mes_producao"):
                    item["mes_producao"] = dt_inicio.month
                if not item.get("ano_producao"):
                    item["ano_producao"] = dt_inicio.year
        data_lib = item.get("data_pa")
        if not data_lib and item.get("data_fim"):
            data_lib = _calcular_data_liberacao(item["data_fim"], produto, codigo, parametros, parametros_produto)
            item["data_pa"] = data_lib
        mes_lib_excel = _to_int(item.get("mes_liberacao"))
        ano_lib_excel = _to_int(item.get("ano_liberacao"))

        # Detecta MÊS LIB manual do Excel ANTES de qualquer cascata.
        # Prioridade 1: se a importação já detectou fórmula vs valor fixo na célula,
        # usa essa informação (mais precisa).
        # Prioridade 2: fallback por divergência entre MÊS LIB e DATA LIB.
        if item.get("mes_lib_manual") is None:
            item["mes_lib_manual"] = _detectar_mes_lib_manual(data_lib, mes_lib_excel, ano_lib_excel, parametros)
        else:
            item["mes_lib_manual"] = bool(item.get("mes_lib_manual"))

        # Marca se a flag veio diretamente do Excel/fallback antes de qualquer herança.
        # Isso evita que a herança propague indefinidamente para todos os lotes do mês.
        item["_mes_lib_manual_origem_direta"] = bool(item.get("mes_lib_manual"))

        # Herança curta de MÊS LIB manual para blocos híbridos do Excel.
        # Exemplos encontrados: uma linha manual seguida por uma linha com fórmula,
        # mas ambas no mesmo bloco de liberação. A fórmula sozinha não detecta esse caso.
        if anterior_processado and not item.get("mes_lib_manual"):
            seq_ant = _to_int(anterior_processado.get("sequencia")) or 0
            seq_atual = _to_int(item.get("sequencia")) or 0
            mes_ant = _to_int(anterior_processado.get("mes_liberacao"))
            mes_atual = _to_int(item.get("mes_liberacao"))
            ano_ant = _to_int(anterior_processado.get("ano_liberacao"))
            ano_atual = _to_int(item.get("ano_liberacao"))
            dt_lib_ant = _str_to_date(anterior_processado.get("data_pa"))
            dt_lib_atual = _str_to_date(item.get("data_pa"))

            mesmo_recurso = anterior_processado.get("recurso") == item.get("recurso")
            sequencia_continua = seq_atual == seq_ant + 1
            mesmo_bucket_lib = mes_ant == mes_atual and ano_ant == ano_atual
            data_lib_proxima = (
                dt_lib_ant is not None
                and dt_lib_atual is not None
                and 0 <= (dt_lib_atual - dt_lib_ant).days <= 1
            )

            if (
                anterior_processado.get("_mes_lib_manual_origem_direta") is True
                and mesmo_recurso
                and sequencia_continua
                and mesmo_bucket_lib
                and data_lib_proxima
            ):
                item["mes_lib_manual"] = True
                item["observacao"] = (
                    item.get("observacao") or ""
                ) + " | MÊS LIB manual herdado do bloco anterior do Excel."

        if mes_lib_excel:
            item["mes_liberacao"] = mes_lib_excel
        if ano_lib_excel:
            item["ano_liberacao"] = ano_lib_excel
        if data_lib and (not item.get("mes_liberacao") or not item.get("ano_liberacao")):
            mes_lib, ano_lib = _calcular_mes_ano_liberacao(data_lib, parametros)
            if not item.get("mes_liberacao"):
                item["mes_liberacao"] = mes_lib
            if not item.get("ano_liberacao"):
                item["ano_liberacao"] = ano_lib
        item["observacao"] = (item.get("observacao") or "") + (" | MÊS LIB manual preservado." if item.get("mes_lib_manual") else " | MÊS LIB automático recalculável.")
        anterior_processado = item

    # Campo interno usado apenas no processamento; não pode ir para o Supabase.
    for item in registros:
        item.pop("_mes_lib_manual_origem_direta", None)

    return registros


def _ler_header_dias(excel_bytes, sheet_name):
    raw = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=None, engine="openpyxl")
    datas_por_coluna = {}
    capacidades_por_coluna = {}
    for col_idx in range(raw.shape[1]):
        data_ref = _parse_data(raw.iat[4, col_idx]) if raw.shape[0] > 4 else None
        if data_ref:
            datas_por_coluna[col_idx] = data_ref
            # Gravar capacidade para TODOS os dias, inclusive 0h (paradas/shutdown)
            capacidade = _to_float(raw.iat[5, col_idx], 0) if raw.shape[0] > 5 else 0
            capacidades_por_coluna[col_idx] = capacidade
    return datas_por_coluna, capacidades_por_coluna


def _ler_aba_mps(excel_bytes, sheet_name, recurso, etapa):
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=5, engine="openpyxl")
    mes_lib_manual_por_linha = _mapear_mes_lib_manual_por_linha(excel_bytes, sheet_name)
    comentarios_gantt_por_linha = _mapear_comentarios_gantt_por_linha(excel_bytes, sheet_name)
    df = df.dropna(how="all")
    col_embalado = _buscar_coluna(df, ["EMBALADO", "EMBALADO?"])
    col_lote = _buscar_coluna(df, ["LOTE"])
    col_codigo = _buscar_coluna(df, ["CÓDIGO", "CODIGO"])
    col_produto = _buscar_coluna(df, ["PRODUTO"])
    col_tempo = _buscar_coluna(df, ["TEMPO (Horas.)", "TEMPO HORAS", "TEMPO", "HORAS"])
    col_un_hora = _buscar_coluna(df, ["UN / HORA", "UN/HORA", "UN HORA"])
    col_qtd = _buscar_coluna(df, ["QTD. (Tubetes)", "QTD Tubetes", "QTD", "QTD TUBETES"])
    col_mes_prod = _buscar_coluna(df, ["MÊS PRODUÇÃO", "MES PRODUÇÃO", "MES PRODUCAO"])
    col_ano_prod = _buscar_coluna(df, ["ANO PRODUÇÃO", "ANO PRODUCAO"])
    col_data_inicio = _buscar_coluna(df, ["DATA INÍCIO", "DATA INICIO"])
    col_data_fim = _buscar_coluna(df, ["DATA FIM"])
    col_data_lib = _buscar_coluna(df, ["DATA LIB.", "DATA LIB", "DATA LIBERAÇÃO", "DATA LIBERACAO"])
    col_mes_lib = _buscar_coluna(df, ["MÊS LIBERAÇÃO", "MES LIBERAÇÃO", "MES LIBERACAO"])
    col_ano_lib = _buscar_coluna(df, ["ANO LIBERAÇÃO", "ANO LIBERACAO"])
    registros = []
    for idx_df, row in df.iterrows():
        try:
            excel_row = int(idx_df) + 7
        except Exception:
            excel_row = None
        lote = _to_str(row.get(col_lote)) if col_lote else None
        codigo = _to_str(row.get(col_codigo)) if col_codigo else None
        produto = _to_str(row.get(col_produto)) if col_produto else None
        if not lote and not codigo and not produto:
            continue
        if produto and produto.upper() in ["PRODUTO", "NAN"]:
            continue
        ano_prod = _to_int(row.get(col_ano_prod)) if col_ano_prod else None
        ano_lib = _to_int(row.get(col_ano_lib)) if col_ano_lib else ano_prod
        data_inicio = _parse_data(row.get(col_data_inicio), ano_prod) if col_data_inicio else None
        data_fim = _parse_data(row.get(col_data_fim), ano_prod) if col_data_fim else None
        data_pa = _parse_data(row.get(col_data_lib), ano_lib) if col_data_lib else None
        if not data_inicio and not data_fim:
            continue
        registros.append({
            "lote": lote, "op": lote, "codigo_produto": codigo, "descricao_produto": produto,
            "etapa": etapa, "recurso": recurso, "linha_origem": recurso,
            "data_inicio": data_inicio, "data_fim": data_fim, "data_pa": data_pa,
            "qtd_planejada": _to_float(row.get(col_qtd), 0) if col_qtd else 0,
            "duracao_horas": _to_float(row.get(col_tempo), 0) if col_tempo else 0,
            "sequencia": len(registros) + 1, "status": "planejada", "origem": "IMPORT_MPS",
            "observacao": (
                f"Importado da aba {sheet_name}"
                + (
                    " | Comentários Gantt: "
                    + " ; ".join(comentarios_gantt_por_linha.get(excel_row, [])[:20])
                    if excel_row and comentarios_gantt_por_linha.get(excel_row)
                    else ""
                )
            ),
            "embalado": _to_str(row.get(col_embalado)) if col_embalado else None,
            "un_hora": _to_float(row.get(col_un_hora), 0) if col_un_hora else 0,
            "mes_producao": _to_int(row.get(col_mes_prod)) if col_mes_prod else None,
            "ano_producao": ano_prod,
            "mes_liberacao": _to_int(row.get(col_mes_lib)) if col_mes_lib else None,
            "ano_liberacao": ano_lib,
            "mes_lib_manual": mes_lib_manual_por_linha.get(excel_row) if excel_row else None,
        })
    return registros


def _ler_alocacoes_dia(excel_bytes, sheet_name, recurso, rodada_id):
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=5, engine="openpyxl")
    df = df.dropna(how="all")
    datas_header, capacidades_header = _ler_header_dias(excel_bytes, sheet_name)
    col_lote = _buscar_coluna(df, ["LOTE"])
    col_codigo = _buscar_coluna(df, ["CÓDIGO", "CODIGO"])
    col_produto = _buscar_coluna(df, ["PRODUTO"])
    colunas_fixas = {
        _buscar_coluna(df, ["EMBALADO", "EMBALADO?"]), _buscar_coluna(df, ["LOTE"]),
        _buscar_coluna(df, ["CÓDIGO", "CODIGO"]), _buscar_coluna(df, ["PRODUTO"]),
        _buscar_coluna(df, ["TEMPO (Horas.)", "TEMPO HORAS", "TEMPO", "HORAS"]),
        _buscar_coluna(df, ["UN / HORA", "UN/HORA", "UN HORA"]),
        _buscar_coluna(df, ["QTD. (Tubetes)", "QTD Tubetes", "QTD", "QTD TUBETES"]),
        _buscar_coluna(df, ["MÊS PRODUÇÃO", "MES PRODUÇÃO", "MES PRODUCAO"]),
        _buscar_coluna(df, ["ANO PRODUÇÃO", "ANO PRODUCAO"]),
        _buscar_coluna(df, ["DATA INÍCIO", "DATA INICIO"]), _buscar_coluna(df, ["DATA FIM"]),
        _buscar_coluna(df, ["DATA LIB.", "DATA LIB", "DATA LIBERAÇÃO", "DATA LIBERACAO"]),
        _buscar_coluna(df, ["MÊS LIBERAÇÃO", "MES LIBERAÇÃO", "MES LIBERACAO"]),
        _buscar_coluna(df, ["ANO LIBERAÇÃO", "ANO LIBERACAO"]),
    }
    colunas_fixas = {c for c in colunas_fixas if c is not None}
    colunas_datas = []
    for idx, col in enumerate(df.columns):
        if col in colunas_fixas:
            continue
        data_ref = _parse_data(col) if _eh_coluna_data(col) else None
        if not data_ref and idx in datas_header:
            data_ref = datas_header[idx]
        if data_ref:
            colunas_datas.append((idx, col, data_ref))
    registros = []
    for idx, _, data_ref in colunas_datas:
        capacidade = capacidades_header.get(idx, 0)
        registros.append({"rodada_id": rodada_id, "recurso": recurso, "lote": None, "codigo_produto": None, "descricao_produto": None, "data": data_ref, "horas_alocadas": 0, "horas_disponiveis_dia": capacidade, "origem": "CAPACIDADE_MPS"})
    for _, row in df.iterrows():
        lote = _to_str(row.get(col_lote)) if col_lote else None
        codigo = _to_str(row.get(col_codigo)) if col_codigo else None
        produto = _to_str(row.get(col_produto)) if col_produto else None
        if not lote and not codigo and not produto:
            continue
        for idx, col_data, data_ref in colunas_datas:
            horas = _to_float(row.get(col_data), 0)
            if horas <= 0:
                continue
            registros.append({"rodada_id": rodada_id, "recurso": recurso, "lote": lote, "codigo_produto": codigo, "descricao_produto": produto, "data": data_ref, "horas_alocadas": horas, "horas_disponiveis_dia": capacidades_header.get(idx, 0), "origem": "IMPORT_MPS"})
    return registros


# ─── Cópia/versionamento ──────────────────────────────────────────────────────

def _remover_campos_controle(registro: Dict[str, Any], remover_rodada: bool = False) -> Dict[str, Any]:
    novo = dict(registro)
    for campo in ["id", "criado_em", "created_at", "updated_at", "atualizado_em"]:
        novo.pop(campo, None)
    if remover_rodada:
        novo.pop("rodada_id", None)
    return novo


def _proxima_versao(mes: int, ano: int) -> int:
    try:
        res = supabase.table("f_mrp_rodadas").select("versao").eq("mes", mes).eq("ano", ano).execute()
        versoes = [_to_int(item.get("versao")) or 0 for item in (res.data or [])]
        return (max(versoes) if versoes else 0) + 1
    except Exception:
        return 1


def _nome_base_rodada(nome: Optional[str]) -> str:
    texto = (nome or "MPS").strip()
    texto = re.sub(r"\s+—\s+[A-Za-zÀ-ÿ]{3}/\d{4}\s+—\s+V\d+\s*$", "", texto)
    texto = re.sub(r"\s*-\s*V\d+\s*$", "", texto)
    texto = re.sub(r"\s+V\d+\s*$", "", texto)
    return texto.strip() or "MPS"


def _mes_label_pt(mes: int) -> str:
    labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    if 1 <= int(mes) <= 12:
        return labels[int(mes) - 1]
    return str(mes)


def _copiar_rodada_base(rodada_origem: Dict[str, Any], payload: RodadaCopiarCreate) -> Dict[str, Any]:
    mes = payload.mes or rodada_origem.get("mes")
    ano = payload.ano or rodada_origem.get("ano")
    versao = payload.versao or _proxima_versao(mes, ano)
    nome_base = _nome_base_rodada(payload.nome or rodada_origem.get("nome") or "MPS")
    observacao = payload.observacao or f"V{versao} criada a partir da V{rodada_origem.get('versao')} de {_mes_label_pt(mes)}/{ano}."
    res = supabase.table("f_mrp_rodadas").insert({"nome": nome_base, "mes": mes, "ano": ano, "versao": versao, "observacao": observacao, "status": "rascunho"}).execute()
    data = res.data or []
    if not data:
        raise HTTPException(status_code=500, detail="Rodada copiada, mas o Supabase não retornou dados.")
    return data[0]


# ─── Produção real Cogtive ────────────────────────────────────────────────────

def _identificar_recurso_real(equipamento, lote=None):
    """
    Identifica se o apontamento pertence a uma linha produtiva real.

    Correção v47:
    - NÃO inferir L1/L2 pelo número do lote quando o equipamento não é produtivo.
      Antes, apontamentos administrativos de GQ/PCP/ALMOX/PROD com lote D1050,
      D1051 etc. eram tratados como produção real da L1 só porque o lote continha
      "D1". Isso fazia lotes que ainda nem passaram pela Cogtive aparecerem como
      realizados/antecipados.
    - Só retorna L1/L2 quando o equipamento for envasadora real.
    - Lavadora, documentação, abertura de OP, separação e conferência não são fim
      Cogtive da linha de envase.
    """
    if not equipamento:
        return None

    equip_compacto = _normalizar_lote_key(equipamento)

    # Apontamentos administrativos/documentais: nunca podem virar produção real
    # da fila de MPS, mesmo quando o lote contém D1/D2 no código.
    termos_administrativos = [
        "PCP",
        "ABERTURADEOP",
        "GQ",
        "IMPRESSAODOCUMENTACAO",
        "IMPRESSAODOC",
        "DOCUMENTACAO",
        "PROD",
        "CONFERENCIADOCUMENTACAO",
        "CONFERENCIA",
        "ALMOX",
        "SEPARACAODEMATERIAL",
        "SEPARACAO",
    ]
    if any(termo in equip_compacto for termo in termos_administrativos):
        return None

    # Lavadora/Fabrima/Bausch não são a etapa de envase usada para empurrar a
    # fila L1/L2 de liberação do PA.
    for ignorar in ["LAVADORA", "LAV", "BAUSCH", "FABRIMA"]:
        if ignorar in equip_compacto:
            return None

    # Linha 1: MAQ 1 e MAQ 2 envasadoras.
    if any(p in equip_compacto for p in [
        "MAQ1ENVASADORA",
        "MAQ2ENVASADORA",
        "ENV001",
        "ENV002",
        "MAQUINA1ENVASADORA",
        "MAQUINA2ENVASADORA",
        "ENVASADORA1",
        "ENVASADORA2",
    ]):
        return "L1"

    # Linha 2: envasadora 3.
    if any(p in equip_compacto for p in [
        "L2ENVASADORA",
        "ENV003",
        "ENVASADORA3",
        "MAQ3ENVASADORA",
        "MAQUINA3ENVASADORA",
    ]):
        return "L2"

    # Em dúvida, não casa com a fila. Melhor deixar como não encontrado do que
    # atualizar um lote que ainda não passou pela linha.
    return None


def _identificar_recurso_contexto_operacional(equipamento, lote=None):
    """
    Identifica a linha para CONTEXTO OPERACIONAL de paradas.

    Regra v49:
    - O contexto do modal também deve considerar somente ENVASADORAS.
    - Lavadora continua ignorada, tanto para fim Cogtive quanto para ranking/paradas.
    - Não inferir linha pelo lote quando o equipamento não for envasadora.
    """
    if not equipamento:
        return None

    equip_compacto = _normalizar_lote_key(equipamento)

    termos_administrativos = [
        "PCP",
        "ABERTURADEOP",
        "GQ",
        "IMPRESSAODOCUMENTACAO",
        "IMPRESSAODOC",
        "DOCUMENTACAO",
        "PROD",
        "CONFERENCIADOCUMENTACAO",
        "CONFERENCIA",
        "ALMOX",
        "SEPARACAODEMATERIAL",
        "SEPARACAO",
    ]
    if any(termo in equip_compacto for termo in termos_administrativos):
        return None

    # Mantém a regra operacional: lavadora não entra nem no contexto do modal.
    for ignorar in ["LAVADORA", "LAV", "BAUSCH", "FABRIMA"]:
        if ignorar in equip_compacto:
            return None

    if any(p in equip_compacto for p in [
        "MAQ1ENVASADORA",
        "MAQ2ENVASADORA",
        "ENV001",
        "ENV002",
        "MAQUINA1ENVASADORA",
        "MAQUINA2ENVASADORA",
        "ENVASADORA1",
        "ENVASADORA2",
    ]):
        return "L1"

    if any(p in equip_compacto for p in [
        "L2ENVASADORA",
        "ENV003",
        "ENVASADORA3",
        "MAQ3ENVASADORA",
        "MAQUINA3ENVASADORA",
    ]):
        return "L2"

    return None

def _lote_base_real(lote, ordem=None):
    lote_extraido = _extrair_lote_do_texto(lote)
    if lote_extraido:
        return lote_extraido
    ordem_extraida = _extrair_lote_do_texto(ordem)
    if ordem_extraida:
        return ordem_extraida
    return _to_str(lote) or _to_str(ordem)


def _chaves_possiveis_lote(lote=None, op=None):
    chaves = set()
    for valor in [lote, op]:
        chave = _normalizar_lote_key(valor)
        extraida = _extrair_lote_do_texto(valor)
        core = _lote_core_key(valor)
        if chave:
            chaves.add(chave)
        if extraida:
            chaves.add(extraida)
        if core:
            chaves.add(core)
    return chaves


def _detectar_linha_header_real(excel_bytes: bytes, sheet_name: str) -> Optional[int]:
    raw = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=None, engine="openpyxl", nrows=40)
    obrigatorias = {"DATA INICIAL", "DATA FINAL", "EQUIPAMENTO"}
    desejaveis = {"LOTE", "ORDEM", "TIPO DE EVENTO"}
    for idx in range(raw.shape[0]):
        valores = {_normalizar_coluna(v) for v in raw.iloc[idx].tolist() if _to_str(v)}
        if obrigatorias.issubset(valores) and len(valores.intersection(desejaveis)) >= 2:
            return idx
    return None


def _preparar_registro_real_para_banco(registro: Dict[str, Any]) -> Dict[str, Any]:
    novo = dict(registro)
    novo.pop("_data_real_fim_dt", None)
    novo.pop("_data_real_inicio_dt", None)
    return novo


def _ler_apontamentos_cogtive(excel_bytes: bytes, rodada_id: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Lê o relatório do Cogtive em dois blocos:
    - producoes: somente eventos de PRODUÇÃO, agrupados por lote/recurso para aplicação do realizado.
    - paradas: eventos diferentes de PRODUÇÃO, preservados linha a linha para contexto operacional.

    Importante:
    As paradas NÃO são tratadas como causa do atraso. Elas servem apenas para exibir
    quais eventos/paradas ocorreram no dia do FIM ANTERIOR do lote.
    """
    xls = pd.ExcelFile(BytesIO(excel_bytes), engine="openpyxl")
    producoes_agrupadas: Dict[Tuple[str, str], Dict[str, Any]] = {}
    producoes_raw: List[Dict[str, Any]] = []
    paradas: List[Dict[str, Any]] = []

    for sheet_name in xls.sheet_names:
        header_row = _detectar_linha_header_real(excel_bytes, sheet_name)
        if header_row is None:
            continue

        df = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=header_row, engine="openpyxl")
        df = df.dropna(how="all")

        if df.empty:
            continue

        col_data_ini = _buscar_coluna(df, ["DATA INICIAL", "DATA INICIO", "DATA INÍCIO"])
        col_data_fim = _buscar_coluna(df, ["DATA FINAL", "DATA FIM"])
        col_equip = _buscar_coluna(df, ["EQUIPAMENTO", "MAQUINA", "MÁQUINA"])
        col_ordem = _buscar_coluna(df, ["ORDEM", "OP"])
        col_lote = _buscar_coluna(df, ["LOTE"])
        col_produto = _buscar_coluna(df, ["PRODUTO"])
        col_sku = _buscar_coluna(df, ["SKU", "CÓDIGO", "CODIGO"])
        col_qtd = _buscar_coluna(df, ["QUANTIDADE PRODUZIDA", "QTD PRODUZIDA", "QUANTIDADE", "QTD"])
        col_tipo_evento = _buscar_coluna(df, ["TIPO DE EVENTO", "TIPO EVENTO"])
        col_evento = _buscar_coluna(df, ["EVENTO", "DESCRIÇÃO", "DESCRICAO"])
        col_duracao = _buscar_coluna(df, ["DURAÇÃO", "DURACAO"])

        if not col_equip or not col_data_fim:
            continue

        for _, row in df.iterrows():
            equipamento = _to_str(row.get(col_equip)) if col_equip else None
            if not equipamento:
                continue

            lote = _to_str(row.get(col_lote)) if col_lote else None
            ordem = _to_str(row.get(col_ordem)) if col_ordem else None
            identificador_lote = _lote_base_real(lote, ordem)

            tipo_evento_raw = _to_str(row.get(col_tipo_evento)) if col_tipo_evento else None
            tipo_evento = _normalizar_texto(tipo_evento_raw)
            evento = _to_str(row.get(col_evento)) if col_evento else None

            recurso_real = _identificar_recurso_real(equipamento, identificador_lote)
            recurso_contexto = _identificar_recurso_contexto_operacional(equipamento, identificador_lote)

            # Produção real só pode vir de equipamento produtivo de envase.
            # Parada/contexto pode vir de lavadora/envasadora da mesma linha.
            if "PRODUCAO" in tipo_evento:
                recurso = recurso_real
            else:
                recurso = recurso_contexto

            if not recurso:
                continue

            data_fim_dt = _parse_datetime(row.get(col_data_fim))
            if not data_fim_dt:
                continue

            data_inicio_dt = _parse_datetime(row.get(col_data_ini)) if col_data_ini else None
            duracao_horas = _to_float(row.get(col_duracao), 0) if col_duracao else 0

            registro_base = {
                "rodada_id": rodada_id,
                "recurso": recurso,
                "lote": identificador_lote,
                "op": ordem,
                "codigo_produto": _to_str(row.get(col_sku)) if col_sku else None,
                "descricao_produto": _to_str(row.get(col_produto)) if col_produto else None,
                "equipamento": equipamento,
                "data_real_inicio": data_inicio_dt.date().isoformat() if data_inicio_dt else None,
                "hora_inicio": _hora_str(data_inicio_dt),
                "data_real_fim": data_fim_dt.date().isoformat(),
                "hora_fim": _hora_str(data_fim_dt),
                "horas_reais": duracao_horas,
                "qtd_real": _to_float(row.get(col_qtd), 0) if col_qtd else 0,
                "tipo_evento": "PRODUCAO" if "PRODUCAO" in tipo_evento else (tipo_evento_raw or tipo_evento or "PARADA"),
                "evento": evento,
                "origem_arquivo": "IMPORT_REAL_COGTIVE",
                "_data_real_inicio_dt": data_inicio_dt,
                "_data_real_fim_dt": data_fim_dt,
            }

            if "PRODUCAO" in tipo_evento:
                # Mantém a produção linha a linha para leitura de horas produtivas reais do dia.
                # IMPORTANTE: para calcular horas reais do dia/recurso, a linha de produção não
                # precisa obrigatoriamente ter lote/OP preenchido. O lote é obrigatório apenas
                # para aplicar o realizado na etapa do MPS.
                producoes_raw.append(registro_base)

                if not identificador_lote:
                    continue

                chave = (_normalizar_lote_key(identificador_lote), recurso)
                qtd_linha = _to_float(row.get(col_qtd), 0) if col_qtd else 0

                if chave not in producoes_agrupadas:
                    producoes_agrupadas[chave] = dict(registro_base)
                    producoes_agrupadas[chave]["qtd_real"] = qtd_linha
                    producoes_agrupadas[chave]["horas_reais"] = duracao_horas
                    producoes_agrupadas[chave]["qtd_eventos_producao"] = 1
                    producoes_agrupadas[chave]["equipamentos_producao"] = sorted({equipamento})
                    continue

                atual = producoes_agrupadas[chave]
                atual["qtd_real"] = _to_float(atual.get("qtd_real"), 0) + qtd_linha
                atual["horas_reais"] = _to_float(atual.get("horas_reais"), 0) + duracao_horas
                atual["qtd_eventos_producao"] = int(atual.get("qtd_eventos_producao") or 0) + 1
                equipamentos = set(atual.get("equipamentos_producao") or [])
                equipamentos.add(equipamento)
                atual["equipamentos_producao"] = sorted(equipamentos)

                if data_inicio_dt and (
                    not atual.get("_data_real_inicio_dt")
                    or data_inicio_dt < atual.get("_data_real_inicio_dt")
                ):
                    atual["_data_real_inicio_dt"] = data_inicio_dt
                    atual["data_real_inicio"] = data_inicio_dt.date().isoformat()
                    atual["hora_inicio"] = _hora_str(data_inicio_dt)

                if data_fim_dt > atual.get("_data_real_fim_dt"):
                    atual["_data_real_fim_dt"] = data_fim_dt
                    atual["data_real_fim"] = data_fim_dt.date().isoformat()
                    atual["hora_fim"] = _hora_str(data_fim_dt)
                    atual["equipamento"] = equipamento
                    atual["evento"] = evento
                    atual["op"] = ordem
                    atual["codigo_produto"] = _to_str(row.get(col_sku)) if col_sku else atual.get("codigo_produto")
                    atual["descricao_produto"] = _to_str(row.get(col_produto)) if col_produto else atual.get("descricao_produto")
            else:
                # Mantém paradas linha a linha. Não usamos como causa, apenas como
                # contexto no dia do FIM ANTERIOR.
                paradas.append(registro_base)

    return {
        "producoes": [_preparar_registro_real_para_banco(item) for item in producoes_agrupadas.values()],
        "producoes_raw": [_preparar_registro_real_para_banco(item) for item in producoes_raw],
        "paradas": [_preparar_registro_real_para_banco(item) for item in paradas],
    }


def _ler_producao_real_cogtive(excel_bytes: bytes, rodada_id: str) -> List[Dict[str, Any]]:
    """
    Compatibilidade com o fluxo antigo: retorna somente as produções agrupadas.
    """
    return _ler_apontamentos_cogtive(excel_bytes, rodada_id).get("producoes", [])


def _data_evento_cobre_referencia(data_inicio_iso: Optional[str], data_fim_iso: Optional[str], data_ref_iso: Optional[str]) -> bool:
    data_ref = _str_to_date(data_ref_iso)
    if not data_ref:
        return False

    data_inicio = _str_to_date(data_inicio_iso) or _str_to_date(data_fim_iso)
    data_fim = _str_to_date(data_fim_iso) or data_inicio

    if not data_inicio or not data_fim:
        return False

    return data_inicio <= data_ref <= data_fim



def _data_evento_intersecta_janela(
    data_inicio_iso: Optional[str],
    data_fim_iso: Optional[str],
    janela_inicio_iso: Optional[str],
    janela_fim_iso: Optional[str],
) -> bool:
    """
    Compatibilidade: valida interseção por data.
    A lógica principal do modal usa _evento_intersecta_janela_operacional,
    que considera hora quando disponível.
    """
    janela_inicio = _str_to_date(janela_inicio_iso)
    janela_fim = _str_to_date(janela_fim_iso) or janela_inicio

    if not janela_inicio:
        return False
    if janela_fim and janela_fim < janela_inicio:
        janela_fim = janela_inicio

    evento_inicio = _str_to_date(data_inicio_iso) or _str_to_date(data_fim_iso)
    evento_fim = _str_to_date(data_fim_iso) or evento_inicio

    if not evento_inicio or not evento_fim:
        return False

    return evento_inicio <= janela_fim and evento_fim >= janela_inicio


def _combinar_data_hora_para_janela(data_iso: Optional[str], hora: Optional[str] = None, fim_do_dia: bool = False) -> Optional[datetime]:
    """
    Combina data + hora para comparar janelas do Cogtive.
    Quando não há hora:
    - início de janela/evento: 00:00
    - fim de janela/evento: 23:59:59
    """
    if not data_iso:
        return None

    data_txt = str(data_iso).strip()
    hora_txt = str(hora or "").strip()

    # Se já veio uma string com data e hora, usa diretamente.
    if ("T" in data_txt or ":" in data_txt) and not hora_txt:
        dt = _parse_datetime(data_txt)
        if dt:
            return dt

    data_base = data_txt[:10]
    if hora_txt:
        dt = _parse_datetime(f"{data_base} {hora_txt[:8]}")
        if dt:
            return dt

    data_dt = _str_to_date(data_base)
    if not data_dt:
        return None

    if fim_do_dia:
        return datetime.combine(data_dt, datetime.max.time()).replace(microsecond=0)
    return datetime.combine(data_dt, datetime.min.time())


def _extrair_data_hora_texto(valor: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """Extrai YYYY-MM-DD e HH:MM:SS de textos salvos na observação."""
    if not valor:
        return None, None
    texto = str(valor).strip()
    dt = _parse_datetime(texto)
    if dt:
        return dt.date().isoformat(), _hora_str(dt)

    data_match = re.search(r"(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})", texto)
    hora_match = re.search(r"(\d{1,2}:\d{2}(?::\d{2})?)", texto)
    data_iso = None
    if data_match:
        parsed = _str_to_date(data_match.group(1)) or (_parse_datetime(data_match.group(1)).date().isoformat() if _parse_datetime(data_match.group(1)) else None)
        data_iso = parsed
    hora = hora_match.group(1) if hora_match else None
    if hora and len(hora.split(":")) == 2:
        hora = f"{hora}:00"
    return data_iso, hora


def _resolver_janela_operacional_mudanca(mudanca: Dict[str, Any]) -> Tuple[Optional[datetime], Optional[datetime], Optional[str], Optional[str]]:
    """
    Define a janela correta do modal de paradas.

    Regra operacional:
    - Para o lote com Cogtive real/parcial: do início planejado ao início real produtivo.
    - Se não houver início real produtivo, usa o fim novo/projetado como fallback.
    - Para cascata: usa a janela herdada do lote âncora que causou o arraste.
    """
    inicio_iso = (
        mudanca.get("janela_gap_inicio")
        or mudanca.get("janela_paradas_inicio")
        or mudanca.get("data_inicio_anterior")
        or mudanca.get("data_inicio_planejada_anterior")
        or mudanca.get("data_inicio")
        or mudanca.get("data_fim_anterior")
    )

    fim_iso = (
        mudanca.get("janela_gap_fim")
        or mudanca.get("janela_paradas_fim")
        or mudanca.get("data_inicio_real_cogtive")
        or mudanca.get("data_fim_nova")
        or mudanca.get("data_fim_anterior")
        or inicio_iso
    )

    hora_fim = (
        mudanca.get("janela_gap_fim_hora")
        or mudanca.get("hora_inicio_real")
        or mudanca.get("hora_inicio_real_cogtive")
    )

    inicio_dt = _combinar_data_hora_para_janela(inicio_iso, None, fim_do_dia=False)

    # Se o fim da janela veio do início real produtivo, a hora importa.
    # Sem hora, tratamos como fim do dia para não esconder paradas do período.
    fim_dt = _combinar_data_hora_para_janela(fim_iso, hora_fim, fim_do_dia=(not bool(hora_fim)))

    if inicio_dt and fim_dt and fim_dt < inicio_dt:
        fim_dt = inicio_dt

    inicio_label = inicio_dt.date().isoformat() if inicio_dt else None
    fim_label = fim_dt.date().isoformat() if fim_dt else None

    return inicio_dt, fim_dt, inicio_label, fim_label


def _intervalo_evento_para_janela(row: Dict[str, Any]) -> Optional[Tuple[datetime, datetime]]:
    """Monta intervalo de um evento/parada considerando data, hora e duração."""
    data_inicio = row.get("data_inicio") or row.get("data_real_inicio") or row.get("data_inicial")
    data_fim = row.get("data_fim") or row.get("data_real_fim") or row.get("data_final")
    hora_inicio = row.get("hora_inicio")
    hora_fim = row.get("hora_fim")
    duracao_horas = _to_float(row.get("duracao_horas", row.get("horas_reais", 0)), 0)

    inicio_dt = _combinar_data_hora_para_janela(data_inicio, hora_inicio, fim_do_dia=False) if (data_inicio and hora_inicio) else None
    fim_dt = _combinar_data_hora_para_janela(data_fim, hora_fim, fim_do_dia=True if not hora_fim else False)

    if fim_dt and not inicio_dt and duracao_horas > 0:
        inicio_dt = fim_dt - timedelta(hours=duracao_horas)

    if inicio_dt and not fim_dt and duracao_horas > 0:
        fim_dt = inicio_dt + timedelta(hours=duracao_horas)

    if not inicio_dt:
        inicio_dt = _combinar_data_hora_para_janela(data_inicio or data_fim, None, fim_do_dia=False)
    if not fim_dt:
        fim_dt = _combinar_data_hora_para_janela(data_fim or data_inicio, None, fim_do_dia=True)

    if not inicio_dt or not fim_dt:
        return None
    if fim_dt < inicio_dt:
        fim_dt = inicio_dt

    return inicio_dt, fim_dt


def _evento_intersecta_janela_operacional(row: Dict[str, Any], janela_inicio: Optional[datetime], janela_fim: Optional[datetime]) -> bool:
    if not janela_inicio or not janela_fim:
        return False
    intervalo = _intervalo_evento_para_janela(row)
    if not intervalo:
        return False
    evento_inicio, evento_fim = intervalo
    return evento_inicio <= janela_fim and evento_fim >= janela_inicio

def _buscar_paradas_cogtive_rodada(rodada_id: str) -> List[Dict[str, Any]]:
    """
    Busca paradas já importadas do Cogtive na tabela f_mrp_producao_real.
    Só considera eventos diferentes de PRODUÇÃO.
    """
    try:
        rows = _select_all(
            supabase.table("f_mrp_producao_real")
            .select("*")
            .eq("rodada_id", rodada_id)
        )
    except Exception:
        return []

    paradas = []

    for row in rows:
        tipo_evento = _normalizar_texto(row.get("tipo_evento"))
        if "PRODUCAO" in tipo_evento:
            continue

        data_inicio = row.get("data_real_inicio")
        data_fim = row.get("data_real_fim")
        paradas.append({
            "id": row.get("id"),
            "recurso": row.get("recurso"),
            "linha": row.get("recurso"),
            "equipamento": row.get("equipamento"),
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "data_inicial": data_inicio,
            "data_final": data_fim,
            "hora_inicio": row.get("hora_inicio"),
            "hora_fim": row.get("hora_fim"),
            "tipo_evento": row.get("tipo_evento"),
            "evento": row.get("evento"),
            "duracao_horas": round(_to_float(row.get("horas_reais"), 0), 2),
            "lote": row.get("lote"),
            "op": row.get("op"),
        })

    return paradas



def _paradas_para_mudanca(mudanca: Dict[str, Any], paradas: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Para uma mudança do Cogtive/cascata, traz as paradas ocorridas na MESMA LINHA
    dentro do GAP OPERACIONAL.

    GAP operacional:
    - lote real/parcial: início planejado → início real produtivo;
    - cascata: janela herdada do lote âncora que arrastou a fila.
    """
    recurso = _normalizar_texto(mudanca.get("recurso"))
    janela_inicio_dt, janela_fim_dt, _, _ = _resolver_janela_operacional_mudanca(mudanca)

    if not recurso or not janela_inicio_dt or not janela_fim_dt:
        return []

    eventos = []

    for parada in paradas:
        if _normalizar_texto(parada.get("recurso")) != recurso:
            continue

        if not _evento_intersecta_janela_operacional(parada, janela_inicio_dt, janela_fim_dt):
            continue

        eventos.append(parada)

    eventos.sort(key=lambda x: (
        str(x.get("data_inicio") or x.get("data_real_inicio") or x.get("data_inicial") or ""),
        str(x.get("hora_inicio") or ""),
        str(x.get("hora_fim") or ""),
        str(x.get("evento") or ""),
    ))

    return eventos

def _evento_intervalo_absoluto(row: Dict[str, Any]) -> Optional[Tuple[datetime, datetime]]:
    """
    Monta o intervalo absoluto do apontamento Cogtive.

    Regra importante:
    - O banco normalmente guarda data_real_inicio/data_real_fim como DATA.
    - A hora confiável que temos é hora_fim.
    - Então, quando não houver hora_inicio, o início deve ser inferido por:
        inicio = fim - duração
    - Nunca devemos usar data_inicio 00:00 como início quando existe duração,
      porque isso transforma qualquer apontamento em um intervalo enorme e pode
      gerar 24h produtivas falsas.
    """
    data_fim = row.get("data_fim") or row.get("data_real_fim") or row.get("data_final")
    data_inicio = row.get("data_inicio") or row.get("data_real_inicio") or row.get("data_inicial")

    hora_fim = row.get("hora_fim")
    hora_inicio = row.get("hora_inicio")

    duracao_horas = _to_float(
        row.get("duracao_horas", row.get("horas_reais", 0)),
        0,
    )

    fim_dt: Optional[datetime] = None
    inicio_dt: Optional[datetime] = None

    # 1) Fim do evento: usa data final + hora final quando existir.
    if data_fim and hora_fim:
        fim_dt = _parse_datetime(f"{str(data_fim)[:10]} {str(hora_fim)[:8]}")
    elif data_fim:
        fim_dt = _parse_datetime(data_fim)

    # 2) Início do evento: só usa data início diretamente se tiver hora início.
    #    Se não tiver hora início, NÃO usa 00:00 como fallback antes de tentar duração.
    if data_inicio and hora_inicio:
        inicio_dt = _parse_datetime(f"{str(data_inicio)[:10]} {str(hora_inicio)[:8]}")

    # 3) Caminho principal para o Cogtive salvo no banco:
    #    data_fim + hora_fim + duração => intervalo real.
    if fim_dt and not inicio_dt and duracao_horas > 0:
        inicio_dt = fim_dt - timedelta(hours=duracao_horas)

    # 4) Caso raro: temos início com hora, mas não temos fim.
    if inicio_dt and not fim_dt and duracao_horas > 0:
        fim_dt = inicio_dt + timedelta(hours=duracao_horas)

    # 5) Último fallback seguro: se vier data início completa com hora embutida,
    #    tenta parsear. Se for só data, isso vira 00:00, então só aceitamos se
    #    não houver duração ou se o intervalo continuar plausível.
    if not inicio_dt and data_inicio:
        inicio_parse = _parse_datetime(data_inicio)
        if inicio_parse and (
            "T" in str(data_inicio)
            or ":" in str(data_inicio)
            or duracao_horas <= 0
        ):
            inicio_dt = inicio_parse

    if not fim_dt or not inicio_dt:
        return None

    if fim_dt <= inicio_dt:
        return None

    return inicio_dt, fim_dt

def _horas_uniao_intervalos(intervalos: List[Tuple[datetime, datetime]]) -> float:
    if not intervalos:
        return 0.0

    ordenados = sorted(intervalos, key=lambda x: x[0])
    total_segundos = 0.0
    ini_atual, fim_atual = ordenados[0]

    for ini, fim in ordenados[1:]:
        if ini <= fim_atual:
            if fim > fim_atual:
                fim_atual = fim
        else:
            total_segundos += (fim_atual - ini_atual).total_seconds()
            ini_atual, fim_atual = ini, fim

    total_segundos += (fim_atual - ini_atual).total_seconds()
    return total_segundos / 3600


def _horas_planejadas_gantt_dia(rodada_id: str, recurso: str, data_ref_iso: Optional[str]) -> float:
    if not data_ref_iso or not recurso:
        return 0.0

    try:
        rows = _select_all(
            supabase.table("f_mrp_alocacoes_dia")
            .select("lote,horas_alocadas,horas_disponiveis_dia,data,recurso")
            .eq("rodada_id", rodada_id)
            .eq("recurso", recurso)
            .eq("data", data_ref_iso)
        )
    except Exception:
        return 0.0

    if not rows:
        return 0.0

    capacidades = [
        _to_float(r.get("horas_disponiveis_dia"), 0)
        for r in rows
        if not r.get("lote") and _to_float(r.get("horas_disponiveis_dia"), 0) > 0
    ]

    if capacidades:
        return round(max(capacidades), 2)

    return round(sum(_to_float(r.get("horas_alocadas"), 0) for r in rows if r.get("lote")), 2)


def _horas_produtivas_reais_cogtive_dia(rodada_id: str, recurso: str, data_ref_iso: Optional[str]) -> float:
    """
    Calcula horas produtivas reais do Cogtive no dia de referência.

    Regra:
    - Considera apenas tipo_evento = PRODUÇÃO.
    - Usa a duração real do apontamento e o intervalo inferido pelo fim + duração.
    - Para L1, MAQ 1 e MAQ 2 operam em paralelo:
        se as duas produzem ao mesmo tempo, conta uma vez;
        se só uma produz em uma janela, essa janela conta.
      Isso é exatamente a união dos intervalos produtivos das máquinas.
    - Para L2, a mesma união também é segura.
    """
    data_ref = _str_to_date(data_ref_iso)
    recurso_norm = _normalizar_texto(recurso)

    if not data_ref or not recurso_norm:
        return 0.0

    inicio_dia = datetime.combine(data_ref, datetime.min.time())
    fim_dia = inicio_dia + timedelta(days=1)

    try:
        rows = _select_all(
            supabase.table("f_mrp_producao_real")
            .select("recurso,tipo_evento,data_real_inicio,data_real_fim,hora_fim,horas_reais,equipamento")
            .eq("rodada_id", rodada_id)
            .eq("recurso", recurso_norm)
        )
    except Exception:
        return 0.0

    intervalos: List[Tuple[datetime, datetime]] = []
    duracao_por_equipamento: Dict[str, float] = {}

    for row in rows:
        tipo_evento = _normalizar_texto(row.get("tipo_evento"))
        if "PRODUCAO" not in tipo_evento:
            continue

        duracao = _to_float(row.get("horas_reais"), 0)
        if duracao <= 0:
            continue

        equipamento = _normalizar_texto(row.get("equipamento")) or "SEM EQUIPAMENTO"
        duracao_por_equipamento[equipamento] = duracao_por_equipamento.get(equipamento, 0.0) + duracao

        intervalo = _evento_intervalo_absoluto(row)
        if not intervalo:
            continue

        ini, fim = intervalo

        # Mantém somente o pedaço que encosta no dia analisado.
        ini_clip = max(ini, inicio_dia)
        fim_clip = min(fim, fim_dia)

        if fim_clip > ini_clip:
            intervalos.append((ini_clip, fim_clip))

    if intervalos:
        return round(_horas_uniao_intervalos(intervalos), 2)

    # Fallback apenas quando não foi possível montar intervalos.
    # Para L1, não soma MAQ 1 + MAQ 2 porque elas operam em paralelo.
    # Pega o maior tempo por equipamento como aproximação conservadora.
    if recurso_norm == "L1" and duracao_por_equipamento:
        return round(max(duracao_por_equipamento.values()), 2)

    return round(sum(duracao_por_equipamento.values()), 2)


def _datas_no_intervalo(inicio_iso: Optional[str], fim_iso: Optional[str]) -> List[str]:
    inicio = _str_to_date(inicio_iso)
    fim = _str_to_date(fim_iso) or inicio
    if not inicio:
        return []
    if fim < inicio:
        fim = inicio
    datas = []
    cursor = inicio
    while cursor <= fim:
        datas.append(cursor.isoformat())
        cursor += timedelta(days=1)
    return datas


def _horas_planejadas_gantt_janela(rodada_id: str, recurso: str, inicio_iso: Optional[str], fim_iso: Optional[str]) -> float:
    total = 0.0
    for data_ref in _datas_no_intervalo(inicio_iso, fim_iso):
        total += _horas_planejadas_gantt_dia(rodada_id, recurso, data_ref)
    return round(total, 2)


def _horas_produtivas_reais_cogtive_janela(
    rodada_id: str,
    recurso: str,
    janela_inicio: Optional[datetime],
    janela_fim: Optional[datetime],
) -> float:
    recurso_norm = _normalizar_texto(recurso)
    if not rodada_id or not recurso_norm or not janela_inicio or not janela_fim:
        return 0.0

    try:
        rows = _select_all(
            supabase.table("f_mrp_producao_real")
            .select("recurso,tipo_evento,data_real_inicio,data_real_fim,hora_inicio,hora_fim,horas_reais,equipamento")
            .eq("rodada_id", rodada_id)
            .eq("recurso", recurso_norm)
        )
    except Exception:
        return 0.0

    intervalos: List[Tuple[datetime, datetime]] = []
    duracao_por_equipamento: Dict[str, float] = {}

    for row in rows:
        tipo_evento = _normalizar_texto(row.get("tipo_evento"))
        if "PRODUCAO" not in tipo_evento:
            continue

        duracao = _to_float(row.get("horas_reais"), 0)
        if duracao <= 0:
            continue

        equipamento = _normalizar_texto(row.get("equipamento")) or "SEM EQUIPAMENTO"
        duracao_por_equipamento[equipamento] = duracao_por_equipamento.get(equipamento, 0.0) + duracao

        intervalo = _evento_intervalo_absoluto(row)
        if not intervalo:
            continue

        ini, fim = intervalo
        ini_clip = max(ini, janela_inicio)
        fim_clip = min(fim, janela_fim)

        if fim_clip > ini_clip:
            intervalos.append((ini_clip, fim_clip))

    if intervalos:
        return round(_horas_uniao_intervalos(intervalos), 2)

    if recurso_norm == "L1" and duracao_por_equipamento:
        return round(max(duracao_por_equipamento.values()), 2)

    return round(sum(duracao_por_equipamento.values()), 2)


def _enriquecer_mudanca_com_horas_operacionais(rodada_id: str, mudanca: Dict[str, Any]) -> Dict[str, Any]:
    item = dict(mudanca)
    recurso = _normalizar_texto(item.get("recurso"))
    janela_inicio_dt, janela_fim_dt, janela_inicio_iso, janela_fim_iso = _resolver_janela_operacional_mudanca(item)

    planejadas = _horas_planejadas_gantt_janela(rodada_id, recurso, janela_inicio_iso, janela_fim_iso)
    reais = _horas_produtivas_reais_cogtive_janela(rodada_id, recurso, janela_inicio_dt, janela_fim_dt)

    item["data_referencia_operacional"] = janela_inicio_iso
    item["janela_paradas_inicio"] = janela_inicio_iso
    item["janela_paradas_fim"] = janela_fim_iso
    item["horas_produtivas_planejadas_dia"] = planejadas
    item["horas_produtivas_reais_dia"] = reais
    item["gap_horas_produtivas_dia"] = round(reais - planejadas, 2)

    return item


def _anexar_paradas_mudancas(mudancas: List[Dict[str, Any]], paradas: List[Dict[str, Any]], rodada_id: Optional[str] = None) -> List[Dict[str, Any]]:
    enriquecidas = []

    for mudanca in mudancas:
        item = dict(mudanca)
        janela_inicio_dt, janela_fim_dt, janela_inicio_iso, janela_fim_iso = _resolver_janela_operacional_mudanca(item)
        eventos = _paradas_para_mudanca(item, paradas)
        total_horas = sum(_to_float(e.get("duracao_horas"), 0) for e in eventos)

        item["paradas_dia_fim_anterior"] = eventos
        item["total_paradas_dia_fim_anterior"] = len(eventos)
        item["horas_paradas_dia_fim_anterior"] = round(total_horas, 2)
        item["janela_paradas_inicio"] = janela_inicio_iso
        item["janela_paradas_fim"] = janela_fim_iso

        if rodada_id:
            item = _enriquecer_mudanca_com_horas_operacionais(rodada_id, item)

        enriquecidas.append(item)

    return enriquecidas

def _buscar_lotes_realizados_em_versoes_anteriores(rodada_atual: Dict[str, Any]) -> set:
    mes = _to_int(rodada_atual.get("mes"))
    ano = _to_int(rodada_atual.get("ano"))
    versao_atual = _to_int(rodada_atual.get("versao")) or 0
    if not mes or not ano or versao_atual <= 1:
        return set()
    try:
        rodadas_mes = _select_all(supabase.table("f_mrp_rodadas").select("*").eq("mes", mes).eq("ano", ano))
    except Exception:
        return set()
    rodadas_anteriores = [r for r in rodadas_mes if (_to_int(r.get("versao")) or 0) < versao_atual and r.get("id")]
    if not rodadas_anteriores:
        return set()
    chaves = set()
    for rodada in rodadas_anteriores:
        try:
            etapas = _select_all(supabase.table("f_mrp_etapas").select("recurso,lote,op,origem,status,observacao").eq("rodada_id", rodada["id"]))
        except Exception:
            continue
        for etapa in etapas:
            obs = etapa.get("observacao") or ""
            origem = _normalizar_texto(etapa.get("origem"))
            status = _normalizar_texto(etapa.get("status"))
            foi_realizado = (origem == "REAL_COGTIVE" or status == "REALIZADO" or "Atualizado pelo real Cogtive" in obs)
            if not foi_realizado:
                continue
            recurso_norm = _normalizar_texto(etapa.get("recurso"))
            if recurso_norm not in ["L1", "L2"]:
                continue
            for chave_lote in _chaves_possiveis_lote(etapa.get("lote"), etapa.get("op")):
                if chave_lote:
                    chaves.add((recurso_norm, chave_lote))
    return chaves


def _etapa_ja_realizada_em_versao_anterior(etapa: Dict[str, Any], chaves_realizadas_anteriores: set) -> bool:
    recurso_norm = _normalizar_texto(etapa.get("recurso"))
    for chave_lote in _chaves_possiveis_lote(etapa.get("lote"), etapa.get("op")):
        if (recurso_norm, chave_lote) in chaves_realizadas_anteriores:
            return True
    return False


def _buscar_etapa_para_real(recurso_real: str, lote_real_norm: str, etapas_por_chave: Dict[Tuple[str, str], Dict[str, Any]], etapas_atualizadas: set):
    recurso_norm = _normalizar_texto(recurso_real)
    lote_full = _normalizar_lote_key(lote_real_norm)
    lote_core = _lote_core_key(lote_real_norm)
    for chave_lote in [lote_full, lote_core]:
        if not chave_lote:
            continue
        etapa = etapas_por_chave.get((recurso_norm, chave_lote))
        if etapa and etapa.get("id") not in etapas_atualizadas:
            return etapa, "recurso_lote" if chave_lote == lote_full else "recurso_lote_core"
    return None, "nao_encontrada"


# ─── Goal seek UN/HORA (binary search — replica o goal seek manual do Excel) ──

def _montar_horas_anteriores_por_data(alocacoes, recurso, sequencia_lote, etapas_por_seq):
    lotes_anteriores: set = set()
    for seq, etapa in etapas_por_seq.items():
        if seq < sequencia_lote:
            lote_key = _normalizar_lote_key(etapa.get("lote") or "")
            if lote_key:
                lotes_anteriores.add(lote_key)
    mapa: Dict[str, float] = {}
    for aloc in alocacoes:
        if _normalizar_texto(aloc.get("recurso")) != recurso:
            continue
        lote_aloc = _normalizar_lote_key(aloc.get("lote") or "")
        if not lote_aloc or lote_aloc not in lotes_anteriores:
            continue
        data = str(aloc.get("data") or "")
        if not data:
            continue
        h = _to_float(aloc.get("horas_alocadas"), 0)
        if h > 0:
            mapa[data] = mapa.get(data, 0.0) + h
    return mapa


def _goal_seek_un_hora(data_inicio, data_fim_real, qtd, horas_disponiveis_por_data, horas_ja_alocadas):
    def data_fim_para_un(un_hora):
        if un_hora <= 0:
            return data_inicio + timedelta(days=999)
        duracao = qtd / un_hora
        return _calcular_data_fim_por_alocacao(
            data_inicio=data_inicio, duracao_horas=duracao,
            horas_disponiveis_por_data=horas_disponiveis_por_data,
            horas_ja_alocadas_por_data=dict(horas_ja_alocadas),
        )
    lo, hi = 1, 500000
    while lo < hi:
        mid = (lo + hi + 1) // 2
        df = data_fim_para_un(float(mid))
        if df >= data_fim_real:
            lo = mid
        else:
            hi = mid - 1
    un_hora_resultado = float(lo)
    duracao_resultado = qtd / un_hora_resultado if un_hora_resultado > 0 else 0.0
    df_check = data_fim_para_un(un_hora_resultado)
    if df_check != data_fim_real:
        return 0.0, 0.0
    return un_hora_resultado, duracao_resultado


# ─── Aplicar realizado Cogtive (CORRIGIDO v5 — cascata contínua) ──────────────

def _atualizar_lotes_com_real(rodada_id: str, registros_reais: List[Dict[str, Any]]):
    """
    Aplica o realizado Cogtive na rodada e reprograma a fila em cascata.

    Correção v47:
    - apontamento administrativo/documental não vira realizado de linha;
    - lote com produção parcial não é tratado como concluído;
    - para lote em andamento, o fim é projetado a partir do início real + ritmo
      real conservador/planejado;
    - todos os lotes posteriores da mesma linha são empurrados em cascata;
    - lotes arrastados pela fila ficam como CASCATA_REAL, sem Fim Cogtive real.
    """
    parametros = _carregar_parametros_globais()
    parametros_produto = _carregar_parametros_produto()

    rodada_res = supabase.table("f_mrp_rodadas").select("*").eq("id", rodada_id).single().execute()
    rodada = rodada_res.data or {}
    rodada_mes = _to_int(rodada.get("mes"))
    rodada_ano = _to_int(rodada.get("ano"))
    if not rodada_mes or not rodada_ano:
        raise HTTPException(status_code=404, detail="Rodada não encontrada ou sem mês/ano definidos.")

    chaves_realizadas_anteriores = _buscar_lotes_realizados_em_versoes_anteriores(rodada)

    etapas_l1 = _select_all(
        supabase.table("f_mrp_etapas")
        .select("*")
        .eq("rodada_id", rodada_id)
        .eq("recurso", "L1")
        .order("sequencia")
    )
    etapas_l2 = _select_all(
        supabase.table("f_mrp_etapas")
        .select("*")
        .eq("rodada_id", rodada_id)
        .eq("recurso", "L2")
        .order("sequencia")
    )
    alocacoes = _select_all(
        supabase.table("f_mrp_alocacoes_dia")
        .select("*")
        .eq("rodada_id", rodada_id)
    )

    def _montar_horas_disponiveis(recurso):
        mapa = {}
        for a in alocacoes:
            if _normalizar_texto(a.get("recurso")) == recurso and not a.get("lote"):
                data = a.get("data")
                h = _to_float(a.get("horas_disponiveis_dia"), 0)
                if data is not None:
                    mapa[str(data)] = h
        return mapa

    horas_disp_por_recurso = {
        "L1": _montar_horas_disponiveis("L1"),
        "L2": _montar_horas_disponiveis("L2"),
    }
    etapas_por_seq_por_recurso = {
        "L1": {(_to_int(e.get("sequencia")) or 0): e for e in etapas_l1},
        "L2": {(_to_int(e.get("sequencia")) or 0): e for e in etapas_l2},
    }
    etapas_por_recurso = {"L1": etapas_l1, "L2": etapas_l2}

    def _montar_mapa_etapas(etapas):
        mapa = {}
        for etapa in etapas:
            for chave_lote in _chaves_possiveis_lote(etapa.get("lote"), etapa.get("op")):
                mapa[(_normalizar_texto(etapa.get("recurso")), chave_lote)] = etapa
        return mapa

    etapas_por_chave = {}
    etapas_por_chave.update(_montar_mapa_etapas(etapas_l1))
    etapas_por_chave.update(_montar_mapa_etapas(etapas_l2))

    atualizadas = []
    nao_encontradas = []
    ignoradas_fora_mes = []
    ignoradas_ja_realizadas = []
    etapas_atualizadas_ids = set()
    menor_seq_impactada_por_recurso: Dict[str, int] = {}

    def _registrar_seq_impactada(recurso: str, sequencia_lote: int):
        if recurso not in {"L1", "L2"} or sequencia_lote <= 0:
            return
        if recurso not in menor_seq_impactada_por_recurso:
            menor_seq_impactada_por_recurso[recurso] = sequencia_lote
        else:
            menor_seq_impactada_por_recurso[recurso] = min(
                menor_seq_impactada_por_recurso[recurso],
                sequencia_lote,
            )

    def _data_inicio_real(real: Dict[str, Any], fallback: Optional[date]) -> Optional[date]:
        return _str_to_date(real.get("data_real_inicio")) or _str_to_date(real.get("data_real_fim")) or fallback

    def _projetar_fim_lote_em_andamento(
        etapa: Dict[str, Any],
        real: Dict[str, Any],
        data_inicio_real: date,
        qtd_base: float,
        recurso_real: str,
    ) -> Tuple[date, float, float, str]:
        """Projeta o fim do lote quando o Cogtive trouxe produção parcial."""
        un_hora_planejada = _to_float(etapa.get("un_hora"), 0)
        horas_reais = _to_float(real.get("horas_reais"), 0)
        qtd_real = _to_float(real.get("qtd_real"), 0)
        un_hora_real = (qtd_real / horas_reais) if horas_reais > 0 and qtd_real > 0 else 0.0

        # Conservador: se o real está mais lento que o plano, usa o real; se o
        # real parcial parecer mais rápido, não antecipa e mantém o planejado.
        if un_hora_planejada > 0 and un_hora_real > 0:
            nova_un_hora = min(un_hora_planejada, un_hora_real)
            origem_un_hora = "min_planejado_real_parcial"
        elif un_hora_real > 0:
            nova_un_hora = un_hora_real
            origem_un_hora = "real_parcial"
        else:
            nova_un_hora = un_hora_planejada
            origem_un_hora = "planejado_sem_qtd_real"

        if nova_un_hora <= 0:
            nova_un_hora = _to_float(etapa.get("un_hora"), 0) or 1.0

        duracao_total_horas = qtd_base / nova_un_hora if nova_un_hora > 0 else _to_float(etapa.get("duracao_horas"), 0)
        sequencia_lote = _to_int(etapa.get("sequencia")) or 0
        etapas_por_seq = etapas_por_seq_por_recurso.get(recurso_real, {})
        horas_anteriores = _montar_horas_anteriores_por_data(
            alocacoes,
            recurso_real,
            sequencia_lote,
            etapas_por_seq,
        )
        data_fim_projetada = _calcular_data_fim_por_alocacao(
            data_inicio=data_inicio_real,
            duracao_horas=duracao_total_horas,
            horas_disponiveis_por_data=horas_disp_por_recurso.get(recurso_real, {}),
            horas_ja_alocadas_por_data=dict(horas_anteriores),
        )
        return data_fim_projetada, nova_un_hora, duracao_total_horas, origem_un_hora

    registros_ordenados = sorted(
        registros_reais,
        key=lambda r: (
            r.get("recurso") or "",
            r.get("data_real_inicio") or r.get("data_real_fim") or "",
            r.get("hora_inicio") or r.get("hora_fim") or "",
        ),
    )

    # ── Fase 1: aplica somente lotes com produção real produtiva da linha ───
    for real in registros_ordenados:
        recurso_real = real.get("recurso")
        lote_real = real.get("lote") or real.get("op")
        lote_norm = _normalizar_lote_key(lote_real)
        etapa, metodo = _buscar_etapa_para_real(
            recurso_real,
            lote_norm,
            etapas_por_chave,
            etapas_atualizadas_ids,
        )

        if not etapa:
            nao_encontradas.append({
                "recurso": recurso_real,
                "lote": lote_real,
                "data_real_inicio": real.get("data_real_inicio"),
                "hora_inicio": real.get("hora_inicio"),
                "data_real_fim": real.get("data_real_fim"),
                "hora_fim": real.get("hora_fim"),
                "qtd_real": _to_float(real.get("qtd_real"), 0),
                "motivo_nao_encontrado": metodo,
            })
            continue

        if etapa.get("id") in etapas_atualizadas_ids:
            continue

        if _etapa_ja_realizada_em_versao_anterior(etapa, chaves_realizadas_anteriores):
            ignoradas_ja_realizadas.append({
                "recurso": recurso_real,
                "lote": etapa.get("lote"),
                "lote_real_cogtive": lote_real,
                "data_real_fim": real.get("data_real_fim"),
                "hora_fim": real.get("hora_fim"),
                "motivo": "Lote já atualizado pelo Cogtive em versão anterior.",
            })
            continue

        data_inicio_planejada = etapa.get("data_inicio")
        data_fim_planejada = etapa.get("data_fim")
        dt_inicio_planejada = _str_to_date(data_inicio_planejada)
        dt_fim_planejada = _str_to_date(data_fim_planejada)
        dt_inicio_real = _data_inicio_real(real, dt_inicio_planejada)
        dt_fim_apontamento = _str_to_date(real.get("data_real_fim")) or dt_fim_planejada

        if not dt_fim_planejada or dt_fim_planejada.month != rodada_mes or dt_fim_planejada.year != rodada_ano:
            ignoradas_fora_mes.append({
                "recurso": recurso_real,
                "lote": etapa.get("lote"),
                "lote_real_cogtive": lote_real,
                "data_fim_planejada": data_fim_planejada,
                "data_real_fim": real.get("data_real_fim"),
                "hora_fim": real.get("hora_fim"),
                "motivo": f"Fora do mês da rodada ({rodada_mes:02d}/{rodada_ano})",
            })
            continue

        qtd_base = _to_float(etapa.get("qtd_planejada"), 0) or _to_float(real.get("qtd_real"), 0)
        qtd_real_total = _to_float(real.get("qtd_real"), 0)
        percentual_realizado = (qtd_real_total / qtd_base * 100.0) if qtd_base > 0 else 0.0

        # Só considera lote concluído quando a quantidade produtiva real está
        # praticamente fechada. Caso contrário, o apontamento é parcial e serve
        # apenas como âncora para projetar a fila.
        lote_concluido = qtd_base <= 0 or qtd_real_total >= (qtd_base * 0.95)
        un_hora_anterior = etapa.get("un_hora")

        if lote_concluido:
            data_fim_nova_dt = dt_fim_apontamento or dt_fim_planejada
            nova_un_hora = etapa.get("un_hora")
            nova_duracao = etapa.get("duracao_horas")

            if dt_inicio_planejada and data_fim_nova_dt and qtd_base > 0 and data_fim_nova_dt != dt_fim_planejada:
                sequencia_lote = _to_int(etapa.get("sequencia")) or 0
                etapas_por_seq = etapas_por_seq_por_recurso.get(recurso_real, {})
                horas_anteriores = _montar_horas_anteriores_por_data(
                    alocacoes,
                    recurso_real,
                    sequencia_lote,
                    etapas_por_seq,
                )
                nova_un_hora, nova_duracao = _goal_seek_un_hora(
                    dt_inicio_planejada,
                    data_fim_nova_dt,
                    qtd_base,
                    horas_disp_por_recurso.get(recurso_real, {}),
                    horas_anteriores,
                )
                if nova_un_hora <= 0 or nova_duracao <= 0:
                    horas_disp = horas_disp_por_recurso.get(recurso_real, {})
                    nova_duracao = 0.0
                    cursor = dt_inicio_planejada
                    while cursor <= data_fim_nova_dt:
                        nova_duracao += horas_disp.get(cursor.isoformat(), 24)
                        cursor += timedelta(days=1)
                    nova_un_hora = qtd_base / nova_duracao if nova_duracao > 0 else etapa.get("un_hora")

            data_inicio_nova = data_inicio_planejada
            data_fim_nova_iso = data_fim_nova_dt.isoformat() if data_fim_nova_dt else data_fim_planejada
            status_novo = "realizado"
            origem_un_hora = "fim_real_cogtive"
            motivo_provavel = "lote concluído no Cogtive"
            texto_obs_real = (
                f"Atualizado pelo real Cogtive. "
                f"Data início planejada anterior: {data_inicio_planejada}. "
                f"Data fim planejada anterior: {data_fim_planejada}. "
                f"Data início real Cogtive: {real.get('data_real_inicio')} {real.get('hora_inicio') or ''}. "
                f"Último apontamento produtivo Cogtive: {real.get('data_real_fim')} {real.get('hora_fim') or ''}. "
                f"Data fim real Cogtive: {data_fim_nova_iso}. "
            )
        else:
            if not dt_inicio_real:
                dt_inicio_real = dt_inicio_planejada or dt_fim_apontamento
            data_fim_projetada, nova_un_hora, nova_duracao, origem_un_hora = _projetar_fim_lote_em_andamento(
                etapa,
                real,
                dt_inicio_real,
                qtd_base,
                recurso_real,
            )
            data_inicio_nova = dt_inicio_real.isoformat() if dt_inicio_real else data_inicio_planejada
            data_fim_nova_iso = data_fim_projetada.isoformat()
            status_novo = "em_producao_real"
            motivo_provavel = "lote em produção; fim projetado pela fila"
            texto_obs_real = (
                f"Atualizado pelo real Cogtive em andamento. "
                f"Data início planejada anterior: {data_inicio_planejada}. "
                f"Data fim planejada anterior: {data_fim_planejada}. "
                f"Data início real Cogtive: {real.get('data_real_inicio')} {real.get('hora_inicio') or ''}. "
                f"Último apontamento produtivo Cogtive: {real.get('data_real_fim')} {real.get('hora_fim') or ''}. "
                f"Data fim projetada pela fila: {data_fim_nova_iso}. "
                f"Qtd real parcial: {round(qtd_real_total, 4)}. "
                f"Qtd planejada: {round(qtd_base, 4)}. "
                f"Percentual realizado: {round(percentual_realizado, 2)}%. "
            )

        data_pa = _calcular_data_liberacao(
            data_fim_nova_iso,
            etapa.get("descricao_produto"),
            etapa.get("codigo_produto"),
            parametros,
            parametros_produto,
        )
        mes_lib, ano_lib = _resolver_mes_ano_liberacao_final(etapa, data_pa, parametros)

        update = {
            "data_inicio": data_inicio_nova,
            "data_fim": data_fim_nova_iso,
            "data_pa": data_pa,
            "qtd_planejada": qtd_base,
            "un_hora": nova_un_hora,
            "duracao_horas": nova_duracao,
            "mes_producao": (_str_to_date(data_inicio_nova).month if _str_to_date(data_inicio_nova) else etapa.get("mes_producao")),
            "ano_producao": (_str_to_date(data_inicio_nova).year if _str_to_date(data_inicio_nova) else etapa.get("ano_producao")),
            "mes_liberacao": mes_lib,
            "ano_liberacao": ano_lib,
            "status": status_novo,
            "origem": "REAL_COGTIVE",
            "observacao": (
                f"{etapa.get('observacao') or ''} | {texto_obs_real}"
                f"Hora fim real Cogtive: {real.get('hora_fim')}. "
                f"UN/HORA anterior: {un_hora_anterior}. "
                f"UN/HORA nova: {nova_un_hora}. "
                f"Origem UN/HORA: {origem_un_hora}. "
                f"Motivo provável: {motivo_provavel}. "
                f"Método casamento: {metodo}."
            ),
        }

        res = supabase.table("f_mrp_etapas").update(update).eq("id", etapa["id"]).execute()
        if not (res.data or []):
            continue

        etapa.update(update)
        etapas_atualizadas_ids.add(etapa["id"])
        sequencia_lote = _to_int(etapa.get("sequencia")) or 0
        _registrar_seq_impactada(recurso_real, sequencia_lote)

        impacto = _calcular_impacto_mudanca(data_fim_planejada, data_fim_nova_iso, un_hora_anterior, nova_un_hora)
        atualizadas.append({
            "lote": etapa.get("lote"),
            "lote_real_cogtive": lote_real,
            "codigo_produto": etapa.get("codigo_produto"),
            "descricao_produto": etapa.get("descricao_produto"),
            "recurso": recurso_real,
            "data_inicio": data_inicio_nova,
            "data_inicio_anterior": data_inicio_planejada,
            "data_fim_anterior": data_fim_planejada,
            "data_fim_nova": data_fim_nova_iso,
            "data_inicio_real_cogtive": real.get("data_real_inicio"),
            "hora_inicio_real": real.get("hora_inicio"),
            "janela_gap_inicio": data_inicio_planejada,
            "janela_gap_fim": real.get("data_real_inicio") or data_fim_nova_iso,
            "janela_gap_fim_hora": real.get("hora_inicio"),
            "data_fim_real_cogtive": real.get("data_real_fim") if lote_concluido else None,
            "hora_fim_real": real.get("hora_fim"),
            "data_lib_nova": data_pa,
            "mes_liberacao_novo": mes_lib,
            "ano_liberacao_novo": ano_lib,
            "un_hora_anterior": un_hora_anterior,
            "un_hora_nova": nova_un_hora,
            "duracao_horas_nova": nova_duracao,
            "qtd_planejada": qtd_base,
            "qtd_real": round(qtd_real_total, 4),
            "percentual_realizado": round(percentual_realizado, 2),
            "tipo_realizacao": "concluido" if lote_concluido else "parcial_em_producao",
            "motivo_provavel": motivo_provavel,
            "metodo_casamento": metodo,
            **impacto,
        })

    # ── Fase 2: cascata contínua por recurso ─────────────────────────────────
    def _rodar_cascata_unica_por_recurso(recurso: str, sequencia_inicio: int):
        etapas_linha = sorted(
            etapas_por_recurso.get(recurso, []),
            key=lambda e: _to_int(e.get("sequencia")) or 0,
        )
        horas_disp = horas_disp_por_recurso.get(recurso, {})
        horas_alocadas: Dict[str, float] = {}

        etapa_anchor = None
        for etapa in etapas_linha:
            if (_to_int(etapa.get("sequencia")) or 0) == sequencia_inicio:
                etapa_anchor = etapa
                break

        if not etapa_anchor:
            return []

        dt_fim_anchor = _str_to_date(etapa_anchor.get("data_fim"))
        if not dt_fim_anchor:
            return []

        obs_anchor = etapa_anchor.get("observacao") or ""
        anchor_gap_inicio = (
            _extrair_valor_observacao(obs_anchor, "Data início planejada anterior")
            or _extrair_valor_observacao(obs_anchor, "Data início anterior")
            or etapa_anchor.get("data_inicio")
        )
        anchor_inicio_real_txt = _extrair_valor_observacao(obs_anchor, "Data início real Cogtive")
        anchor_inicio_real_data, anchor_inicio_real_hora = _extrair_data_hora_texto(anchor_inicio_real_txt)
        anchor_gap_fim = anchor_inicio_real_data or etapa_anchor.get("data_inicio") or etapa_anchor.get("data_fim")
        anchor_gap_fim_hora = anchor_inicio_real_hora
        anchor_lote = etapa_anchor.get("lote")

        alocacoes_originais_por_lote: Dict[str, List[Tuple[str, float]]] = {}
        for aloc in alocacoes:
            if _normalizar_texto(aloc.get("recurso")) != recurso:
                continue
            lote_key = _normalizar_lote_key(aloc.get("lote") or "")
            if not lote_key:
                continue
            data_aloc = str(aloc.get("data") or "")
            horas_aloc = _to_float(aloc.get("horas_alocadas"), 0)
            if data_aloc and horas_aloc > 0:
                alocacoes_originais_por_lote.setdefault(lote_key, []).append((data_aloc, horas_aloc))

        for lote_key in list(alocacoes_originais_por_lote.keys()):
            alocacoes_originais_por_lote[lote_key].sort(key=lambda x: x[0])

        def _capacidade_dia(data_ref: date) -> float:
            data_str = data_ref.isoformat()
            if data_str in horas_disp:
                return _to_float(horas_disp.get(data_str), 0)
            if data_ref.weekday() < 5:
                return 21.0
            return 0.0

        def _ocupar_horas_em_matriz(data_inicio_base: date, horas_por_bloco: List[float]) -> date:
            cursor = data_inicio_base
            ultima_data = data_inicio_base
            limite = data_inicio_base + timedelta(days=400)

            for horas_bloco in horas_por_bloco:
                restante = _to_float(horas_bloco, 0)
                if restante <= 0:
                    continue

                while restante > 0.001 and cursor <= limite:
                    data_str = cursor.isoformat()
                    capacidade = _capacidade_dia(cursor)
                    ja_alocado = horas_alocadas.get(data_str, 0)
                    saldo = max(0, capacidade - ja_alocado)

                    if saldo > 0:
                        usado = min(restante, saldo)
                        horas_alocadas[data_str] = ja_alocado + usado
                        restante -= usado
                        ultima_data = cursor

                    if restante > 0.001:
                        cursor += timedelta(days=1)

            return ultima_data

        def _ocupar_etapa_real_ou_anchor(etapa: Dict[str, Any]) -> Optional[date]:
            dt_ini = _str_to_date(etapa.get("data_inicio"))
            qtd = _to_float(etapa.get("qtd_planejada"), 0)
            un_hora = _to_float(etapa.get("un_hora"), 0)
            if not dt_ini or qtd <= 0 or un_hora <= 0:
                return _str_to_date(etapa.get("data_fim"))
            duracao = qtd / un_hora
            return _calcular_data_fim_por_alocacao(
                data_inicio=dt_ini,
                duracao_horas=duracao,
                horas_disponiveis_por_data=horas_disp,
                horas_ja_alocadas_por_data=horas_alocadas,
            )

        # Antes do anchor: preserva a ocupação original dos lotes anteriores.
        lotes_antes = set()
        for etapa in etapas_linha:
            seq = _to_int(etapa.get("sequencia")) or 0
            if seq >= sequencia_inicio:
                break
            lote_key = _normalizar_lote_key(etapa.get("lote") or "")
            if lote_key:
                lotes_antes.add(lote_key)

        for aloc in alocacoes:
            if _normalizar_texto(aloc.get("recurso")) != recurso:
                continue
            lote_aloc = _normalizar_lote_key(aloc.get("lote") or "")
            if not lote_aloc or lote_aloc not in lotes_antes:
                continue
            data = str(aloc.get("data") or "")
            h = _to_float(aloc.get("horas_alocadas"), 0)
            if data and h > 0:
                horas_alocadas[data] = horas_alocadas.get(data, 0) + h

        # Anchor real/parcial: ocupa a matriz com a data já recalculada.
        dt_fim_anchor_ocupado = _ocupar_etapa_real_ou_anchor(etapa_anchor)
        if dt_fim_anchor_ocupado:
            dt_fim_anchor = dt_fim_anchor_ocupado

        data_inicio_cursor = dt_fim_anchor
        cascatas = []

        for etapa in etapas_linha:
            seq = _to_int(etapa.get("sequencia")) or 0
            if seq <= sequencia_inicio:
                continue

            # Se outro lote da mesma linha também veio com produção real, ele já
            # foi recalculado na fase 1. Apenas ocupa a matriz e segue a fila.
            if etapa.get("id") in etapas_atualizadas_ids:
                dt_fim_realizado = _ocupar_etapa_real_ou_anchor(etapa)
                if dt_fim_realizado:
                    data_inicio_cursor = dt_fim_realizado
                continue

            data_inicio_anterior = etapa.get("data_inicio")
            data_fim_anterior = etapa.get("data_fim")
            dt_inicio_atual = _str_to_date(data_inicio_anterior)

            # Regra de cascata: nunca antecipa. Se o original estava depois do
            # cursor, mantém; se o cursor passou do original, empurra.
            if dt_inicio_atual and dt_inicio_atual > data_inicio_cursor:
                novo_inicio = dt_inicio_atual
            else:
                novo_inicio = data_inicio_cursor

            lote_key = _normalizar_lote_key(etapa.get("lote") or "")
            alocs_lote = alocacoes_originais_por_lote.get(lote_key, [])
            qtd = _to_float(etapa.get("qtd_planejada"), 0)
            un_hora = _to_float(etapa.get("un_hora"), 0)

            if alocs_lote:
                horas_blocos = [h for _, h in alocs_lote if _to_float(h, 0) > 0]
                duracao_horas = sum(_to_float(h, 0) for h in horas_blocos)
                data_fim_nova = _ocupar_horas_em_matriz(novo_inicio, horas_blocos)
            elif qtd <= 0 or un_hora <= 0:
                dt_fim_atual = _str_to_date(data_fim_anterior)
                if dt_inicio_atual and dt_fim_atual and dt_fim_atual >= dt_inicio_atual:
                    duracao_original_dias = (dt_fim_atual - dt_inicio_atual).days
                    data_fim_nova = novo_inicio + timedelta(days=duracao_original_dias)
                    data_fim_nova = _ajustar_para_dia_util(data_fim_nova)
                else:
                    data_fim_nova = novo_inicio
                duracao_horas = _to_float(etapa.get("duracao_horas"), 0)
            else:
                duracao_horas = qtd / un_hora
                data_fim_nova = _calcular_data_fim_por_alocacao(
                    data_inicio=novo_inicio,
                    duracao_horas=duracao_horas,
                    horas_disponiveis_por_data=horas_disp,
                    horas_ja_alocadas_por_data=horas_alocadas,
                )

            data_pa = _calcular_data_liberacao(
                data_fim_nova.isoformat(),
                etapa.get("descricao_produto"),
                etapa.get("codigo_produto"),
                parametros,
                parametros_produto,
            )
            mes_lib_final, ano_lib_final = _resolver_mes_ano_liberacao_final(etapa, data_pa, parametros)

            update = {
                "data_inicio": novo_inicio.isoformat(),
                "data_fim": data_fim_nova.isoformat(),
                "data_pa": data_pa,
                "duracao_horas": duracao_horas,
                "mes_producao": novo_inicio.month,
                "ano_producao": novo_inicio.year,
                "mes_liberacao": mes_lib_final,
                "ano_liberacao": ano_lib_final,
                "status": "cascata",
                "origem": "CASCATA_REAL",
                "observacao": (
                    f"{etapa.get('observacao') or ''} | "
                    f"Cascata por real Cogtive. "
                    f"Data início anterior: {data_inicio_anterior}. "
                    f"Data fim planejada anterior: {data_fim_anterior}. "
                    f"Data início recalculada cascata: {novo_inicio.isoformat()}. "
                    f"Data fim recalculada cascata: {data_fim_nova.isoformat()}. "
                    f"Lote âncora cascata: {anchor_lote}. "
                    f"Janela gap início: {anchor_gap_inicio}. "
                    f"Janela gap fim: {anchor_gap_fim} {anchor_gap_fim_hora or ''}. "
                    f"Motivo provável: arrastado pela fila da {recurso}. "
                    f"Método casamento: cascata_real."
                ),
            }

            supabase.table("f_mrp_etapas").update(update).eq("id", etapa["id"]).execute()
            etapa.update(update)

            impacto = _calcular_impacto_mudanca(data_fim_anterior, data_fim_nova.isoformat(), etapa.get("un_hora"), etapa.get("un_hora"))
            cascatas.append({
                "etapa_id": etapa["id"],
                "lote": etapa.get("lote"),
                "recurso": recurso,
                "data_inicio_anterior": data_inicio_anterior,
                "data_inicio_nova": novo_inicio.isoformat(),
                "data_fim_anterior": data_fim_anterior,
                "data_fim_nova": data_fim_nova.isoformat(),
                "janela_gap_inicio": anchor_gap_inicio,
                "janela_gap_fim": anchor_gap_fim,
                "janela_gap_fim_hora": anchor_gap_fim_hora,
                "lote_ancora_cascata": anchor_lote,
                "data_lib_nova": data_pa,
                "mes_liberacao": mes_lib_final,
                "ano_liberacao": ano_lib_final,
                "motivo_provavel": "arrastado pela fila",
                **impacto,
            })

            data_inicio_cursor = data_fim_nova

        return cascatas

    cascatas_aplicadas = []
    for recurso, seq_inicio in menor_seq_impactada_por_recurso.items():
        cascatas_aplicadas.extend(_rodar_cascata_unica_por_recurso(recurso, seq_inicio))

    return {
        "atualizadas": atualizadas,
        "nao_encontradas": nao_encontradas,
        "ignoradas_fora_mes": ignoradas_fora_mes,
        "ignoradas_ja_realizadas": ignoradas_ja_realizadas,
        "cascatas_aplicadas": cascatas_aplicadas,
    }

def _extrair_valor_observacao(obs: str, chave: str):
    if not obs:
        return None
    pattern = rf"{re.escape(chave)}:\s*([^\.\|]+)"
    match = re.search(pattern, obs, flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1).strip() or None


def _montar_mudanca_realizado_da_etapa(etapa: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    obs = etapa.get("observacao") or ""
    origem = _normalizar_texto(etapa.get("origem"))

    eh_real = "Atualizado pelo real Cogtive" in obs or origem == "REAL_COGTIVE"
    eh_cascata = "Cascata por real Cogtive" in obs or origem == "CASCATA_REAL"

    if not eh_real and not eh_cascata:
        return None

    data_inicio_anterior = _extrair_valor_observacao(obs, "Data início anterior") or _extrair_valor_observacao(obs, "Data início planejada anterior")
    data_fim_anterior = _extrair_valor_observacao(obs, "Data fim planejada anterior")
    data_inicio_real_txt = _extrair_valor_observacao(obs, "Data início real Cogtive")
    data_inicio_real_cogtive, hora_inicio_real = _extrair_data_hora_texto(data_inicio_real_txt)
    janela_gap_inicio = _extrair_valor_observacao(obs, "Janela gap início")
    janela_gap_fim_txt = _extrair_valor_observacao(obs, "Janela gap fim")
    janela_gap_fim, janela_gap_fim_hora = _extrair_data_hora_texto(janela_gap_fim_txt)
    lote_ancora_cascata = _extrair_valor_observacao(obs, "Lote âncora cascata")

    if eh_cascata:
        data_fim_nova = _extrair_valor_observacao(obs, "Data fim recalculada cascata") or etapa.get("data_fim")
        data_inicio_nova = _extrair_valor_observacao(obs, "Data início recalculada cascata") or etapa.get("data_inicio")
        hora_fim_real = None
        motivo = _extrair_valor_observacao(obs, "Motivo provável") or "arrastado pela fila"
        metodo = _extrair_valor_observacao(obs, "Método casamento") or "cascata_real"
        tipo_realizacao = "cascata"
        lote_real_cogtive = None
    else:
        data_fim_nova = (
            _extrair_valor_observacao(obs, "Data fim projetada pela fila")
            or _extrair_valor_observacao(obs, "Data fim real Cogtive")
            or etapa.get("data_fim")
        )
        data_inicio_nova = etapa.get("data_inicio")
        hora_fim_real = _extrair_valor_observacao(obs, "Hora fim real Cogtive")
        motivo = _extrair_valor_observacao(obs, "Motivo provável") or "não identificado"
        metodo = _extrair_valor_observacao(obs, "Método casamento")
        tipo_realizacao = "parcial_em_producao" if "em andamento" in obs else "concluido"
        lote_real_cogtive = etapa.get("lote")

    un_hora_anterior_txt = _extrair_valor_observacao(obs, "UN/HORA anterior")
    un_hora_nova_txt = _extrair_valor_observacao(obs, "UN/HORA nova")
    un_hora_anterior = _to_float(un_hora_anterior_txt, None) if un_hora_anterior_txt is not None else None
    un_hora_nova = _to_float(un_hora_nova_txt, None) if un_hora_nova_txt is not None else etapa.get("un_hora")
    impacto = _calcular_impacto_mudanca(data_fim_anterior, data_fim_nova, un_hora_anterior, un_hora_nova)

    return {
        "lote": etapa.get("lote"),
        "lote_real_cogtive": lote_real_cogtive,
        "codigo_produto": etapa.get("codigo_produto"),
        "descricao_produto": etapa.get("descricao_produto"),
        "recurso": etapa.get("recurso"),
        "data_inicio": data_inicio_nova,
        "data_inicio_anterior": data_inicio_anterior,
        "data_fim_anterior": data_fim_anterior,
        "data_fim_nova": data_fim_nova,
        "data_inicio_real_cogtive": data_inicio_real_cogtive,
        "hora_inicio_real": hora_inicio_real,
        "janela_gap_inicio": janela_gap_inicio or data_inicio_anterior,
        "janela_gap_fim": janela_gap_fim or data_inicio_real_cogtive or data_fim_nova,
        "janela_gap_fim_hora": janela_gap_fim_hora or hora_inicio_real,
        "lote_ancora_cascata": lote_ancora_cascata,
        "hora_fim_real": hora_fim_real,
        "data_lib_nova": etapa.get("data_pa"),
        "mes_liberacao_novo": etapa.get("mes_liberacao"),
        "ano_liberacao_novo": etapa.get("ano_liberacao"),
        "un_hora_anterior": un_hora_anterior,
        "un_hora_nova": un_hora_nova,
        "duracao_horas_nova": etapa.get("duracao_horas"),
        "qtd_planejada": etapa.get("qtd_planejada"),
        "motivo_provavel": motivo,
        "metodo_casamento": metodo,
        "tipo_realizacao": tipo_realizacao,
        **impacto,
    }

# ─── Comparativo de liberação ─────────────────────────────────────────────────

def _agrupar_liberacao_por_mes(etapas: List[Dict[str, Any]], ano_sentinela: int = 2017) -> Dict[Tuple[int, int], Dict[str, Any]]:
    grupos: Dict[Tuple[int, int], Dict[str, Any]] = {}
    for etapa in etapas:
        mes_lib = _to_int(etapa.get("mes_liberacao"))
        ano_lib = _to_int(etapa.get("ano_liberacao"))
        if not mes_lib or not ano_lib or ano_lib == ano_sentinela:
            continue
        chave = (ano_lib, mes_lib)
        qtd = _to_float(etapa.get("qtd_planejada"), 0)
        if chave not in grupos:
            grupos[chave] = {"mes_liberacao": mes_lib, "ano_liberacao": ano_lib, "qtd_tubetes": 0, "caixas": 0, "lotes": 0}
        grupos[chave]["qtd_tubetes"] += qtd
        grupos[chave]["caixas"] = grupos[chave]["qtd_tubetes"] / 500
        grupos[chave]["lotes"] += 1
    return grupos


def _buscar_rodada_anterior(rodada: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    mes = rodada.get("mes")
    ano = rodada.get("ano")
    versao = _to_int(rodada.get("versao")) or 0
    if not mes or not ano or versao <= 1:
        return None
    rodadas = _select_all(supabase.table("f_mrp_rodadas").select("*").eq("mes", mes).eq("ano", ano))
    candidatas = [r for r in rodadas if (_to_int(r.get("versao")) or 0) < versao]
    if not candidatas:
        return None
    candidatas.sort(key=lambda r: _to_int(r.get("versao")) or 0, reverse=True)
    return candidatas[0]


def _montar_comparativo_liberacao(rodada_atual, rodada_anterior, etapas_atual, etapas_anterior):
    grupos_atual = _agrupar_liberacao_por_mes(etapas_atual)
    grupos_anterior = _agrupar_liberacao_por_mes(etapas_anterior)
    chaves = sorted(set(grupos_atual.keys()) | set(grupos_anterior.keys()))
    linhas = []
    for ano_lib, mes_lib in chaves:
        atual = grupos_atual.get((ano_lib, mes_lib), {"qtd_tubetes": 0, "caixas": 0, "lotes": 0})
        anterior = grupos_anterior.get((ano_lib, mes_lib), {"qtd_tubetes": 0, "caixas": 0, "lotes": 0})
        qtd_ant = _to_float(anterior.get("qtd_tubetes"), 0)
        qtd_at = _to_float(atual.get("qtd_tubetes"), 0)
        dif_tb = qtd_at - qtd_ant
        linhas.append({
            "mes_liberacao": mes_lib, "ano_liberacao": ano_lib,
            "qtd_tubetes_anterior": qtd_ant, "caixas_anterior": qtd_ant / 500,
            "qtd_tubetes_atual": qtd_at, "caixas_atual": qtd_at / 500,
            "dif_tubetes": dif_tb, "dif_caixas": dif_tb / 500,
            "variacao_pct": (dif_tb / qtd_ant * 100) if qtd_ant else None,
            "lotes_anterior": anterior.get("lotes", 0), "lotes_atual": atual.get("lotes", 0),
        })
    total_ant = sum(_to_float(l.get("qtd_tubetes_anterior"), 0) for l in linhas)
    total_at = sum(_to_float(l.get("qtd_tubetes_atual"), 0) for l in linhas)
    dif_total = total_at - total_ant
    return {
        "rodada_atual": rodada_atual, "rodada_anterior": rodada_anterior,
        "tem_rodada_anterior": rodada_anterior is not None,
        "total_qtd_tubetes_anterior": total_ant, "total_caixas_anterior": total_ant / 500,
        "total_qtd_tubetes_atual": total_at, "total_caixas_atual": total_at / 500,
        "dif_total_tubetes": dif_total, "dif_total_caixas": dif_total / 500,
        "linhas": linhas,
    }



def _processar_calendario_gantt_background(rodada_id: str, excel_bytes: bytes):
    """
    Processa comentários/calendário do Gantt fora do request principal.
    Assim o upload MPS volta a responder rápido e o proxy do Fly não derruba.
    """
    abas = {
        "PROGRAMAÇÃO MACRO - LINHA 1": {"recurso": "L1"},
        "PROGRAMAÇÃO MACRO - LINHA 2": {"recurso": "L2"},
        "PROGRAMAÇÃO MACRO - FABRIMA": {"recurso": "FABRIMA"},
    }

    wb_comentarios = None
    try:
        wb_comentarios = load_workbook(
            BytesIO(excel_bytes),
            data_only=False,
            read_only=False,
            keep_links=False,
        )

        calendario_dia = []

        for aba, config in abas.items():
            if aba not in wb_comentarios.sheetnames:
                continue
            dados_calendario = _ler_calendario_dia_gantt(
                excel_bytes,
                aba,
                config["recurso"],
                rodada_id,
                wb_full=wb_comentarios,
            )
            calendario_dia.extend(dados_calendario)

        try:
            supabase.table("f_mrp_calendario_dia").delete().eq("rodada_id", rodada_id).execute()
        except Exception:
            pass

        if calendario_dia:
            _insert_em_lotes("f_mrp_calendario_dia", calendario_dia)

    except Exception as e:
        print(f"[MRP][calendario_gantt_background] erro rodada={rodada_id}: {str(e)[:500]}")

    finally:
        if wb_comentarios:
            try:
                wb_comentarios.close()
            except Exception:
                pass



# ─── Hotfix importação MPS v8 — header dinâmico e validação pós-gravação ──────

def _sheet_norm(value):
    return _normalizar_texto(value).replace(" ", "")


def _encontrar_aba_mps(xls, nomes_possiveis):
    """
    Localiza a aba mesmo quando o Excel vem com pequena variação no nome.
    Ex.: "PROGRAMAÇÃO MACRO - LINHA 1", "PROG MACRO LINHA 1", etc.
    """
    if not xls or not getattr(xls, "sheet_names", None):
        return None

    nomes_norm = [_sheet_norm(n) for n in nomes_possiveis]
    for sheet in xls.sheet_names:
        sheet_norm = _sheet_norm(sheet)
        if sheet_norm in nomes_norm:
            return sheet

    tokens_por_nome = []
    for nome in nomes_possiveis:
        tokens = [t for t in re.split(r"[^A-Z0-9]+", _normalizar_texto(nome)) if t]
        tokens_por_nome.append(tokens)

    for sheet in xls.sheet_names:
        sheet_text = _normalizar_texto(sheet)
        for tokens in tokens_por_nome:
            if tokens and all(t in sheet_text for t in tokens):
                return sheet

    return None


def _detectar_linha_header_mps(excel_bytes, sheet_name):
    """
    Detecta a linha de cabeçalho real do MPS.

    O parser antigo usava header fixo na linha 6 (header=5). Quando o arquivo
    de julho veio com deslocamento/layout diferente, o endpoint retornava 200,
    mas gravava a rodada sem volume útil. Aqui procuramos a linha que contém os
    campos fixos do Gantt/MPS.
    """
    try:
        raw = pd.read_excel(
            BytesIO(excel_bytes),
            sheet_name=sheet_name,
            header=None,
            engine="openpyxl",
            nrows=80,
        )
    except Exception:
        return 5

    melhores = []
    for idx in range(raw.shape[0]):
        valores = [_normalizar_coluna(v) for v in raw.iloc[idx].tolist() if _to_str(v)]
        valores_set = set(valores)
        texto = " | ".join(valores)

        score = 0
        if "LOTE" in valores_set:
            score += 3
        if "PRODUTO" in valores_set:
            score += 3
        if "CODIGO" in valores_set or "CÓDIGO" in valores_set:
            score += 3
        if any("QTD" in v and ("TUBETE" in v or v == "QTD") for v in valores):
            score += 2
        if any("DATA INICIO" in v or "DATA INÍCIO" in v for v in valores):
            score += 2
        if any("DATA FIM" in v for v in valores):
            score += 2
        if any("MES LIB" in v or "MÊS LIB" in v for v in valores):
            score += 1
        if any("ANO LIB" in v for v in valores):
            score += 1

        # Header antigo e esperado fica com score alto. Aceita 7+ para tolerar
        # coluna de código/produto ligeiramente diferente.
        if score >= 7:
            melhores.append((score, idx, texto[:300]))

    if melhores:
        melhores.sort(reverse=True)
        return melhores[0][1]

    return 5


def _ler_df_mps_dinamico(excel_bytes, sheet_name):
    header_idx = _detectar_linha_header_mps(excel_bytes, sheet_name)
    df = pd.read_excel(
        BytesIO(excel_bytes),
        sheet_name=sheet_name,
        header=header_idx,
        engine="openpyxl",
    )
    # Remove colunas completamente vazias/Unnamed sem conteúdo.
    df = df.dropna(axis=1, how="all")
    return df, header_idx


def _mapear_mes_lib_manual_por_linha(excel_bytes, sheet_name, header_idx=None):
    mapa = {}
    wb = None
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=False, read_only=True)
        if sheet_name not in wb.sheetnames:
            return mapa
        ws = wb[sheet_name]
        header_row = int(header_idx if header_idx is not None else _detectar_linha_header_mps(excel_bytes, sheet_name)) + 1
        col_mes_lib = None

        rows = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))
        if rows:
            for cell in rows[0]:
                header_norm = _normalizar_coluna(cell.value)
                if header_norm in [
                    _normalizar_coluna("MÊS LIBERAÇÃO"),
                    _normalizar_coluna("MES LIBERAÇÃO"),
                    _normalizar_coluna("MES LIBERACAO"),
                    _normalizar_coluna("MÊS LIB."),
                    _normalizar_coluna("MES LIB."),
                ]:
                    col_mes_lib = cell.column
                    break
        if not col_mes_lib:
            return mapa

        for row in ws.iter_rows(min_row=header_row + 1, min_col=col_mes_lib, max_col=col_mes_lib, values_only=False):
            cell = row[0]
            valor = cell.value
            if valor is None or str(valor).strip() == "":
                continue
            eh_formula = (isinstance(valor, str) and valor.strip().startswith("="))
            mapa[cell.row] = not eh_formula
    except Exception:
        pass
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
    return mapa


def _ler_header_dias(excel_bytes, sheet_name, header_idx=None):
    raw = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=None, engine="openpyxl")
    header_idx = int(header_idx if header_idx is not None else _detectar_linha_header_mps(excel_bytes, sheet_name))
    data_row_idx = max(0, header_idx - 1)
    capacidade_row_idx = header_idx

    datas_por_coluna = {}
    capacidades_por_coluna = {}
    for col_idx in range(raw.shape[1]):
        data_ref = _parse_data(raw.iat[data_row_idx, col_idx]) if raw.shape[0] > data_row_idx else None

        # Fallback: alguns arquivos trazem a própria data como nome da coluna no
        # cabeçalho, não na linha anterior.
        if not data_ref and raw.shape[0] > header_idx:
            data_ref = _parse_data(raw.iat[header_idx, col_idx])

        if data_ref:
            datas_por_coluna[col_idx] = data_ref
            capacidade = _to_float(raw.iat[capacidade_row_idx, col_idx], 0) if raw.shape[0] > capacidade_row_idx else 0
            capacidades_por_coluna[col_idx] = capacidade

    return datas_por_coluna, capacidades_por_coluna


def _coluna_data_flexivel(df, possibilidades):
    col = _buscar_coluna(df, possibilidades)
    if col:
        return col

    # Fallback por contém, para casos "GANTT - INÍCIO", "INÍCIO", etc.
    possibilidades_norm = [_normalizar_coluna(p) for p in possibilidades]
    for c in df.columns:
        c_norm = _normalizar_coluna(c)
        if any(p in c_norm for p in possibilidades_norm):
            return c
    return None


def _ler_aba_mps(excel_bytes, sheet_name, recurso, etapa):
    df, header_idx = _ler_df_mps_dinamico(excel_bytes, sheet_name)
    mes_lib_manual_por_linha = _mapear_mes_lib_manual_por_linha(excel_bytes, sheet_name, header_idx)
    comentarios_gantt_por_linha = _mapear_comentarios_gantt_por_linha(excel_bytes, sheet_name)
    df = df.dropna(how="all")

    col_embalado = _buscar_coluna(df, ["EMBALADO", "EMBALADO?"])
    col_lote = _buscar_coluna(df, ["LOTE", "LOTE/OP", "LOTE OP"])
    col_codigo = _buscar_coluna(df, ["CÓDIGO", "CODIGO", "COD.", "COD"])
    col_produto = _buscar_coluna(df, ["PRODUTO", "DESCRIÇÃO PRODUTO", "DESCRICAO PRODUTO"])
    col_tempo = _buscar_coluna(df, ["TEMPO (Horas.)", "TEMPO (H)", "TEMPO HORAS", "TEMPO", "HORAS"])
    col_un_hora = _buscar_coluna(df, ["UN / HORA", "UN/HORA", "UN HORA", "UN. HORA"])
    col_qtd = _buscar_coluna(df, ["QTD. (Tubetes)", "QTD (Tubetes)", "QTD Tubetes", "QTD", "QTD TUBETES", "QUANTIDADE"])
    col_mes_prod = _buscar_coluna(df, ["MÊS PRODUÇÃO", "MES PRODUÇÃO", "MES PRODUCAO", "MÊS PROD.", "MES PROD."])
    col_ano_prod = _buscar_coluna(df, ["ANO PRODUÇÃO", "ANO PRODUCAO", "ANO PROD."])
    col_data_inicio = _coluna_data_flexivel(df, ["DATA INÍCIO", "DATA INICIO", "INÍCIO", "INICIO"])
    col_data_fim = _coluna_data_flexivel(df, ["DATA FIM", "FIM"])
    col_data_lib = _buscar_coluna(df, ["DATA LIB.", "DATA LIB", "DATA LIBERAÇÃO", "DATA LIBERACAO", "LIBERAÇÃO", "LIBERACAO"])
    col_mes_lib = _buscar_coluna(df, ["MÊS LIBERAÇÃO", "MES LIBERAÇÃO", "MES LIBERACAO", "MÊS LIB.", "MES LIB."])
    col_ano_lib = _buscar_coluna(df, ["ANO LIBERAÇÃO", "ANO LIBERACAO", "ANO LIB."])

    registros = []
    linhas_descartadas = 0

    for idx_df, row in df.iterrows():
        try:
            excel_row = int(idx_df) + header_idx + 2
        except Exception:
            excel_row = None

        lote = _to_str(row.get(col_lote)) if col_lote else None
        codigo = _to_str(row.get(col_codigo)) if col_codigo else None
        produto = _to_str(row.get(col_produto)) if col_produto else None

        if not lote and not codigo and not produto:
            continue

        produto_norm = _normalizar_texto(produto)
        codigo_norm = _normalizar_texto(codigo)
        lote_norm = _normalizar_texto(lote)

        if produto_norm in ["PRODUTO", "NAN", "TOTAL", "TOTAIS"] or codigo_norm in ["CODIGO", "CÓDIGO", "TOTAL", "TOTAIS"]:
            continue
        if "TOTAL" in produto_norm and not codigo:
            continue

        ano_prod = _to_int(row.get(col_ano_prod)) if col_ano_prod else None
        ano_lib = _to_int(row.get(col_ano_lib)) if col_ano_lib else ano_prod

        data_inicio = _parse_data(row.get(col_data_inicio), ano_prod) if col_data_inicio else None
        data_fim = _parse_data(row.get(col_data_fim), ano_prod) if col_data_fim else None
        data_pa = _parse_data(row.get(col_data_lib), ano_lib) if col_data_lib else None

        # Mantém a regra: etapa sem início/fim não entra como linha produtiva.
        if not data_inicio and not data_fim:
            linhas_descartadas += 1
            continue

        registros.append({
            "lote": lote,
            "op": lote,
            "codigo_produto": codigo,
            "descricao_produto": produto,
            "etapa": etapa,
            "recurso": recurso,
            "linha_origem": recurso,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "data_pa": data_pa,
            "qtd_planejada": _to_float(row.get(col_qtd), 0) if col_qtd else 0,
            "duracao_horas": _to_float(row.get(col_tempo), 0) if col_tempo else 0,
            "sequencia": len(registros) + 1,
            "status": "planejada",
            "origem": "IMPORT_MPS",
            "observacao": (
                f"Importado da aba {sheet_name}"
                + (
                    " | Comentários Gantt: "
                    + " ; ".join(comentarios_gantt_por_linha.get(excel_row, [])[:20])
                    if excel_row and comentarios_gantt_por_linha.get(excel_row)
                    else ""
                )
            ),
            "embalado": _to_str(row.get(col_embalado)) if col_embalado else None,
            "un_hora": _to_float(row.get(col_un_hora), 0) if col_un_hora else 0,
            "mes_producao": _to_int(row.get(col_mes_prod)) if col_mes_prod else None,
            "ano_producao": ano_prod,
            "mes_liberacao": _to_int(row.get(col_mes_lib)) if col_mes_lib else None,
            "ano_liberacao": ano_lib,
            "mes_lib_manual": mes_lib_manual_por_linha.get(excel_row) if excel_row else None,
        })

    # Debug leve anexado como atributo para o endpoint devolver sem quebrar a assinatura antiga.
    try:
        _ler_aba_mps._ultimo_debug = {
            "aba": sheet_name,
            "header_idx_zero_based": header_idx,
            "header_excel_row": header_idx + 1,
            "colunas_detectadas": {
                "lote": str(col_lote) if col_lote is not None else None,
                "codigo": str(col_codigo) if col_codigo is not None else None,
                "produto": str(col_produto) if col_produto is not None else None,
                "qtd": str(col_qtd) if col_qtd is not None else None,
                "data_inicio": str(col_data_inicio) if col_data_inicio is not None else None,
                "data_fim": str(col_data_fim) if col_data_fim is not None else None,
                "data_lib": str(col_data_lib) if col_data_lib is not None else None,
                "mes_lib": str(col_mes_lib) if col_mes_lib is not None else None,
                "ano_lib": str(col_ano_lib) if col_ano_lib is not None else None,
            },
            "linhas_dataframe": int(len(df)),
            "linhas_validas": int(len(registros)),
            "linhas_descartadas_sem_data": int(linhas_descartadas),
        }
    except Exception:
        pass

    return registros


def _ler_alocacoes_dia(excel_bytes, sheet_name, recurso, rodada_id):
    df, header_idx = _ler_df_mps_dinamico(excel_bytes, sheet_name)
    df = df.dropna(how="all")
    datas_header, capacidades_header = _ler_header_dias(excel_bytes, sheet_name, header_idx)

    col_lote = _buscar_coluna(df, ["LOTE", "LOTE/OP", "LOTE OP"])
    col_codigo = _buscar_coluna(df, ["CÓDIGO", "CODIGO", "COD.", "COD"])
    col_produto = _buscar_coluna(df, ["PRODUTO", "DESCRIÇÃO PRODUTO", "DESCRICAO PRODUTO"])

    colunas_fixas = {
        _buscar_coluna(df, ["EMBALADO", "EMBALADO?"]),
        col_lote,
        col_codigo,
        col_produto,
        _buscar_coluna(df, ["TEMPO (Horas.)", "TEMPO (H)", "TEMPO HORAS", "TEMPO", "HORAS"]),
        _buscar_coluna(df, ["UN / HORA", "UN/HORA", "UN HORA", "UN. HORA"]),
        _buscar_coluna(df, ["QTD. (Tubetes)", "QTD (Tubetes)", "QTD Tubetes", "QTD", "QTD TUBETES", "QUANTIDADE"]),
        _buscar_coluna(df, ["MÊS PRODUÇÃO", "MES PRODUÇÃO", "MES PRODUCAO", "MÊS PROD.", "MES PROD."]),
        _buscar_coluna(df, ["ANO PRODUÇÃO", "ANO PRODUCAO", "ANO PROD."]),
        _coluna_data_flexivel(df, ["DATA INÍCIO", "DATA INICIO", "INÍCIO", "INICIO"]),
        _coluna_data_flexivel(df, ["DATA FIM", "FIM"]),
        _buscar_coluna(df, ["DATA LIB.", "DATA LIB", "DATA LIBERAÇÃO", "DATA LIBERACAO"]),
        _buscar_coluna(df, ["MÊS LIBERAÇÃO", "MES LIBERAÇÃO", "MES LIBERACAO", "MÊS LIB.", "MES LIB."]),
        _buscar_coluna(df, ["ANO LIBERAÇÃO", "ANO LIBERACAO", "ANO LIB."]),
    }
    colunas_fixas = {c for c in colunas_fixas if c is not None}

    colunas_datas = []
    for idx, col in enumerate(df.columns):
        if col in colunas_fixas:
            continue
        data_ref = _parse_data(col) if _eh_coluna_data(col) else None
        if not data_ref and idx in datas_header:
            data_ref = datas_header[idx]
        if data_ref:
            colunas_datas.append((idx, col, data_ref))

    registros = []

    # Registros de capacidade do calendário.
    for idx, _, data_ref in colunas_datas:
        capacidade = capacidades_header.get(idx, 0)
        registros.append({
            "rodada_id": rodada_id,
            "recurso": recurso,
            "lote": None,
            "codigo_produto": None,
            "descricao_produto": None,
            "data": data_ref,
            "horas_alocadas": 0,
            "horas_disponiveis_dia": capacidade,
            "origem": "CAPACIDADE_MPS",
        })

    # Alocações por lote/dia.
    for _, row in df.iterrows():
        lote = _to_str(row.get(col_lote)) if col_lote else None
        codigo = _to_str(row.get(col_codigo)) if col_codigo else None
        produto = _to_str(row.get(col_produto)) if col_produto else None
        if not lote and not codigo and not produto:
            continue

        produto_norm = _normalizar_texto(produto)
        if produto_norm in ["PRODUTO", "TOTAL", "TOTAIS"]:
            continue

        for idx, col_data, data_ref in colunas_datas:
            horas = _to_float(row.get(col_data), 0)
            if horas <= 0:
                continue
            registros.append({
                "rodada_id": rodada_id,
                "recurso": recurso,
                "lote": lote,
                "codigo_produto": codigo,
                "descricao_produto": produto,
                "data": data_ref,
                "horas_alocadas": horas,
                "horas_disponiveis_dia": capacidades_header.get(idx, 0),
                "origem": "IMPORT_MPS",
            })

    return registros


def _contar_rows_rodada(tabela, rodada_id):
    try:
        rows = _select_all(
            supabase.table(tabela)
            .select("id")
            .eq("rodada_id", rodada_id)
        )
        return len(rows)
    except Exception:
        return 0


def _amostra_rows_rodada(tabela, rodada_id, limit=5):
    try:
        return _select_all(
            supabase.table(tabela)
            .select("*")
            .eq("rodada_id", rodada_id)
            .limit(limit)
        )[:limit]
    except Exception:
        return []


def _diagnostico_importacao_mps(excel_bytes, xls, abas_config):
    diagnostico = {
        "abas_arquivo": list(getattr(xls, "sheet_names", []) or []),
        "abas_esperadas": [],
    }
    for label, config in abas_config.items():
        sheet_real = _encontrar_aba_mps(xls, config.get("nomes") or [label])
        item = {
            "label": label,
            "aba_encontrada": sheet_real,
            "recurso": config.get("recurso"),
            "etapa": config.get("etapa"),
        }
        if sheet_real:
            try:
                header_idx = _detectar_linha_header_mps(excel_bytes, sheet_real)
                df = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_real, header=header_idx, engine="openpyxl", nrows=3)
                item["header_excel_row"] = header_idx + 1
                item["colunas_lidas"] = [str(c) for c in df.columns[:30]]
            except Exception as exc:
                item["erro_preview"] = str(exc)
        diagnostico["abas_esperadas"].append(item)
    return diagnostico


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/rodadas")
def listar_rodadas():
    try:
        res = supabase.table("f_mrp_rodadas").select("*").order("criado_em", desc=True).execute()
        return res.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rodadas")
def criar_rodada(payload: RodadaCreate):
    try:
        res = supabase.table("f_mrp_rodadas").insert({"nome": payload.nome, "mes": payload.mes, "ano": payload.ano, "versao": payload.versao, "observacao": payload.observacao, "status": "rascunho"}).execute()
        data = res.data or []
        if not data:
            raise HTTPException(status_code=500, detail="Rodada criada, mas o Supabase não retornou dados.")
        return data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/rodadas/{rodada_id}")
def excluir_rodada(rodada_id: str):
    try:
        rodada_res = supabase.table("f_mrp_rodadas").select("*").eq("id", rodada_id).execute()
        if not (rodada_res.data or []):
            raise HTTPException(status_code=404, detail="Rodada não encontrada.")
        try:
            supabase.table("f_mrp_producao_real").delete().eq("rodada_id", rodada_id).execute()
        except Exception:
            pass
        aloc_res = supabase.table("f_mrp_alocacoes_dia").delete().eq("rodada_id", rodada_id).execute()
        etapas_res = supabase.table("f_mrp_etapas").delete().eq("rodada_id", rodada_id).execute()
        rodada_delete_res = supabase.table("f_mrp_rodadas").delete().eq("id", rodada_id).execute()
        return {"ok": True, "rodada_id": rodada_id, "alocacoes_excluidas": len(aloc_res.data or []), "etapas_excluidas": len(etapas_res.data or []), "rodada_excluida": rodada_delete_res.data[0] if rodada_delete_res.data else None}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rodadas/{rodada_id}/copiar")
def copiar_rodada(rodada_id: str, payload: RodadaCopiarCreate):
    try:
        res_origem = supabase.table("f_mrp_rodadas").select("*").eq("id", rodada_id).single().execute()
        rodada_origem = res_origem.data
        if not rodada_origem:
            raise HTTPException(status_code=404, detail="Rodada origem não encontrada.")
        nova_rodada = _copiar_rodada_base(rodada_origem, payload)
        nova_rodada_id = nova_rodada["id"]
        etapas_origem = _select_all(supabase.table("f_mrp_etapas").select("*").eq("rodada_id", rodada_id))
        alocacoes_origem = _select_all(supabase.table("f_mrp_alocacoes_dia").select("*").eq("rodada_id", rodada_id))
        etapas_novas = []
        for item in etapas_origem:
            novo = _remover_campos_controle(item)
            novo["rodada_id"] = nova_rodada_id
            novo["origem"] = "COPIA_RODADA"
            novo["observacao"] = f"{novo.get('observacao') or ''} | Copiado da rodada V{rodada_origem.get('versao')}."
            etapas_novas.append(novo)
        alocacoes_novas = []
        for item in alocacoes_origem:
            novo = _remover_campos_controle(item)
            novo["rodada_id"] = nova_rodada_id
            novo["origem"] = "COPIA_RODADA"
            alocacoes_novas.append(novo)
        total_etapas = _insert_em_lotes("f_mrp_etapas", etapas_novas)
        total_alocacoes = _insert_em_lotes("f_mrp_alocacoes_dia", alocacoes_novas)
        return {"ok": True, "rodada_origem_id": rodada_id, "nova_rodada": nova_rodada, "total_etapas": total_etapas, "total_alocacoes": total_alocacoes}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rodadas/{rodada_id}/importar-mps")
async def importar_mps(
    rodada_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
):
    """
    Importa o MPS para a rodada.

    Hotfix v8:
    - detecta header dinamicamente;
    - aceita pequenas variações de nome das abas;
    - valida se a rodada ficou gravada no Supabase;
    - se gravar 0 etapas, retorna erro em vez de "MPS importado" falso.
    """
    try:
        if not file.filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="Envie um arquivo Excel .xlsx ou .xlsm.")

        excel_bytes = await file.read()
        parametros = _carregar_parametros_globais()
        parametros_produto = _carregar_parametros_produto()

        abas = {
            "PROGRAMAÇÃO MACRO - LINHA 1": {
                "nomes": ["PROGRAMAÇÃO MACRO - LINHA 1", "PROGRAMACAO MACRO - LINHA 1", "LINHA 1", "L1"],
                "recurso": "L1",
                "etapa": "ENVASE",
            },
            "PROGRAMAÇÃO MACRO - LINHA 2": {
                "nomes": ["PROGRAMAÇÃO MACRO - LINHA 2", "PROGRAMACAO MACRO - LINHA 2", "LINHA 2", "L2"],
                "recurso": "L2",
                "etapa": "ENVASE",
            },
            "PROGRAMAÇÃO MACRO - FABRIMA": {
                "nomes": ["PROGRAMAÇÃO MACRO - FABRIMA", "PROGRAMACAO MACRO - FABRIMA", "FABRIMA"],
                "recurso": "FABRIMA",
                "etapa": "FABRIMA",
            },
        }

        xls = pd.ExcelFile(BytesIO(excel_bytes), engine="openpyxl")
        registros = []
        alocacoes = []
        abas_lidas = []
        diagnostico = _diagnostico_importacao_mps(excel_bytes, xls, abas)

        for aba_label, config in abas.items():
            sheet_name = _encontrar_aba_mps(xls, config.get("nomes") or [aba_label])
            if not sheet_name:
                abas_lidas.append({
                    "aba": aba_label,
                    "aba_encontrada": None,
                    "qtd_registros": 0,
                    "qtd_alocacoes": 0,
                    "status": "aba_nao_encontrada",
                })
                continue

            dados_aba = _ler_aba_mps(excel_bytes, sheet_name, config["recurso"], config["etapa"])
            debug_aba = getattr(_ler_aba_mps, "_ultimo_debug", {}) or {}

            for item in dados_aba:
                item["rodada_id"] = rodada_id

            dados_alocacoes = _ler_alocacoes_dia(excel_bytes, sheet_name, config["recurso"], rodada_id)
            dados_aba = _recalcular_registros_mrp(dados_aba, dados_alocacoes, parametros, parametros_produto)

            registros.extend(dados_aba)
            alocacoes.extend(dados_alocacoes)
            abas_lidas.append({
                "aba": aba_label,
                "aba_encontrada": sheet_name,
                "qtd_registros": len(dados_aba),
                "qtd_alocacoes": len(dados_alocacoes),
                "debug": debug_aba,
                "status": "ok",
            })

        if not registros:
            raise HTTPException(
                status_code=400,
                detail={
                    "erro": "Nenhum registro encontrado nas abas esperadas do MPS.",
                    "arquivo": file.filename,
                    "abas_lidas": abas_lidas,
                    "diagnostico": diagnostico,
                    "orientacao": "Verifique se o arquivo tem as abas de L1/L2/Fabrima e se o cabeçalho contém LOTE, CÓDIGO, PRODUTO, DATA INÍCIO e DATA FIM.",
                },
            )

        # Limpa a rodada e grava de novo.
        supabase.table("f_mrp_etapas").delete().eq("rodada_id", rodada_id).execute()
        supabase.table("f_mrp_alocacoes_dia").delete().eq("rodada_id", rodada_id).execute()

        total_inserido_resposta = _insert_em_lotes("f_mrp_etapas", registros)
        total_alocacoes_resposta = _insert_em_lotes("f_mrp_alocacoes_dia", alocacoes)

        # Não confiar apenas no retorno do insert: dependendo da configuração do
        # Supabase, res.data pode vir vazio. A fonte de verdade é consultar a rodada.
        total_etapas_gravadas = _contar_rows_rodada("f_mrp_etapas", rodada_id)
        total_alocacoes_gravadas = _contar_rows_rodada("f_mrp_alocacoes_dia", rodada_id)

        if total_etapas_gravadas <= 0:
            raise HTTPException(
                status_code=500,
                detail={
                    "erro": "O arquivo foi lido, mas nenhuma etapa ficou gravada na rodada.",
                    "arquivo": file.filename,
                    "rodada_id": rodada_id,
                    "total_registros_lidos": len(registros),
                    "total_inserido_resposta_supabase": total_inserido_resposta,
                    "total_etapas_gravadas": total_etapas_gravadas,
                    "total_alocacoes_gravadas": total_alocacoes_gravadas,
                    "abas_lidas": abas_lidas,
                    "diagnostico": diagnostico,
                    "amostra_primeiro_registro": registros[:2],
                },
            )

        # Comentários/calendário rodam fora do request para não estourar timeout no Fly.
        background_tasks.add_task(_processar_calendario_gantt_background, rodada_id, excel_bytes)

        return {
            "ok": True,
            "rodada_id": rodada_id,
            "arquivo": file.filename,
            "abas_lidas": abas_lidas,
            "total_registros": len(registros),
            "total_inserido_resposta_supabase": total_inserido_resposta,
            "total_etapas_gravadas": total_etapas_gravadas,
            "total_alocacoes_lidas": len(alocacoes),
            "total_alocacoes_inseridas_resposta_supabase": total_alocacoes_resposta,
            "total_alocacoes_gravadas": total_alocacoes_gravadas,
            "amostra_etapas_gravadas": _amostra_rows_rodada("f_mrp_etapas", rodada_id, limit=3),
            "comentarios_gantt_status": "processando_background",
            "validar_comentarios_em": f"/mrp/rodadas/{rodada_id}/calendario-comentarios-resumo",
            "diagnostico": diagnostico,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rodadas/{rodada_id}/calendario-comentarios-resumo")
def resumo_calendario_comentarios(rodada_id: str):
    """
    Diagnóstico rápido pós-upload para validar se f_mrp_calendario_dia recebeu
    calendário e comentários do Gantt.
    """
    try:
        rows = _select_all(
            supabase.table("f_mrp_calendario_dia")
            .select("*")
            .eq("rodada_id", rodada_id)
        )

        total = len(rows)
        com_comentario = [
            r for r in rows
            if str(r.get("comentario_calendario") or "").strip()
        ]

        exemplos = []
        for r in com_comentario[:20]:
            exemplos.append({
                "recurso": r.get("recurso"),
                "data": r.get("data"),
                "horas_disponiveis_dia": r.get("horas_disponiveis_dia"),
                "comentario_calendario": r.get("comentario_calendario"),
            })

        return {
            "ok": True,
            "rodada_id": rodada_id,
            "total_calendario_dia": total,
            "total_dias_com_comentario_gantt": len(com_comentario),
            "exemplos": exemplos,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rodadas/{rodada_id}/importar-producao-real")
async def importar_producao_real(rodada_id: str, file: UploadFile = File(...)):
    try:
        if not file.filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="Envie um arquivo Excel .xlsx ou .xlsm.")
        excel_bytes = await file.read()

        apontamentos = _ler_apontamentos_cogtive(excel_bytes, rodada_id)
        registros_reais = apontamentos.get("producoes", [])
        registros_producao_raw = apontamentos.get("producoes_raw", registros_reais)
        registros_paradas = apontamentos.get("paradas", [])
        # Banco recebe apontamentos detalhados (produção linha a linha + paradas).
        # A aplicação do realizado usa registros_reais agrupados por lote/recurso.
        registros_para_banco = registros_producao_raw + registros_paradas

        if not registros_reais:
            raise HTTPException(status_code=400, detail="Nenhum apontamento de produção encontrado no relatório.")

        total_real_inserido = 0
        try:
            supabase.table("f_mrp_producao_real").delete().eq("rodada_id", rodada_id).execute()
            total_real_inserido = _insert_em_lotes("f_mrp_producao_real", registros_para_banco)
        except Exception:
            pass

        resultado = _atualizar_lotes_com_real(rodada_id, registros_reais)
        mudancas = _anexar_paradas_mudancas(resultado["atualizadas"], registros_paradas, rodada_id)
        resumo_por_linha = {}
        for item in mudancas:
            recurso = item.get("recurso") or "OUTROS"
            resumo_por_linha.setdefault(recurso, {"total": 0, "atrasou": 0, "antecipou": 0, "sem_mudanca_data": 0, "sem_comparativo": 0})
            resumo_por_linha[recurso]["total"] += 1
            tipo = item.get("tipo_impacto") or "sem_comparativo"
            resumo_por_linha[recurso][tipo] = resumo_por_linha[recurso].get(tipo, 0) + 1
        return {
            "ok": True, "rodada_id": rodada_id, "arquivo": file.filename,
            "total_apontamentos_lidos": len(registros_para_banco),
            "total_apontamentos_producao": len(registros_reais),
            "total_apontamentos_paradas": len(registros_paradas),
            "total_real_inserido": total_real_inserido,
            "total_lotes_atualizados": len(mudancas),
            "total_lotes_nao_encontrados": len(resultado["nao_encontradas"]),
            "total_lotes_ignorados_fora_mes": len(resultado.get("ignoradas_fora_mes", [])),
            "total_lotes_ignorados_ja_realizados": len(resultado.get("ignoradas_ja_realizadas", [])),
            "resumo_por_linha": resumo_por_linha,
            "mudancas_realizado": mudancas,
            "lotes_atualizados": mudancas,
            "lotes_nao_encontrados": resultado["nao_encontradas"],
            "lotes_ignorados_fora_mes": resultado.get("ignoradas_fora_mes", []),
            "lotes_ignorados_ja_realizados": resultado.get("ignoradas_ja_realizadas", []),
            "cascatas_aplicadas": resultado.get("cascatas_aplicadas", []),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rodadas/{rodada_id}/mudancas-realizado")
def listar_mudancas_realizado(rodada_id: str):
    try:
        etapas = _select_all(supabase.table("f_mrp_etapas").select("*").eq("rodada_id", rodada_id).order("recurso").order("data_inicio").order("sequencia"))
        mudancas_base = [m for e in etapas if (m := _montar_mudanca_realizado_da_etapa(e))]
        paradas_cogtive = _buscar_paradas_cogtive_rodada(rodada_id)
        mudancas = _anexar_paradas_mudancas(mudancas_base, paradas_cogtive, rodada_id)
        resumo_por_linha = {}
        for item in mudancas:
            recurso = item.get("recurso") or "OUTROS"
            resumo_por_linha.setdefault(recurso, {"total": 0, "atrasou": 0, "antecipou": 0, "sem_mudanca_data": 0, "sem_comparativo": 0})
            resumo_por_linha[recurso]["total"] += 1
            tipo = item.get("tipo_impacto") or "sem_comparativo"
            resumo_por_linha[recurso][tipo] = resumo_por_linha[recurso].get(tipo, 0) + 1
        return {"ok": True, "rodada_id": rodada_id, "total": len(mudancas), "resumo_por_linha": resumo_por_linha, "mudancas_realizado": mudancas, "lotes_atualizados": mudancas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rodadas/{rodada_id}/comparativo-liberacao")
def comparar_liberacao_rodadas(rodada_id: str):
    try:
        rodada_res = supabase.table("f_mrp_rodadas").select("*").eq("id", rodada_id).single().execute()
        rodada_atual = rodada_res.data
        if not rodada_atual:
            raise HTTPException(status_code=404, detail="Rodada não encontrada.")
        rodada_anterior = _buscar_rodada_anterior(rodada_atual)
        etapas_atual = _select_all(supabase.table("f_mrp_etapas").select("*").eq("rodada_id", rodada_id))
        etapas_anterior = []
        if rodada_anterior and rodada_anterior.get("id"):
            etapas_anterior = _select_all(supabase.table("f_mrp_etapas").select("*").eq("rodada_id", rodada_anterior["id"]))
        comparativo = _montar_comparativo_liberacao(rodada_atual, rodada_anterior, etapas_atual, etapas_anterior)
        return {"ok": True, "rodada_id": rodada_id, **comparativo}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rodadas/{rodada_id}/etapas")
def listar_etapas(rodada_id: str):
    try:
        query = supabase.table("f_mrp_etapas").select("*").eq("rodada_id", rodada_id).order("recurso").order("data_inicio").order("sequencia")
        return _select_all(query)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rodadas/{rodada_id}/alocacoes")
def listar_alocacoes(rodada_id: str):
    try:
        query = supabase.table("f_mrp_alocacoes_dia").select("*").eq("rodada_id", rodada_id).order("recurso").order("data")
        return _select_all(query)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/parametros")
def listar_parametros_mrp():
    try:
        globais = _select_all(supabase.table("d_mrp_parametros").select("*").order("chave"))
        produtos = _select_all(supabase.table("d_mrp_parametros_produto").select("*").order("grupo_produto"))
        return {"globais": globais, "produtos": produtos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/etapas")
def criar_etapa(payload: EtapaCreate):
    try:
        res = supabase.table("f_mrp_etapas").insert({
            "rodada_id": payload.rodada_id, "lote": payload.lote, "op": payload.op,
            "codigo_produto": payload.codigo_produto, "descricao_produto": payload.descricao_produto,
            "etapa": payload.etapa, "recurso": payload.recurso, "linha_origem": payload.linha_origem,
            "data_inicio": _date_to_str(payload.data_inicio), "data_fim": _date_to_str(payload.data_fim),
            "data_pa": _date_to_str(payload.data_pa), "qtd_planejada": payload.qtd_planejada,
            "duracao_horas": payload.duracao_horas, "sequencia": payload.sequencia,
            "status": payload.status, "origem": payload.origem, "observacao": payload.observacao,
            "embalado": payload.embalado, "un_hora": payload.un_hora,
            "mes_producao": payload.mes_producao, "ano_producao": payload.ano_producao,
            "mes_liberacao": payload.mes_liberacao, "ano_liberacao": payload.ano_liberacao,
            "mes_lib_manual": payload.mes_lib_manual,
        }).execute()
        data = res.data or []
        if not data:
            raise HTTPException(status_code=500, detail="Etapa criada, mas o Supabase não retornou dados.")
        return data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/etapas/{etapa_id}")
def atualizar_etapa(etapa_id: str, payload: EtapaUpdate):
    try:
        dados = payload.model_dump(exclude_unset=True)
        for campo in ["data_inicio", "data_fim", "data_pa"]:
            if campo in dados:
                dados[campo] = _date_to_str(dados[campo])
        if not dados:
            raise HTTPException(status_code=400, detail="Nenhum campo enviado para atualização.")
        res = supabase.table("f_mrp_etapas").update(dados).eq("id", etapa_id).execute()
        data = res.data or []
        if not data:
            raise HTTPException(status_code=404, detail="Etapa não encontrada.")
        return data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/etapas/{etapa_id}")
def excluir_etapa(etapa_id: str):
    try:
        res = supabase.table("f_mrp_etapas").delete().eq("id", etapa_id).execute()
        return {"ok": True, "deleted": res.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sd3-realizado")
def sd3_realizado(ano: int):
    try:
        rows = _select_all(
            supabase.table("f_sd3_entradas")
            .select("mes, ano, quantidade, lote, descr_prod, armazem, grupo")
            .eq("ano", ano)
        )
        resultado: dict = {}
        for r in rows:
            armazem = str(r.get("armazem") or "").strip()
            grupo = str(r.get("grupo") or "").strip()
            descr = str(r.get("descr_prod") or "").upper()
            lote = str(r.get("lote") or "")
            qtd = _to_float(r.get("quantidade"), 0)
            mes = _to_int(r.get("mes"))
            ano_r = _to_int(r.get("ano"))
            if not mes or not ano_r:
                continue
            if armazem not in ("04", "07"):
                continue
            if not ("0101" <= grupo <= "0116"):
                continue
            if "AVULSO" in descr:
                continue
            if " AG " in descr or descr.startswith("AG ") or descr.endswith(" AG"):
                continue
            linha = None
            m = re.search(r"[A-Za-z]([12])", lote)
            if m:
                linha = f"L{m.group(1)}"
            chave = (mes, ano_r)
            if chave not in resultado:
                resultado[chave] = {"mes": mes, "ano": ano_r, "caixas": 0.0, "caixas_l1": 0.0, "caixas_l2": 0.0}
            resultado[chave]["caixas"] += qtd
            if linha == "L1":
                resultado[chave]["caixas_l1"] += qtd
            elif linha == "L2":
                resultado[chave]["caixas_l2"] += qtd
        return sorted(resultado.values(), key=lambda x: (x["ano"], x["mes"]))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))